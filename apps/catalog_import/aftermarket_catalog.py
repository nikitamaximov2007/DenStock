"""Import an independent USD aftermarket supplier catalog.

The dealer format is intentionally not an analog-relation file.  It creates
normal catalog cards, records source facts in ``AftermarketCatalogPart`` and
never guesses an original BRP/Polaris part or changes warehouse state.
"""

from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.db import transaction
from django.db.models import Max
from django.utils import timezone

from apps.catalog.models import Category, Manufacturer, PartNumber, PartType, Unit, normalize_number
from apps.inventory.presentation import EXACT_NUMBER_KINDS

from .models import AftermarketCatalogPart

MAX_PROBLEMS = 200
MAX_TEXT = 10_000
SOURCE = AftermarketCatalogPart.SOURCE_DEALER_2023
AFTERMARKET_CATEGORY_NAME = "Aftermarket"
DEFAULT_UNIT_NAME = "Штука"

HEADERS = {
    "manufacturer": {"manufacturer"},
    "supplier_sku": {"item sku", "itemsku"},
    "manufacturer_number": {"manufacturer number", "manufacturernumber"},
    "description": {"description"},
    "msrp": {"msrp"},
    "dealer_cost": {"dlr cost", "dlr. cost", "dealer cost"},
}
REQUIRED = ("manufacturer", "manufacturer_number", "description")


class AftermarketCatalogError(RuntimeError):
    """A supplier workbook cannot be safely interpreted."""


@dataclass(frozen=True)
class Problem:
    row: int
    reason: str
    detail: str = ""


@dataclass(frozen=True)
class IncomingRow:
    number: int
    manufacturer: str
    manufacturer_key: str
    manufacturer_number: str
    normalized_number: str
    supplier_sku: str
    description: str
    msrp_usd: Decimal | None
    dealer_cost_usd: Decimal | None

    @property
    def identity(self) -> tuple[str, str]:
        return self.manufacturer_key, self.normalized_number


@dataclass
class Plan:
    sheet: str = ""
    rows_scanned: int = 0
    valid: int = 0
    blank_rows: int = 0
    new_parts: int = 0
    updates: int = 0
    unchanged: int = 0
    ambiguous: int = 0
    invalid: int = 0
    duplicate_source_identities: int = 0
    problems: list[Problem] = field(default_factory=list)

    def problem(self, row: int, reason: str, detail: str = "") -> None:
        if len(self.problems) < MAX_PROBLEMS:
            self.problems.append(Problem(row, reason, detail))

    def as_summary(self) -> dict:
        return {
            "format": "AFTERMARKET_SUPPLIER_CATALOG",
            "format_label": "Каталог аналогов / aftermarket",
            "sheet": self.sheet,
            "rows_scanned": self.rows_scanned,
            "valid": self.valid,
            "blank_rows": self.blank_rows,
            "new_parts": self.new_parts,
            "updates": self.updates,
            "unchanged": self.unchanged,
            "ambiguous": self.ambiguous,
            "invalid": self.invalid,
            "duplicate_source_identities": self.duplicate_source_identities,
            "currency": "USD",
            "msrp_label": "MSRP, USD",
            "dealer_cost_label": "Dlr Cost, USD",
            "stock_changes": False,
            "problems": [item.__dict__ for item in self.problems],
            "problems_total": self.invalid + self.ambiguous + self.duplicate_source_identities,
        }


def _header_key(value: object) -> str:
    return " ".join(str(value or "").replace("\xa0", " ").strip().lower().split())


def _text(value: object) -> str:
    text = "" if value is None else str(value).strip()
    if len(text) > MAX_TEXT:
        raise AftermarketCatalogError("В ячейке слишком длинное значение.")
    return text


def _identifier(cell) -> str:
    value = cell.value
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    if isinstance(value, int):
        # A numeric Excel cell cannot retain leading zeroes in its value.  A
        # simple all-zero number format is the one safe recoverable exception.
        fmt = str(cell.number_format or "")
        if fmt and set(fmt) == {"0"}:
            return f"{value:0{len(fmt)}d}"
    return _text(value)


def _price(value: object, label: str) -> Decimal | None:
    text = _text(value).replace(" ", "").replace(",", ".")
    if not text:
        return None
    try:
        result = Decimal(text)
    except (InvalidOperation, ValueError) as exc:
        raise AftermarketCatalogError(f"{label}: «{text}» не является ценой.") from exc
    if not result.is_finite() or result < 0:
        raise AftermarketCatalogError(f"{label}: цена должна быть конечной и неотрицательной.")
    try:
        return result.quantize(Decimal("0.01"))
    except InvalidOperation as exc:
        raise AftermarketCatalogError(f"{label}: цена слишком велика.") from exc


PRICE_SHEET = "priceupdate"


def _price_sheet_name(names) -> str:
    """Имя листа с прайсом среди листов книги.

    Поставщики называют вкладку своим именем: «diorlight priceupdate»,
    «priceupdate 2026». Ключевое слово остаётся, поэтому лист ищется по
    вхождению, а не по точному совпадению.

    Угадывать нельзя: если подходящих листов несколько, книга отклоняется с
    перечислением - выбрать нужный должен человек, а не импортёр.
    """
    matches = [name for name in names if PRICE_SHEET in _header_key(name).replace(" ", "")]
    if len(matches) == 1:
        return matches[0]
    if not matches:
        raise AftermarketCatalogError(
            "В каталоге aftermarket нужен лист priceupdate. Найдены листы: "
            + (", ".join(names) or "ни одного")
        )
    raise AftermarketCatalogError(
        "В книге несколько листов с прайсом, нужен ровно один: " + ", ".join(matches)
    )


def _mapping(cells) -> dict[str, int]:
    found: dict[str, int] = {}
    for index, cell in enumerate(cells):
        key = _header_key(cell.value)
        for name, aliases in HEADERS.items():
            if key in aliases and name not in found:
                found[name] = index
                break
    missing = [name for name in REQUIRED if name not in found]
    if missing:
        values = ", ".join(_text(cell.value) for cell in cells if _text(cell.value)) or "пусто"
        raise AftermarketCatalogError(
            "Не распознан каталог aftermarket. Нужны колонки Manufacturer, "
            "Manufacturer Number и Description. Найдены: " + values
        )
    return found


def _read(path: Path) -> tuple[str, list[IncomingRow], Plan]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise AftermarketCatalogError("Не установлен openpyxl.") from exc
    if path.suffix.lower() != ".xlsx" or not path.is_file():
        raise AftermarketCatalogError("Ожидается доступный файл .xlsx.")
    book = None
    try:
        book = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise AftermarketCatalogError(f"Файл не читается как Excel: {exc}") from exc
    try:
        sheet = book[_price_sheet_name(book.sheetnames)]
        rows = sheet.iter_rows()
        try:
            header_cells = next(rows)
        except StopIteration as exc:
            raise AftermarketCatalogError("Файл пустой.") from exc
        mapping = _mapping(header_cells)
        plan = Plan(sheet=sheet.title)
        records: list[IncomingRow] = []
        seen: set[tuple[str, str]] = set()
        for row_number, cells in enumerate(rows, start=2):
            values = [_text(cell.value) for cell in cells]
            if not any(values):
                plan.blank_rows += 1
                continue
            plan.rows_scanned += 1

            mapped_cells = {
                name: cells[index] for name, index in mapping.items() if index < len(cells)
            }
            manufacturer = " ".join(
                _text(
                    mapped_cells["manufacturer"].value if "manufacturer" in mapped_cells else ""
                ).split()
            )
            number = (
                _identifier(mapped_cells["manufacturer_number"])
                if "manufacturer_number" in mapped_cells
                else ""
            )
            description = " ".join(
                _text(
                    mapped_cells["description"].value if "description" in mapped_cells else ""
                ).split()
            )
            if not manufacturer or not number or not description:
                plan.invalid += 1
                plan.problem(row_number, "Не заполнены обязательные поля")
                continue
            normalized = normalize_number(number)
            if not normalized:
                plan.invalid += 1
                plan.problem(row_number, "Некорректный номер производителя")
                continue
            try:
                record = IncomingRow(
                    number=row_number,
                    manufacturer=manufacturer[:150],
                    manufacturer_key=manufacturer.casefold(),
                    manufacturer_number=number[:100],
                    normalized_number=normalized,
                    supplier_sku=(
                        _identifier(mapped_cells["supplier_sku"])
                        if "supplier_sku" in mapped_cells
                        else ""
                    )[:100],
                    description=description[:200],
                    msrp_usd=(
                        _price(mapped_cells["msrp"].value, "MSRP")
                        if "msrp" in mapped_cells
                        else None
                    ),
                    dealer_cost_usd=(
                        _price(mapped_cells["dealer_cost"].value, "Dlr Cost")
                        if "dealer_cost" in mapped_cells
                        else None
                    ),
                )
            except AftermarketCatalogError as exc:
                plan.invalid += 1
                plan.problem(row_number, "Некорректная цена", str(exc))
                continue
            if record.identity in seen:
                plan.duplicate_source_identities += 1
                plan.problem(row_number, "Повторяется идентичность поставщика", number)
                continue
            seen.add(record.identity)
            records.append(record)
        return sheet.title, records, plan
    finally:
        if book is not None:
            book.close()


def _manufacturer_map() -> dict[str, Manufacturer]:
    return {item.name.casefold(): item for item in Manufacturer.objects.all()}


def _unit() -> Unit:
    unit = Unit.objects.filter(name__iexact=DEFAULT_UNIT_NAME, is_active=True).first()
    return unit or Unit.objects.filter(is_active=True).first() or _missing_unit()


def _missing_unit():
    raise AftermarketCatalogError("В справочниках нет активной единицы измерения.")


def _source_index(records: list[IncomingRow], manufacturers: dict[str, Manufacturer]):
    ids = [manufacturer.pk for manufacturer in manufacturers.values()]
    rows = AftermarketCatalogPart.objects.filter(
        source=SOURCE, manufacturer_id__in=ids
    ).select_related("part")
    return {(item.manufacturer_id, item.normalized_manufacturer_number): item for item in rows}


def _chunks(values, size: int = 500):
    for start in range(0, len(values), size):
        yield values[start : start + size]


def _manual_article_index(records: list[IncomingRow]):
    numbers = {record.normalized_number for record in records}
    result: dict[str, list[PartType]] = defaultdict(list)
    # SQLite has a much lower bind-variable limit than a dealer price file has
    # rows. Batching is bounded and avoids a query per source record.
    for chunk in _chunks(sorted(numbers)):
        queryset = PartNumber.objects.filter(
            normalized_value__in=chunk,
            kind__in=EXACT_NUMBER_KINDS,
        ).select_related("part__manufacturer")
        for number in queryset:
            if all(part.pk != number.part_id for part in result[number.normalized_value]):
                result[number.normalized_value].append(number.part)
    return result


def _same_values(entry: AftermarketCatalogPart, record: IncomingRow) -> bool:
    return (
        entry.manufacturer_number == record.manufacturer_number
        and entry.supplier_sku == record.supplier_sku
        and entry.source_description == record.description
        and entry.msrp_usd == _keep_positive(entry.msrp_usd, record.msrp_usd)
        and entry.dealer_cost_usd == _keep_positive(entry.dealer_cost_usd, record.dealer_cost_usd)
        and entry.part.name == record.description
    )


def _keep_positive(previous: Decimal | None, incoming: Decimal | None) -> Decimal | None:
    """A blank or zero source price never destroys a known positive price."""
    return (
        previous if incoming is None or (incoming == 0 and previous and previous > 0) else incoming
    )


def _classify(records: list[IncomingRow], plan: Plan):
    manufacturers = _manufacturer_map()
    sources = _source_index(records, manufacturers)
    manual = _manual_article_index(records)
    usable: list[IncomingRow] = []
    for record in records:
        manufacturer = manufacturers.get(record.manufacturer_key)
        entry = sources.get((manufacturer.pk, record.normalized_number)) if manufacturer else None
        if entry:
            plan.valid += 1
            if _same_values(entry, record):
                plan.unchanged += 1
            else:
                plan.updates += 1
            usable.append(record)
            continue
        collisions = [
            part
            for part in manual.get(record.normalized_number, [])
            if part.manufacturer and part.manufacturer.name.casefold() == record.manufacturer_key
        ]
        if collisions:
            plan.ambiguous += 1
            plan.problem(
                record.number, "Неоднозначная существующая карточка", record.manufacturer_number
            )
            continue
        plan.valid += 1
        plan.new_parts += 1
        usable.append(record)
    return usable


def build_plan(path) -> Plan:
    _, records, plan = _read(Path(path))
    _classify(records, plan)
    return plan


def catalog_fingerprint() -> str:
    import hashlib

    source_state = AftermarketCatalogPart.objects.aggregate(
        total=Max("pk"), touched=Max("updated_at")
    )
    # A preview can also become unsafe when a new manual card/number appears
    # between checking the file and applying it: that would turn a proposed
    # creation into an ambiguous identity.  Counts and last PKs are cheap
    # bounded guards for those catalog changes.
    payload = "|".join(
        str(value)
        for value in (
            AftermarketCatalogPart.objects.count(),
            source_state["total"],
            source_state["touched"],
            PartType.objects.count(),
            PartType.objects.aggregate(total=Max("pk"))["total"],
            PartNumber.objects.count(),
            PartNumber.objects.aggregate(total=Max("pk"))["total"],
            Manufacturer.objects.count(),
            Manufacturer.objects.aggregate(total=Max("pk"))["total"],
        )
    )
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


@transaction.atomic
def apply_file(path) -> dict:
    _, records, plan = _read(Path(path))
    usable = _classify(records, plan)
    category, _ = Category.objects.get_or_create(name=AFTERMARKET_CATEGORY_NAME, parent=None)
    unit = _unit()
    manufacturers = _manufacturer_map()
    missing = {
        record.manufacturer_key: record.manufacturer
        for record in usable
        if record.manufacturer_key not in manufacturers
    }
    if missing:
        Manufacturer.objects.bulk_create([Manufacturer(name=name) for name in missing.values()])
        manufacturers = _manufacturer_map()
    sources = _source_index(usable, manufacturers)
    created_parts = 0
    updated_parts = 0
    updated_entries: list[AftermarketCatalogPart] = []
    updated_cards: list[PartType] = []
    new_rows: list[tuple[IncomingRow, Manufacturer]] = []
    existing_rows: list[tuple[IncomingRow, AftermarketCatalogPart]] = []
    for record in usable:
        manufacturer = manufacturers[record.manufacturer_key]
        key = (manufacturer.pk, record.normalized_number)
        entry = sources.get(key)
        if entry is None:
            new_rows.append((record, manufacturer))
            continue
        existing_rows.append((record, entry))

    for chunk in _chunks(new_rows):
        cards = [
            PartType(
                name=record.description,
                description=record.description,
                category=category,
                manufacturer=manufacturer,
                unit=unit,
                tracking_mode=PartType.TrackingMode.BULK,
            )
            for record, manufacturer in chunk
        ]
        PartType.objects.bulk_create(cards)
        articles = []
        skus = []
        entries = []
        for (record, manufacturer), part in zip(chunk, cards, strict=True):
            articles.append(
                PartNumber(
                    part=part,
                    value=record.manufacturer_number,
                    normalized_value=record.normalized_number,
                    kind=PartNumber.Kind.ARTICLE,
                    is_primary=True,
                )
            )
            if record.supplier_sku:
                skus.append(
                    PartNumber(
                        part=part,
                        value=record.supplier_sku,
                        normalized_value=normalize_number(record.supplier_sku),
                        kind=PartNumber.Kind.INTERNAL_REF,
                        note="SKU aftermarket-поставщика",
                    )
                )
            entries.append(
                AftermarketCatalogPart(
                    source=SOURCE,
                    part=part,
                    manufacturer=manufacturer,
                    manufacturer_number=record.manufacturer_number,
                    normalized_manufacturer_number=record.normalized_number,
                    supplier_sku=record.supplier_sku,
                    source_description=record.description,
                    msrp_usd=record.msrp_usd,
                    dealer_cost_usd=record.dealer_cost_usd,
                )
            )
        PartNumber.objects.bulk_create(articles + skus)
        AftermarketCatalogPart.objects.bulk_create(entries)
        created_parts += len(chunk)

    sku_numbers = {
        item.part_id: item
        for item in PartNumber.objects.filter(
            part_id__in=[entry.part_id for _, entry in existing_rows],
            kind=PartNumber.Kind.INTERNAL_REF,
            note="SKU aftermarket-поставщика",
        )
    }
    sku_updates: list[PartNumber] = []
    sku_creates: list[PartNumber] = []
    for record, entry in existing_rows:
        changed = False
        for field_name, value in (
            ("manufacturer_number", record.manufacturer_number),
            ("supplier_sku", record.supplier_sku),
            ("source_description", record.description),
            ("msrp_usd", _keep_positive(entry.msrp_usd, record.msrp_usd)),
            ("dealer_cost_usd", _keep_positive(entry.dealer_cost_usd, record.dealer_cost_usd)),
        ):
            if getattr(entry, field_name) != value:
                setattr(entry, field_name, value)
                changed = True
        if entry.part.name != record.description or entry.part.description != record.description:
            entry.part.name = record.description
            entry.part.description = record.description
            updated_cards.append(entry.part)
            changed = True
        if changed:
            entry.updated_at = timezone.now()
            updated_entries.append(entry)
            updated_parts += 1
        if record.supplier_sku:
            sku = sku_numbers.get(entry.part_id)
            if sku is None:
                sku_creates.append(
                    PartNumber(
                        part=entry.part,
                        value=record.supplier_sku,
                        normalized_value=normalize_number(record.supplier_sku),
                        kind=PartNumber.Kind.INTERNAL_REF,
                        note="SKU aftermarket-поставщика",
                    )
                )
            elif sku.value != record.supplier_sku:
                sku.value = record.supplier_sku
                sku.normalized_value = normalize_number(record.supplier_sku)
                sku_updates.append(sku)
    if updated_cards:
        PartType.objects.bulk_update(updated_cards, ["name", "description"])
    if updated_entries:
        AftermarketCatalogPart.objects.bulk_update(
            updated_entries,
            [
                "manufacturer_number",
                "supplier_sku",
                "source_description",
                "msrp_usd",
                "dealer_cost_usd",
                "updated_at",
            ],
        )
    if sku_creates:
        PartNumber.objects.bulk_create(sku_creates)
    if sku_updates:
        PartNumber.objects.bulk_update(sku_updates, ["value", "normalized_value"])
    result = plan.as_summary()
    result.update(
        {"created_parts": created_parts, "updated_parts": updated_parts, "stock_changes": False}
    )
    return result
