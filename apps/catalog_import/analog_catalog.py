"""Каталог аналогов из Excel: разбор, план и применение.

Формат намеренно универсальный. Файлов конкретных производителей сейчас нет, и
угадывать их разметку было бы гаданием. Здесь описан канонический набор колонок
и несколько понятных синонимов заголовков; когда придёт настоящий файл, к нему
добавится синоним, а не второй разборщик.

Главное правило разбора: артикул НЕ определяет деталь. У аналога он часто
совпадает с исходной, поэтому строку вида `Part.objects.get(article=...)`
писать нельзя нигде. Если по артикулу нашлось несколько деталей, строка
показывается человеку, а не решается за него выбором первой.

Остатков этот импорт не создаёт вовсе: он трогает только справочник.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.db import transaction

from apps.catalog.models import (
    PartAnalog,
    PartNumber,
    PartType,
    normalize_number,
)
from apps.catalog.services import (
    AnalogLinkError,
    ManualPartError,
    create_manual_part,
    link_analog,
)
from apps.inventory.presentation import EXACT_NUMBER_KINDS

MAX_PROBLEMS = 200
MAX_ROWS = 100_000

# Канонические колонки и их синонимы. Сравнение идёт по строчным буквам без
# лишних пробелов, поэтому регистр и отступы в заголовке значения не имеют.
COLUMNS = {
    "original_article": (
        "original_article", "исходный артикул", "оригинальный артикул",
        "артикул оригинала", "артикул исходной детали",
    ),
    "analog_article": (
        "analog_article", "артикул аналога", "артикул заменителя",
    ),
    "analog_name": (
        "analog_name", "название аналога", "наименование аналога", "название",
    ),
    "analog_price": (
        "analog_price", "цена", "цена аналога", "рекомендуемая цена",
    ),
    "analog_manufacturer": (
        "analog_manufacturer", "производитель аналога", "производитель", "бренд",
    ),
    "analog_barcode": (
        "analog_barcode", "штрихкод аналога", "штрихкод",
    ),
    "original_manufacturer": (
        "original_manufacturer", "производитель оригинала",
        "производитель исходной детали",
    ),
}
REQUIRED = ("original_article", "analog_article", "analog_name")


class AnalogCatalogError(RuntimeError):
    """Файл не разобран: понятная человеку причина."""


@dataclass
class Row:
    number: int
    original_article: str = ""
    analog_article: str = ""
    analog_name: str = ""
    analog_price: str = ""
    analog_manufacturer: str = ""
    analog_barcode: str = ""
    original_manufacturer: str = ""


@dataclass
class Problem:
    row: int
    reason: str
    detail: str = ""


@dataclass
class Plan:
    """Что произойдёт, если применить файл. Ничего не пишет."""

    rows_total: int = 0
    will_create_parts: int = 0
    will_reuse_parts: int = 0
    will_create_links: int = 0
    already_linked: int = 0
    ambiguous: int = 0
    invalid: int = 0
    problems: list[Problem] = field(default_factory=list)

    def as_summary(self) -> dict:
        return {
            "rows_total": self.rows_total,
            "will_create_parts": self.will_create_parts,
            "will_reuse_parts": self.will_reuse_parts,
            "will_create_links": self.will_create_links,
            "already_linked": self.already_linked,
            "ambiguous": self.ambiguous,
            "invalid": self.invalid,
            "needs_attention": self.ambiguous + self.invalid,
            "problems": [
                {"row": item.row, "reason": item.reason, "detail": item.detail}
                for item in self.problems[:MAX_PROBLEMS]
            ],
            "problems_shown": min(len(self.problems), MAX_PROBLEMS),
            "problems_total": len(self.problems),
        }


# --- Чтение файла ---------------------------------------------------------------


def _cell(value) -> str:
    """Значение ячейки как текст. Формулы не исполняются: книга открыта по значениям."""
    if value is None:
        return ""
    return str(value).strip()


def read_rows(path) -> tuple[list[Row], list[str]]:
    """Прочитать книгу. Возвращает строки и фактические заголовки."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - зависимость есть в проекте
        raise AnalogCatalogError("Не установлен openpyxl.") from exc

    path = Path(path)
    if path.suffix.lower() != ".xlsx":
        raise AnalogCatalogError("Ожидается файл .xlsx.")
    if not path.exists():
        raise AnalogCatalogError("Файл не найден.")

    try:
        # data_only=True отдаёт посчитанные значения, а не тексты формул, и
        # ничего не вычисляет само. read_only не держит книгу в памяти целиком.
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 - причина показывается человеку
        raise AnalogCatalogError(f"Файл не читается как Excel: {exc}") from exc

    try:
        worksheet = workbook[workbook.sheetnames[0]]
        headers: list[str] = []
        mapping: dict[int, str] = {}
        rows: list[Row] = []
        for index, raw in enumerate(worksheet.iter_rows(values_only=True), start=1):
            values = [_cell(cell) for cell in raw]
            if index == 1:
                headers = values
                mapping = _map_headers(values)
                missing = [name for name in REQUIRED if name not in mapping.values()]
                if missing:
                    raise AnalogCatalogError(
                        "В файле не хватает обязательных колонок: "
                        + ", ".join(_human(name) for name in missing)
                        + ". Заголовки первой строки: "
                        + (", ".join(part for part in values if part) or "пусто")
                    )
                continue
            if not any(values):
                continue
            if len(rows) >= MAX_ROWS:
                raise AnalogCatalogError(
                    f"В файле больше {MAX_ROWS} строк. Разделите его на части."
                )
            row = Row(number=index)
            for position, name in mapping.items():
                if position < len(values):
                    setattr(row, name, values[position])
            rows.append(row)
        if not headers:
            raise AnalogCatalogError("Файл пустой.")
        return rows, headers
    finally:
        workbook.close()


def _map_headers(values: list[str]) -> dict[int, str]:
    mapping: dict[int, str] = {}
    for position, title in enumerate(values):
        key = " ".join(title.lower().split())
        if not key:
            continue
        for name, aliases in COLUMNS.items():
            if key in aliases and name not in mapping.values():
                mapping[position] = name
                break
    return mapping


def _human(name: str) -> str:
    return COLUMNS[name][1] if len(COLUMNS[name]) > 1 else name


# --- Разрешение строки ------------------------------------------------------------


class Index:
    """Всё, что нужно для разбора файла, взятое из базы заранее.

    Без него разбор стоил бы двух обращений к базе на каждую строку: десять
    тысяч строк - двадцать тысяч запросов. Измерено до и после.

    Указатель живёт в пределах одного разбора и пополняется по ходу: детали,
    заведённые предыдущими строками того же файла, обязаны быть видны
    следующим, иначе они превращались бы в неоднозначность.
    """

    def __init__(self, articles: set[str]):
        self.by_article: dict[str, list[PartType]] = {}
        if articles:
            numbers = (
                PartNumber.objects.filter(
                    normalized_value__in=articles, kind__in=EXACT_NUMBER_KINDS
                )
                .select_related("part", "part__manufacturer")
            )
            for number in numbers:
                bucket = self.by_article.setdefault(number.normalized_value, [])
                if all(part.pk != number.part_id for part in bucket):
                    bucket.append(number.part)
        self.known_analogs: set[int] = set(
            PartAnalog.objects.values_list("analog_id", flat=True)
        )
        self.links: set[tuple[int, int]] = set(
            PartAnalog.objects.values_list("original_id", "analog_id")
        )

    def parts_by_article(self, article: str) -> list[PartType]:
        normalized = normalize_number(article)
        if not normalized:
            return []
        return list(self.by_article.get(normalized, ()))

    def remember_part(self, part: PartType, article: str) -> None:
        normalized = normalize_number(article)
        if normalized:
            self.by_article.setdefault(normalized, []).append(part)

    def remember_link(self, original_id: int, analog_id: int) -> None:
        self.known_analogs.add(analog_id)
        self.links.add((original_id, analog_id))


def build_index(rows: list[Row]) -> Index:
    articles = set()
    for row in rows:
        for value in (row.original_article, row.analog_article):
            normalized = normalize_number(value)
            if normalized:
                articles.add(normalized)
    return Index(articles)


def _narrow_by_manufacturer(parts, manufacturer: str):
    """Сузить по заводу, если он указан. Пустой завод ничего не отбрасывает."""
    wanted = " ".join((manufacturer or "").split()).lower()
    if not wanted:
        return list(parts)
    narrowed = [
        part for part in parts
        if part.manufacturer and part.manufacturer.name.strip().lower() == wanted
    ]
    return narrowed or list(parts)


def _prefer_not_already_analogs(parts: list, known_analogs: set[int]) -> list:
    """Из деталей с одним номером предпочесть те, что сами не числятся аналогами.

    Зачем это нужно. Каталог аналогов почти всегда содержит несколько строк для
    одной исходной детали, и у аналогов тот же номер. После первой же строки по
    этому номеру находятся уже две карточки, и дальше файл переставал бы
    разбираться вовсе.

    Здесь ничего не угадывается: используется только то, что система уже
    записала. Деталь, про которую кто-то сказал «это аналог вот той», исходной в
    новой строке не считается. Если после такого сужения остаётся больше одной
    детали, строка по-прежнему уходит человеку.
    """
    if len(parts) < 2:
        return list(parts)
    narrowed = [part for part in parts if part.pk not in known_analogs]
    return narrowed if len(narrowed) == 1 else list(parts)


def _clean_price(value: str):
    text = (value or "").strip().replace(",", ".")
    if not text:
        return None
    try:
        price = Decimal(text)
    except (InvalidOperation, ValueError) as exc:
        raise AnalogCatalogError(f"«{value}» не число") from exc
    if not price.is_finite() or price < 0:
        raise AnalogCatalogError(f"«{value}» не может быть ценой")
    return price.quantize(Decimal("0.01"))


@dataclass
class Resolution:
    """Что делать с одной строкой файла."""

    original: PartType | None = None
    analog: PartType | None = None
    create_analog: bool = False
    problem: Problem | None = None
    price = None


def resolve_row(row: Row, index: Index) -> Resolution:
    """Решить судьбу строки, ничего не записывая.

    Никогда не выбирает первую попавшуюся деталь: неоднозначность возвращается
    как проблема, чтобы её увидел человек.
    """
    name = " ".join((row.analog_name or "").split())
    if not row.original_article.strip():
        return Resolution(problem=Problem(row.number, "Не указан исходный артикул"))
    if not row.analog_article.strip() and not name:
        return Resolution(
            problem=Problem(row.number, "Не указан ни артикул аналога, ни название")
        )
    if not name:
        return Resolution(problem=Problem(row.number, "Не указано название аналога"))

    try:
        price = _clean_price(row.analog_price)
    except AnalogCatalogError as exc:
        return Resolution(problem=Problem(row.number, "Некорректная цена", str(exc)))

    # Аналог разрешается ПЕРВЫМ, и вот почему. У аналога артикул часто тот же,
    # что у исходной детали, поэтому обе карточки находятся по одному номеру.
    # Если сначала искать исходную, она окажется «неоднозначной» из-за самого
    # же аналога из этой строки - то есть флагманский случай перестал бы
    # работать. Деталь не может быть аналогом самой себя, поэтому найденный
    # аналог исключается из кандидатов на исходную.
    #
    # Совпадение по артикулу само по себе ничего не доказывает. Совпавшей
    # считается только деталь, у которой сошлись ВСЕ заполненные признаки:
    # артикул, название и, если он указан, завод.
    candidates = index.parts_by_article(row.analog_article)
    wanted_manufacturer = " ".join((row.analog_manufacturer or "").split()).lower()
    matched = []
    for part in candidates:
        if part.name.strip().lower() != name.lower():
            continue
        if wanted_manufacturer:
            actual = part.manufacturer.name.strip().lower() if part.manufacturer else ""
            if actual != wanted_manufacturer:
                continue
        matched.append(part)

    if len(matched) > 1:
        return Resolution(problem=Problem(
            row.number, "Несколько одинаковых деталей-аналогов",
            ", ".join(part.name for part in matched[:5]),
        ))
    analog = matched[0] if matched else None

    originals = _narrow_by_manufacturer(
        index.parts_by_article(row.original_article), row.original_manufacturer
    )
    if analog is not None:
        originals = [part for part in originals if part.pk != analog.pk]
    originals = _prefer_not_already_analogs(originals, index.known_analogs)
    if not originals:
        return Resolution(problem=Problem(
            row.number, "Не найдена исходная деталь",
            f"артикул {row.original_article}",
        ))
    if len(originals) > 1:
        return Resolution(problem=Problem(
            row.number, "Найдено несколько деталей с исходным артикулом",
            f"{row.original_article}: " + ", ".join(part.name for part in originals[:5])
            + ". Укажите колонку «Производитель оригинала», чтобы различить их.",
        ))

    resolution = Resolution(original=originals[0], analog=analog)
    resolution.price = price
    resolution.create_analog = analog is None
    return resolution


# --- План и применение --------------------------------------------------------------


def build_plan(path) -> Plan:
    """Полный разбор файла без единой записи."""
    rows, _ = read_rows(path)
    plan = Plan(rows_total=len(rows))
    index = build_index(rows)
    linked = index.links
    planned_pairs: set[tuple[int, int]] = set()

    for row in rows:
        resolution = resolve_row(row, index)
        if resolution.problem is not None and resolution.original is None:
            _count_problem(plan, resolution.problem)
            continue
        if resolution.problem is not None:
            _count_problem(plan, resolution.problem)
            continue

        if resolution.create_analog:
            plan.will_create_parts += 1
            plan.will_create_links += 1
            continue

        plan.will_reuse_parts += 1
        pair = (resolution.original.pk, resolution.analog.pk)
        if pair in linked or pair in planned_pairs:
            plan.already_linked += 1
        elif (pair[1], pair[0]) in linked:
            _count_problem(plan, Problem(
                row.number, "Эти детали уже связаны в обратную сторону",
                f"{resolution.analog.name} / {resolution.original.name}",
            ))
        else:
            plan.will_create_links += 1
            planned_pairs.add(pair)
    return plan


def _count_problem(plan: Plan, problem: Problem) -> None:
    if problem.reason.startswith("Найдено несколько") or problem.reason.startswith(
        "Несколько"
    ):
        plan.ambiguous += 1
    else:
        plan.invalid += 1
    plan.problems.append(problem)


@transaction.atomic
def apply_file(path, *, by=None) -> dict:
    """Применить разрешимые строки. Остатков не создаёт вовсе.

    Строки, требующие внимания, пропускаются и остаются в сводке: иначе из-за
    трёх спорных строк не удалось бы загрузить остальные тысячу двести.
    """
    rows, _ = read_rows(path)
    index = build_index(rows)
    created_parts = 0
    reused_parts = 0
    created_links = 0
    already_linked = 0
    skipped = 0
    problems: list[Problem] = []

    for row in rows:
        resolution = resolve_row(row, index)
        if resolution.problem is not None or resolution.original is None:
            skipped += 1
            if resolution.problem is not None:
                problems.append(resolution.problem)
            continue

        analog = resolution.analog
        if resolution.create_analog:
            try:
                analog = create_manual_part(
                    name=row.analog_name,
                    article=row.analog_article,
                    price=resolution.price,
                    barcode=row.analog_barcode,
                    manufacturer_name=row.analog_manufacturer,
                )
            except ManualPartError as exc:
                skipped += 1
                problems.append(Problem(row.number, "Деталь не создана", str(exc)))
                continue
            created_parts += 1
            # Следующие строки того же файла обязаны видеть заведённую деталь:
            # иначе она превратилась бы для них в неоднозначность.
            index.remember_part(analog, row.analog_article)
        else:
            reused_parts += 1

        try:
            _, created = link_analog(original=resolution.original, analog=analog, by=by)
        except AnalogLinkError as exc:
            skipped += 1
            problems.append(Problem(row.number, "Связь не создана", str(exc)))
            continue
        index.remember_link(resolution.original.pk, analog.pk)
        if created:
            created_links += 1
        else:
            already_linked += 1

    return {
        "rows_total": len(rows),
        "created_parts": created_parts,
        "reused_parts": reused_parts,
        "created_links": created_links,
        "already_linked": already_linked,
        "skipped": skipped,
        "problems": [
            {"row": item.row, "reason": item.reason, "detail": item.detail}
            for item in problems[:MAX_PROBLEMS]
        ],
        "problems_total": len(problems),
    }


def catalog_fingerprint() -> str:
    """Слепок справочника: поймать «каталог изменили между проверкой и применением»."""
    import hashlib

    from django.db.models import Max

    parts = PartType.objects.aggregate(total=Max("pk"), touched=Max("updated_at"))
    payload = "|".join([
        str(PartType.objects.count()),
        str(parts["total"]),
        str(parts["touched"]),
        str(PartAnalog.objects.count()),
        str(PartAnalog.objects.aggregate(total=Max("pk"))["total"]),
        str(PartNumber.objects.count()),
    ])
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()
