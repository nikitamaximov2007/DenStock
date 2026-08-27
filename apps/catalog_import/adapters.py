"""Адаптеры каталогов поставщиков.

Слой намеренно тонкий: разбор прайса BRP уже реализован и доказан на реальном
файле в `apps.brp.importer`, поэтому адаптер его ВЫЗЫВАЕТ, а не повторяет.
Второй парсер и вторая формула цены не создаются.

Добавить поставщика значит добавить один класс с тремя методами, а не менять
код рабочего процесса.
"""

from __future__ import annotations

import hashlib
from collections.abc import Mapping
from dataclasses import fields, is_dataclass
from decimal import Decimal
from pathlib import Path

from django.db.models import Max

BRP = "brp"
ANALOGS = "analogs"
AFTERMARKET = "aftermarket"


class CatalogAdapterError(RuntimeError):
    """Файл не разобран: понятная причина для пользователя."""


class CatalogAdapter:
    """Контракт адаптера каталога."""

    key = ""
    label = ""

    def check(self, path: Path) -> dict:
        """Разобрать файл БЕЗ записи и вернуть сводку. Обязан быть read-only."""
        raise NotImplementedError

    def apply(self, path: Path) -> dict:
        """Применить разобранный файл к справочнику."""
        raise NotImplementedError

    def fingerprint(self) -> str:
        """Слепок текущего состояния каталога для защиты от устаревшего предпросмотра."""
        raise NotImplementedError

    def validation_error(self, summary: dict) -> str | None:
        """Return a user-facing reason why a checked file must not be applied."""
        return None


def _summary_dict(summary) -> dict:
    """Сводка импортёра в JSON-совместимый вид.

    dataclasses.asdict() нельзя, он ломает Counter. Внутри asdict
    словарь-подкласс пересобирается как `Counter(последовательность пар)`, и
    ключами становятся кортежи вида ('OBS', 1). Поэтому поля читаются напрямую,
    а любые отображения приводятся к обычному словарю со строковыми ключами.
    """
    if is_dataclass(summary):
        data = {item.name: getattr(summary, item.name) for item in fields(summary)}
    else:
        data = dict(summary)
    result = {}
    for key, value in data.items():
        if isinstance(value, Mapping):
            result[key] = {str(inner): value[inner] for inner in value}
        elif isinstance(value, Decimal):
            result[key] = str(value)
        else:
            result[key] = value
    return result


class BrpCatalogAdapter(CatalogAdapter):
    key = BRP
    label = "BRP"

    def check(self, path: Path) -> dict:
        from apps.brp.importer import BrpImportError, import_catalog

        try:
            summary = import_catalog(str(path), commit=False)
        except BrpImportError as exc:
            raise CatalogAdapterError(str(exc)) from exc
        return _summary_dict(summary)

    def apply(self, path: Path) -> dict:
        from apps.brp.importer import BrpImportError, import_catalog

        try:
            summary = import_catalog(str(path), commit=True)
        except BrpImportError as exc:
            raise CatalogAdapterError(str(exc)) from exc
        return _summary_dict(summary)

    def fingerprint(self) -> str:
        """Слепок каталога BRP: сколько позиций и когда последняя правка.

        Этого достаточно, чтобы поймать «каталог поменяли между проверкой и
        применением»: любой импорт, ручная правка или удаление сдвигают либо
        счётчик, либо отметку времени.
        """
        from apps.brp.models import BrpCatalogPart

        aggregate = BrpCatalogPart.objects.aggregate(
            total=Max("pk"), touched=Max("updated_at")
        )
        count = BrpCatalogPart.objects.count()
        payload = f"{count}|{aggregate['total']}|{aggregate['touched']}"
        return hashlib.sha256(payload.encode("utf-8")).hexdigest()

    def validation_error(self, summary: dict) -> str | None:
        ambiguous = int(summary.get("ambiguous_nonzero_wholesale", 0) or 0)
        invalid = int(summary.get("invalid_wholesale_price", 0) or 0)
        negative = int(summary.get("negative_wholesale_price", 0) or 0)
        if ambiguous or invalid or negative:
            return (
                "В прайсе есть неоднозначные, отрицательные или некорректные "
                "оптовые цены. Применение заблокировано до разбора поставщиком."
            )
        return None

    def inspect(self, path: Path, *, sample_rows: int = 5) -> dict:
        """Read-only разбор структуры книги: листы, заголовки, первые строки.

        Нужен, когда поставщик пришлёт файл другой формы: инспектор показывает
        фактическую структуру, вместо того чтобы парсер угадывал колонки.
        """
        return inspect_workbook(path, sample_rows=sample_rows)


def inspect_workbook(path, *, sample_rows: int = 5) -> dict:
    """Показать структуру xlsx, ничего не меняя и не загружая книгу целиком."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - зависимость есть в проекте
        raise CatalogAdapterError("Не установлен openpyxl.") from exc

    path = Path(path)
    if path.suffix.lower() != ".xlsx":
        raise CatalogAdapterError("Ожидается файл .xlsx.")
    if not path.exists():
        raise CatalogAdapterError("Файл не найден.")
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 - причина показывается пользователю
        raise CatalogAdapterError(f"Файл не читается как Excel: {exc}") from exc
    try:
        sheets = list(workbook.sheetnames)
        worksheet = workbook[sheets[0]]
        headers: list[str] = []
        samples: list[list[str]] = []
        # read_only-итератор не держит книгу в памяти: берём только начало.
        for index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            values = ["" if cell is None else str(cell).strip() for cell in row]
            if index == 1:
                headers = values
            elif len(samples) < sample_rows:
                samples.append(values)
            else:
                break
        return {
            "sheets": sheets,
            "sheet": sheets[0],
            "headers": headers,
            "sample_rows": samples,
            "declared_max_row": worksheet.max_row,
            "declared_max_column": worksheet.max_column,
        }
    finally:
        workbook.close()


class AnalogCatalogAdapter(CatalogAdapter):
    """Каталог аналогов: связывает уже заведённые детали и заводит недостающие.

    Разбор живёт отдельным модулем, адаптер только соединяет его с общим
    рабочим процессом: проверка, предпросмотр, применение, защита от устаревшего
    предпросмотра. Второй такой процесс не заводится.
    """

    key = ANALOGS
    label = "Аналоги"

    def check(self, path: Path) -> dict:
        from apps.catalog_import.analog_catalog import AnalogCatalogError, build_plan

        try:
            return build_plan(path).as_summary()
        except AnalogCatalogError as exc:
            raise CatalogAdapterError(str(exc)) from exc

    def apply(self, path: Path) -> dict:
        from apps.catalog_import.analog_catalog import AnalogCatalogError, apply_file

        try:
            return apply_file(path)
        except AnalogCatalogError as exc:
            raise CatalogAdapterError(str(exc)) from exc

    def fingerprint(self) -> str:
        from apps.catalog_import.analog_catalog import catalog_fingerprint

        return catalog_fingerprint()

    def validation_error(self, summary: dict) -> str | None:
        """Спорные строки применение не блокируют.

        Из-за трёх неоднозначных строк не должно срываться заведение остальной
        тысячи: они пропускаются и остаются в сводке. Блокировать имеет смысл
        только файл, в котором применять нечего вовсе.
        """
        useful = int(summary.get("will_create_links", 0) or 0)
        reused = int(summary.get("already_linked", 0) or 0)
        if useful == 0 and reused == 0:
            return (
                "В файле нет ни одной строки, которую можно применить. "
                "Проверьте колонки и артикулы исходных деталей."
            )
        return None


class AftermarketCatalogAdapter(CatalogAdapter):
    """Known dealer format with independent aftermarket parts and USD facts."""

    key = AFTERMARKET
    label = "Каталог аналогов / aftermarket"

    def check(self, path: Path) -> dict:
        from apps.catalog_import.aftermarket_catalog import AftermarketCatalogError, build_plan

        try:
            return build_plan(path).as_summary()
        except AftermarketCatalogError as exc:
            raise CatalogAdapterError(str(exc)) from exc

    def apply(self, path: Path) -> dict:
        from apps.catalog_import.aftermarket_catalog import AftermarketCatalogError, apply_file

        try:
            return apply_file(path)
        except AftermarketCatalogError as exc:
            raise CatalogAdapterError(str(exc)) from exc

    def fingerprint(self) -> str:
        from apps.catalog_import.aftermarket_catalog import catalog_fingerprint

        return catalog_fingerprint()

    def validation_error(self, summary: dict) -> str | None:
        if int(summary.get("valid", 0) or 0) == 0:
            return "В файле нет ни одной безопасно применимой строки aftermarket-каталога."
        return None

    def inspect(self, path: Path, *, sample_rows: int = 5) -> dict:
        return inspect_workbook(path, sample_rows=sample_rows)


ADAPTERS: dict[str, CatalogAdapter] = {
    BRP: BrpCatalogAdapter(),
    ANALOGS: AnalogCatalogAdapter(),
    AFTERMARKET: AftermarketCatalogAdapter(),
}


def get_adapter(catalog: str) -> CatalogAdapter:
    adapter = ADAPTERS.get(catalog)
    if adapter is None:
        raise CatalogAdapterError("Неизвестный каталог.")
    return adapter


def detect_catalog(path: Path, selected: str) -> str:
    """Recognize the known dealer sheet before the normal analog parser rejects it."""
    if selected != ANALOGS:
        return selected
    try:
        from openpyxl import load_workbook

        book = load_workbook(path, read_only=True, data_only=True)
        try:
            # Поставщик приписывает к вкладке своё имя («diorlight priceupdate»),
            # поэтому лист ищется по вхождению ключевого слова. Если подходящих
            # листов несколько, книгу всё равно ведём в aftermarket: там она
            # получит внятный отказ с перечислением, а не чужую ошибку разбора.
            candidates = [
                name for name in book.sheetnames
                if "priceupdate" in " ".join(str(name).lower().split()).replace(" ", "")
            ]
            if not candidates:
                return selected
            for name in candidates:
                headers = {
                    " ".join(str(cell.value or "").replace("\xa0", " ").lower().split())
                    for cell in next(book[name].iter_rows())
                }
                if {"manufacturer", "manufacturer number", "description"} <= headers:
                    return AFTERMARKET
        finally:
            book.close()
    except Exception:  # The selected adapter renders the user-facing parse error.
        return selected
    return selected


def file_sha256(path) -> str:
    digest = hashlib.sha256()
    with open(path, "rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()
