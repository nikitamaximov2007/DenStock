"""Import Polaris dealer price Excel into reference catalog rows only."""
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.utils import timezone

from apps.catalog.models import normalize_number

from .models import PolarisCatalogPart

CHUNK_SIZE = 1000
EXPECTED_HEADERS = [
    "part_number",
    "part_name",
    "superseded_number",
    "ОПТОВАЯ",
    "РОЗНИЦА",
    "uom",
]
DATA_COLUMNS = len(EXPECTED_HEADERS)
ZERO = Decimal("0")
SYNC_FIELDS = (
    "part_name",
    "superseded_number",
    "wholesale_price_usd",
    "retail_price_usd",
    "uom",
)
UPDATE_FIELDS = SYNC_FIELDS + (
    "part_number_norm",
    "superseded_number_norm",
    "source_file",
    "source_row",
    "import_batch",
    "updated_at",
)


class PolarisImportError(Exception):
    """The Excel file cannot be imported."""


@dataclass
class ImportSummary:
    mode: str = "dry-run"
    total_rows: int = 0
    data_rows: int = 0
    created: int = 0
    updated: int = 0
    skipped_unchanged: int = 0
    skipped_empty: int = 0
    no_retail_price: int = 0
    with_superseded: int = 0
    errors: int = 0
    error_examples: list[str] = field(default_factory=list)


def _text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _dec(value, row_no: int, column: str, summary: ImportSummary):
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value).replace(",", ".").replace(" ", ""))
    except InvalidOperation:
        _add_error(summary, f"строка {row_no}: некорректная цена {column}={value!r}")
        return None


def _add_error(summary: ImportSummary, text: str) -> None:
    summary.errors += 1
    if len(summary.error_examples) < 10:
        summary.error_examples.append(text)


def _row_dict(cells, row_no: int, summary: ImportSummary) -> dict | None:
    padded = list(cells[:DATA_COLUMNS]) + [None] * (DATA_COLUMNS - len(cells))
    if all(_text(value) == "" for value in padded):
        summary.skipped_empty += 1
        return None
    part_number = _text(padded[0])
    if not part_number:
        _add_error(summary, f"строка {row_no}: пустой part_number")
        return None
    return {
        "part_number": part_number[:40],
        "part_name": _text(padded[1])[:255],
        "superseded_number": _text(padded[2])[:40],
        "wholesale_price_usd": _dec(padded[3], row_no, "ОПТОВАЯ", summary),
        "retail_price_usd": _dec(padded[4], row_no, "РОЗНИЦА", summary),
        "uom": _text(padded[5])[:40],
        "source_row": row_no,
    }


def _differs(obj: PolarisCatalogPart, row: dict) -> bool:
    return any(getattr(obj, name) != row[name] for name in SYNC_FIELDS)


def _apply(obj: PolarisCatalogPart, row: dict, *, source_file: str, batch: str) -> None:
    for name in SYNC_FIELDS + ("source_row",):
        setattr(obj, name, row[name])
    obj.source_file = source_file
    obj.import_batch = batch
    obj.part_number_norm = normalize_number(obj.part_number)
    obj.superseded_number_norm = normalize_number(obj.superseded_number)
    obj.updated_at = timezone.now()


def _flush(chunk: list[dict], summary: ImportSummary, *,
           commit: bool, source_file: str, batch: str) -> None:
    keys = [row["part_number"] for row in chunk]
    existing = {
        obj.part_number: obj
        for obj in PolarisCatalogPart.objects.filter(part_number__in=keys)
    }
    to_create, to_update = [], []
    for row in chunk:
        obj = existing.get(row["part_number"])
        if obj is None:
            obj = PolarisCatalogPart(part_number=row["part_number"])
            _apply(obj, row, source_file=source_file, batch=batch)
            to_create.append(obj)
            summary.created += 1
        elif _differs(obj, row):
            _apply(obj, row, source_file=source_file, batch=batch)
            to_update.append(obj)
            summary.updated += 1
        else:
            summary.skipped_unchanged += 1
    if commit:
        if to_create:
            PolarisCatalogPart.objects.bulk_create(to_create, batch_size=CHUNK_SIZE)
        if to_update:
            PolarisCatalogPart.objects.bulk_update(
                to_update, UPDATE_FIELDS, batch_size=CHUNK_SIZE
            )


def _open_worksheet(path: Path, sheet: str | None):
    import openpyxl

    try:
        workbook = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise PolarisImportError(f"Не удалось открыть Excel: {exc}") from exc
    worksheet = workbook[sheet] if sheet else workbook[workbook.sheetnames[0]]
    return workbook, worksheet


def _validate_headers(worksheet) -> None:
    headers = [_text(value) for value in next(worksheet.iter_rows(max_row=1, values_only=True))]
    actual = headers[:DATA_COLUMNS]
    if actual != EXPECTED_HEADERS:
        expected = ", ".join(EXPECTED_HEADERS)
        got = ", ".join(actual)
        raise PolarisImportError(
            f"Неверные заголовки Polaris-файла. Ожидалось: {expected}. Получено: {got}."
        )


def import_catalog(path, *, commit: bool = False, sheet: str | None = None) -> ImportSummary:
    """Parse and upsert Polaris catalog rows. Dry-run does not write."""
    path = Path(path)
    if not path.exists():
        raise PolarisImportError(f"Файл не найден: {path}")
    summary = ImportSummary(mode="commit" if commit else "dry-run")
    batch = timezone.now().strftime("%Y-%m-%d_%H-%M-%S")
    workbook, worksheet = _open_worksheet(path, sheet)
    try:
        _validate_headers(worksheet)
        chunk: list[dict] = []
        for row_no, cells in enumerate(
            worksheet.iter_rows(min_row=2, values_only=True), start=2
        ):
            summary.total_rows += 1
            row = _row_dict(cells, row_no, summary)
            if row is None:
                continue
            summary.data_rows += 1
            retail = row["retail_price_usd"]
            if retail is None or retail <= ZERO:
                summary.no_retail_price += 1
            if row["superseded_number"]:
                summary.with_superseded += 1
            chunk.append(row)
            if len(chunk) >= CHUNK_SIZE:
                _flush(chunk, summary, commit=commit, source_file=path.name, batch=batch)
                chunk = []
        if chunk:
            _flush(chunk, summary, commit=commit, source_file=path.name, batch=batch)
        return summary
    finally:
        workbook.close()

