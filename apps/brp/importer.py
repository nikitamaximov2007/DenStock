"""Layer 31/32.3 — импорт дилерского прайса BRP из Excel в справочник.

СТРОГО справочник: импорт не создаёт остатков, движений, поступлений и не
удаляет складские карточки. Идемпотентен: повторный запуск обновляет только
изменившиеся строки. 127 тысяч строк обрабатываются чанками через bulk-операции.

Дубликаты Material_No (hotfix 32.3): в файле встречаются повторы одного
номера, где у одной строки оптовая цена 0, а у другой — реальная цена. Правило
выбора лучшей строки детерминировано:
1) предпочесть строку с wholesale_price_usd > 0;
2) при равенстве — строку с retail_price_usd > 0;
3) при полном равенстве — первую по порядку в файле;
4) если у всех дубликатов нет пригодной оптовой цены, цена считается
   отсутствующей; при обновлении существующей позиции сохраняется её прежняя
   ненулевая wholesale.
Повторный импорт того же файла ЧИНИТ существующие записи с нулевой ценой:
логика «пропустить без изменений» сравнивает выбранную лучшую строку с базой
и обновляет отличающиеся записи (счётчик zero_wholesale_price_repaired).

Формат файла (проверен на реальном прайсе):
- первый лист; строка 1 — заголовки; данные начинаются со строки 2. Старые
  файлы с пустой строкой примечаний после заголовка поддерживаются;
- значимые колонки определяются по заголовку, а не позиции: Material_No,
  Part_Desc, Last_Yr_Util, Status, РОЗНИЦА, ОПТОВАЯ и две ЗАМЕНА НОМЕРА.
  Поддерживаются оба официальных порядка цен и замен;
- колонки правее H (легенда статусов) игнорируются;
- Material_No хранится СТРОКОЙ (ведущие нули не теряются), пробелы обрезаются,
  пустые ячейки нормализуются.

Файл читается в два прохода, чтобы не держать 127k строк в памяти:
проход 1 выбирает номер лучшей строки на каждый Material_No, проход 2
обрабатывает только выбранные строки чанками.
"""
from collections import Counter
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.utils import timezone

from apps.catalog.models import normalize_number

from .models import BrpCatalogPart

# Reading and writing a supplier workbook uses bounded chunks. Changed existing
# rows are written through an upsert instead of Django's CASE-based bulk_update:
# a full 130k snapshot otherwise creates excessively expensive SQL statements.
CHUNK_SIZE = 4000
ZERO = Decimal("0")
STATUS_ALIASES = {"UCP": "USE"}

# Поля, которые синхронизируются из файла при обновлении существующей строки.
SYNC_FIELDS = (
    "part_desc", "last_year_util", "brp_status",
    "retail_price_usd", "wholesale_price_usd",
    "replacement_no_1", "replacement_no_2",
)
UPDATE_FIELDS = SYNC_FIELDS + (
    "material_no_norm", "replacement_no_1_norm", "replacement_no_2_norm",
    "source_file", "source_row", "import_batch", "is_current", "updated_at",
)


class BrpImportError(Exception):
    """Файл не может быть разобран (нет файла/листа/колонок)."""


@dataclass
class ImportSummary:
    mode: str = "dry-run"
    total_rows_scanned: int = 0
    data_rows: int = 0
    created: int = 0
    updated: int = 0
    reactivated: int = 0
    deactivated: int = 0
    skipped_unchanged: int = 0
    skipped_empty: int = 0
    duplicates: int = 0
    duplicates_price_resolved: int = 0  # дубликат выиграл по правилу ненулевой цены
    zero_wholesale_price_repaired: int = 0
    new_file_nonzero_price: int = 0
    same_file_nonzero_fallback: int = 0
    previous_catalog_price_retained: int = 0
    no_usable_price: int = 0
    ambiguous_nonzero_wholesale: int = 0
    conflicting_nonzero_wholesale: int = 0
    invalid_wholesale_price: int = 0
    negative_wholesale_price: int = 0
    unique_materials: int = 0
    with_retail_price: int = 0
    with_wholesale_price: int = 0
    with_replacement: int = 0
    status_counts: Counter = field(default_factory=Counter)
    recommended_prices_refreshed: int = 0


def _text(value) -> str:
    """Ячейка -> строка: None -> '', числа -> str без потери, пробелы обрезаны."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))  # Excel любит превращать номера в 460041.0
    return str(value).strip()


def _dec(value):
    """Ячейка -> Decimal или None (пустое/нечисловое)."""
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value).replace(",", ".").replace(" ", ""))
    except InvalidOperation:
        return None


def normalize_status(value) -> str:
    """Привести supplier status к canonical BRP-коду на входе импорта."""
    status = _text(value).upper()
    return STATUS_ALIASES.get(status, status)


def _header_name(value) -> str:
    """Нормализовать заголовок Excel без привязки к регистру и пробелам."""
    return "".join(char for char in _text(value).casefold() if char.isalnum())


def _column_map(worksheet) -> dict[str, int]:
    """Вернуть индексы обязательных колонок поддерживаемого BRP-листа."""
    headers = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    positions: dict[str, list[int]] = {}
    for index, value in enumerate(headers):
        positions.setdefault(_header_name(value), []).append(index)

    aliases = {
        "material_no": ("materialno",),
        "part_desc": ("partdesc",),
        "last_year_util": ("lastyrutil",),
        "brp_status": ("status",),
        "retail_price_usd": ("розница", "розницаusd", "retail", "retailusd"),
        "wholesale_price_usd": ("оптовая", "оптоваяusd", "wholesale", "wholesaleusd"),
    }
    columns: dict[str, int] = {}
    missing = []
    for field_name, names in aliases.items():
        indexes = [index for name in names for index in positions.get(name, [])]
        if len(indexes) != 1:
            missing.append(names[0])
        else:
            columns[field_name] = indexes[0]

    replacements = positions.get("заменаномера", [])
    if len(replacements) >= 2:
        columns["replacement_no_1"], columns["replacement_no_2"] = replacements[:2]
    else:
        first = positions.get("replacement1", [])
        second = positions.get("replacement2", [])
        if len(first) == 1 and len(second) == 1:
            columns["replacement_no_1"] = first[0]
            columns["replacement_no_2"] = second[0]
        else:
            missing.append("ЗАМЕНА НОМЕРА (две колонки)")

    if missing:
        raise BrpImportError(
            "Формат Excel не поддерживается: отсутствуют или повторяются колонки "
            + ", ".join(missing)
            + "."
        )
    return columns


def _cell(cells, index: int):
    return cells[index] if index < len(cells) else None


def _row_dict(cells, row_no: int, columns: dict[str, int]) -> dict:
    wholesale_raw = _cell(cells, columns["wholesale_price_usd"])
    wholesale = _dec(wholesale_raw)
    return {
        "material_no": _text(_cell(cells, columns["material_no"])),
        "part_desc": _text(_cell(cells, columns["part_desc"]))[:255],
        "last_year_util": _text(_cell(cells, columns["last_year_util"]))[:20],
        "brp_status": normalize_status(_cell(cells, columns["brp_status"]))[:20],
        "retail_price_usd": _dec(_cell(cells, columns["retail_price_usd"])),
        "wholesale_price_usd": wholesale,
        "wholesale_price_invalid": bool(_text(wholesale_raw)) and wholesale is None,
        "replacement_no_1": _text(_cell(cells, columns["replacement_no_1"]))[:40],
        "replacement_no_2": _text(_cell(cells, columns["replacement_no_2"]))[:40],
        "source_row": row_no,
    }


def _price_score(row: dict) -> tuple[int, int]:
    """Ранг строки для выбора среди дубликатов: только Decimal, без float.

    (оптовая > 0, розница > 0); лексикографическое сравнение кортежей даёт
    правило «сначала ненулевая оптовая, затем ненулевая розница». Побеждает
    строго больший ранг, при равенстве остаётся более ранняя строка файла.
    """
    wholesale = row["wholesale_price_usd"]
    retail = row["retail_price_usd"]
    return (
        1 if wholesale is not None and wholesale > ZERO else 0,
        1 if retail is not None and retail > ZERO else 0,
    )


def _differs(obj: BrpCatalogPart, row: dict) -> bool:
    return any(getattr(obj, name) != row[name] for name in SYNC_FIELDS)


def _apply(obj: BrpCatalogPart, row: dict, *, source_file: str, batch: str) -> None:
    for name in SYNC_FIELDS + ("source_row",):
        setattr(obj, name, row[name])
    obj.source_file = source_file
    obj.import_batch = batch
    obj.is_current = True
    # bulk_create/bulk_update не вызывают save(): нормализацию считаем сами.
    obj.material_no_norm = normalize_number(obj.material_no)
    obj.replacement_no_1_norm = normalize_number(obj.replacement_no_1)
    obj.replacement_no_2_norm = normalize_number(obj.replacement_no_2)
    obj.updated_at = timezone.now()


def _choose_wholesale(row: dict, obj: BrpCatalogPart | None) -> str:
    """Resolve the supplier raw price without letting a zero erase a known price.

    Metadata remains from ``row``. The selected raw wholesale is either a
    usable value from the current workbook, a previous positive catalog price,
    or ``None`` when the supplier did not provide any usable price.
    """
    incoming = row["wholesale_price_usd"]
    if incoming is not None and incoming > ZERO:
        row["price_source"] = "new_file"
        return "new_file"
    previous = getattr(obj, "wholesale_price_usd", None)
    if previous is not None and previous > ZERO:
        row["wholesale_price_usd"] = previous
        row["price_source"] = "previous_catalog"
        return "previous_catalog"
    row["wholesale_price_usd"] = None
    row["price_source"] = "missing"
    return "missing"


def _flush(chunk: list[dict], summary: ImportSummary, *,
           commit: bool, source_file: str, batch: str) -> None:
    keys = [row["material_no"] for row in chunk]
    existing = {
        obj.material_no: obj
        for obj in BrpCatalogPart.objects.filter(material_no__in=keys)
    }
    to_create, to_update = [], []
    for row in chunk:
        obj = existing.get(row["material_no"])
        price_source = _choose_wholesale(row, obj)
        if price_source == "new_file":
            summary.new_file_nonzero_price += 1
        elif price_source == "previous_catalog":
            summary.previous_catalog_price_retained += 1
        else:
            summary.no_usable_price += 1
        if obj is None:
            obj = BrpCatalogPart(material_no=row["material_no"])
            _apply(obj, row, source_file=source_file, batch=batch)
            to_create.append(obj)
            summary.created += 1
        elif _differs(obj, row) or not obj.is_current:
            was_current = obj.is_current
            old_wholesale = obj.wholesale_price_usd
            new_wholesale = row["wholesale_price_usd"]
            if (old_wholesale is None or old_wholesale == ZERO) and (
                new_wholesale is not None and new_wholesale > ZERO
            ):
                summary.zero_wholesale_price_repaired += 1
            _apply(obj, row, source_file=source_file, batch=batch)
            to_update.append(obj)
            summary.updated += 1
            if not was_current:
                summary.reactivated += 1
        else:
            summary.skipped_unchanged += 1
    if commit:
        if to_create:
            BrpCatalogPart.objects.bulk_create(to_create, batch_size=CHUNK_SIZE)
        if to_update:
            BrpCatalogPart.objects.bulk_create(
                to_update,
                batch_size=CHUNK_SIZE,
                update_conflicts=True,
                update_fields=UPDATE_FIELDS,
                unique_fields=["material_no"],
            )


def _open_worksheet(path: Path, sheet: str | None):
    import openpyxl

    try:
        workbook = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 — любой битый xlsx -> понятная ошибка
        raise BrpImportError(f"Не удалось открыть Excel: {exc}") from exc
    worksheet = workbook[sheet] if sheet else workbook[workbook.sheetnames[0]]
    return workbook, worksheet


def _select_best_rows(
    worksheet, summary: ImportSummary, columns: dict[str, int]
) -> dict[str, int]:
    """Проход 1: для каждого Material_No выбрать номер лучшей строки файла."""
    best: dict[str, tuple[tuple[int, int], int]] = {}
    positive_prices: dict[str, set[Decimal]] = {}
    for row_no, cells in enumerate(
        worksheet.iter_rows(min_row=2, values_only=True), start=2
    ):
        summary.total_rows_scanned += 1
        row = _row_dict(cells, row_no, columns)
        material = row["material_no"]
        if not material:
            summary.skipped_empty += 1
            continue
        if row["wholesale_price_invalid"]:
            summary.invalid_wholesale_price += 1
        elif row["wholesale_price_usd"] is not None and row["wholesale_price_usd"] < ZERO:
            summary.negative_wholesale_price += 1
        elif row["wholesale_price_usd"] is not None and row["wholesale_price_usd"] > ZERO:
            positive_prices.setdefault(material, set()).add(row["wholesale_price_usd"])
        score = _price_score(row)
        kept = best.get(material)
        if kept is None:
            best[material] = (score, row_no)
            continue
        summary.duplicates += 1
        if score > kept[0]:  # строго лучше по цене: дубликат побеждает
            best[material] = (score, row_no)
            summary.duplicates_price_resolved += 1
    # The supplier format has an explicit stable tie-breaker: when two rows
    # have the same rank, the first source row wins. Keep reporting differing
    # positive duplicates for operator review, but they are not ambiguous to
    # this deterministic importer.
    summary.conflicting_nonzero_wholesale = sum(
        len(prices) > 1 for prices in positive_prices.values()
    )
    return {material: row_no for material, (_score, row_no) in best.items()}


def _has_blocking_price_issues(summary: ImportSummary) -> bool:
    return bool(
        summary.ambiguous_nonzero_wholesale
        or summary.invalid_wholesale_price
        or summary.negative_wholesale_price
    )


def selected_wholesale_prices(
    path, *, sheet: str | None = None
) -> tuple[dict[str, Decimal | None], ImportSummary]:
    """Read the deterministic supplier winner for each material without DB writes.

    The correction workflow uses this same parser for the applied supplier file
    and the authoritative previous file. It deliberately returns raw supplier
    values: choosing a fallback belongs to the caller that knows both sources.
    """
    path = Path(path)
    if not path.exists():
        raise BrpImportError(f"Файл не найден: {path}")
    summary = ImportSummary(mode="dry-run")
    workbook, worksheet = _open_worksheet(path, sheet)
    try:
        selected_rows = _select_best_rows(worksheet, summary, _column_map(worksheet))
    finally:
        workbook.close()
    winners = set(selected_rows.values())
    prices: dict[str, Decimal | None] = {}
    workbook, worksheet = _open_worksheet(path, sheet)
    try:
        columns = _column_map(worksheet)
        for row_no, cells in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if row_no not in winners:
                continue
            row = _row_dict(cells, row_no, columns)
            prices[row["material_no"]] = row["wholesale_price_usd"]
    finally:
        workbook.close()
    summary.unique_materials = len(prices)
    return prices, summary


def import_catalog(path, *, commit: bool = False, sheet: str | None = None) -> ImportSummary:
    """Разобрать Excel и синхронизировать справочник. dry-run ничего не пишет."""
    path = Path(path)
    if not path.exists():
        raise BrpImportError(f"Файл не найден: {path}")

    summary = ImportSummary(mode="commit" if commit else "dry-run")
    batch = timezone.now().strftime("%Y-%m-%d_%H-%M-%S")

    # Проход 1: выбор лучших строк (в памяти только номера строк и ранги).
    workbook, worksheet = _open_worksheet(path, sheet)
    try:
        selected_rows = _select_best_rows(worksheet, summary, _column_map(worksheet))
    finally:
        workbook.close()
    if commit and _has_blocking_price_issues(summary):
        raise BrpImportError(
            "В файле есть неоднозначные, отрицательные или некорректные оптовые цены. "
            "Применение заблокировано до их разбора."
        )
    summary.unique_materials = len(selected_rows)
    summary.deactivated = BrpCatalogPart.objects.filter(is_current=True).exclude(
        material_no__in=selected_rows
    ).count()
    winners = set(selected_rows.values())

    # Проход 2: обработать только выбранные строки, чанками.
    workbook, worksheet = _open_worksheet(path, sheet)
    try:
        columns = _column_map(worksheet)
        chunk: list[dict] = []
        for row_no, cells in enumerate(
            worksheet.iter_rows(min_row=2, values_only=True), start=2
        ):
            if row_no not in winners:
                continue
            row = _row_dict(cells, row_no, columns)
            summary.data_rows += 1
            if row["brp_status"]:
                summary.status_counts[row["brp_status"]] += 1
            if row["retail_price_usd"] is not None:
                summary.with_retail_price += 1
            if row["wholesale_price_usd"] is not None:
                summary.with_wholesale_price += 1
            if row["replacement_no_1"] or row["replacement_no_2"]:
                summary.with_replacement += 1
            chunk.append(row)
            if len(chunk) >= CHUNK_SIZE:
                _flush(chunk, summary, commit=commit,
                       source_file=path.name, batch=batch)
                chunk = []
        if chunk:
            _flush(chunk, summary, commit=commit, source_file=path.name, batch=batch)
        if commit:
            # Every BRP workbook is a complete supplier snapshot. This runs in
            # the caller's apply transaction, so readers never observe a
            # partially replaced current catalog after a failed apply.
            summary.deactivated = BrpCatalogPart.objects.filter(is_current=True).exclude(
                material_no__in=selected_rows
            ).update(is_current=False, updated_at=timezone.now())
            from apps.catalog.services import get_current_price_settings, refresh_linked_part_prices

            pricing = get_current_price_settings()
            summary.recommended_prices_refreshed = refresh_linked_part_prices(
                usd_rate=pricing.current_usd_rate,
                brp_markup=pricing.brp_markup_percent,
                polaris_markup=pricing.polaris_markup_percent,
                catalogs=frozenset({"brp"}),
            )
        return summary
    finally:
        workbook.close()
