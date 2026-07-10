"""Таможенный Excel-экспорт: регрессия production-500 и контракт файла.

Первопричина 500 на production: шаблон «Формы для заказа» лежал в docs/,
а docs/ исключён из Docker-образа (.dockerignore) — openpyxl.load_workbook
падал с FileNotFoundError. Шаблон перенесён в пакет приложения.

Здесь же закреплён фактический контракт файла (лист, колонки, группировка,
exact identity, фильтры) и защита Excel: управляющие символы и formula
injection. Экспорт read-only: ни движений, ни действий, ни лотов не меняет.
"""
import datetime
from decimal import Decimal
from io import BytesIO

import openpyxl
import pytest
from django.contrib.auth.models import Group
from django.urls import reverse

from apps.actions.models import PartCustomsInfo, WarehouseAction
from apps.actions.services import (
    TEMPLATE_PATH,
    ActionError,
    cancel_warehouse_action,
    excel_safe_text,
    perform_action,
)
from apps.brp.models import BrpCatalogPart
from apps.brp.services import promote_to_warehouse as promote_brp
from apps.catalog.models import Category, PartNumber, PartType, Unit
from apps.inventory.models import StockLot, StockMovement
from apps.inventory.services import create_stock_lot, receive_stock_lot
from apps.polaris.models import PolarisCatalogPart
from apps.polaris.services import promote_to_warehouse as promote_polaris
from apps.procurement.models import Batch, BatchLine
from apps.procurement.services import finalize_cost
from apps.suppliers.models import Supplier
from apps.warehouse.models import StorageLocation

PASSWORD = "parol-12345"
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
SHEET = "Лист1"
DATA_ROW = 10


@pytest.fixture
def make_user(db, django_user_model):
    def _make(username, *, role=None, is_superuser=False):
        if is_superuser:
            user = django_user_model.objects.create_superuser(username=username, password=PASSWORD)
        else:
            user = django_user_model.objects.create_user(username=username, password=PASSWORD)
        if role:
            user.groups.add(Group.objects.get(name=role))
        return user

    return _make


@pytest.fixture
def admin(make_user):
    return make_user("admin", is_superuser=True)


def _stock(part, location, qty, sup, admin):
    batch = Batch.objects.create(supplier=sup, shipping_cost=Decimal("0"))
    line = BatchLine.objects.create(
        batch=batch, part_type=part, quantity=Decimal(str(qty)),
        unit_cost_currency=Decimal("1"),
    )
    batch.status = Batch.Status.ACCEPTED
    batch.save(update_fields=["status"])
    finalize_cost(batch, admin)
    line.refresh_from_db()
    lot = create_stock_lot(line, location, Decimal(str(qty)))
    receive_stock_lot(lot, by=admin)
    return lot


@pytest.fixture
def env(db, admin):
    sup = Supplier.objects.create(name="Стартовый ввод")
    loc = StorageLocation.objects.create(
        name="C04", code="S04-L03-D01-C04", storage_allowed=True, is_active=True
    )
    loc2 = StorageLocation.objects.create(
        name="C03", code="S04-L03-D01-C03", storage_allowed=True, is_active=True
    )
    return {"sup": sup, "loc": loc, "loc2": loc2, "admin": admin}


def _brp(env, *, material, retail="10", wholesale="7", replacement="", desc="BELT DRIVE", qty=5):
    brp = BrpCatalogPart.objects.create(
        material_no=material, part_desc=desc,
        retail_price_usd=Decimal(retail), replacement_no_1=replacement,
        wholesale_price_usd=Decimal(wholesale) if wholesale is not None else None,
    )
    part = promote_brp(brp, by=env["admin"])
    _stock(part, env["loc"], qty, env["sup"], env["admin"])
    return part, brp


def _polaris(env, *, number, wholesale="6", retail="20", superseded="", desc="SEAL", qty=5):
    pol = PolarisCatalogPart.objects.create(
        part_number=number, part_name=desc, superseded_number=superseded,
        wholesale_price_usd=Decimal(wholesale) if wholesale is not None else None,
        retail_price_usd=Decimal(retail) if retail is not None else None,
    )
    part = promote_polaris(pol, by=env["admin"])
    _stock(part, env["loc"], qty, env["sup"], env["admin"])
    return part, pol


def _warehouse_only(env, *, number="WH-500", name="РУЧНАЯ ДЕТАЛЬ", qty=5):
    part = PartType.objects.create(
        name=name, category=Category.objects.create(name=f"cat-{number}"),
        unit=Unit.objects.get(name="Штука"),
        tracking_mode=PartType.TrackingMode.BULK, recommended_price=Decimal("100"),
    )
    PartNumber.objects.create(part=part, value=number, kind=PartNumber.Kind.OEM, is_primary=True)
    _stock(part, env["loc"], qty, env["sup"], env["admin"])
    return part


def _sell(env, part, *, qty="1", number="", comment="Иванов", location=None):
    return perform_action(
        part=part, location=location or env["loc"], action_type="sale",
        quantity=qty, customer_comment=comment, scanned_number=number, by=env["admin"],
    )


def _login(client, make_user, *, superuser=True, name="boss"):
    make_user(name, is_superuser=superuser)
    client.login(username=name, password=PASSWORD)


def _sheet(content: bytes):
    return openpyxl.load_workbook(BytesIO(content))[SHEET]


def _b_column(sheet, count=6):
    return [sheet[f"B{DATA_ROW + i}"].value for i in range(count)]


# --- Регрессия production-500: шаблон должен ехать в образ -----------------------------


def test_template_lives_inside_app_package_not_docs():
    """Первопричина 500: шаблон был в docs/, а docs/ исключён из Docker-образа."""
    assert TEMPLATE_PATH.exists(), TEMPLATE_PATH
    parts = TEMPLATE_PATH.parts
    assert "docs" not in parts, "рантайм-ассет не должен лежать в docs/"
    assert "actions" in parts and "customs_template" in parts


def test_template_path_not_excluded_by_dockerignore():
    """Regression: любой путь шаблона, попавший под .dockerignore, ломает export."""
    from django.conf import settings

    root = TEMPLATE_PATH.parents[3]  # .../apps/actions/customs_template/file -> repo root
    ignore = (root / ".dockerignore").read_text(encoding="utf-8").splitlines()
    excluded = {
        line.strip().rstrip("/")
        for line in ignore
        if line.strip() and not line.startswith(("#", "!"))
    }
    relative = TEMPLATE_PATH.relative_to(settings.BASE_DIR).parts
    assert not set(relative) & excluded, (
        f"шаблон {relative} исключён из образа правилами {excluded & set(relative)}"
    )


def test_missing_template_raises_clear_error(env, monkeypatch):
    """Без шаблона — понятная ActionError, а не голый FileNotFoundError."""
    import apps.actions.services as svc

    monkeypatch.setattr(svc, "TEMPLATE_PATH", TEMPLATE_PATH.with_name("нет-файла.xlsx"))
    with pytest.raises(ActionError, match="Шаблон таможенной формы не найден"):
        svc.export_customs_xlsx([])


# --- HTTP-контракт ---------------------------------------------------------------------


def test_export_returns_valid_xlsx(client, make_user, env):
    part, _ = _brp(env, material="219800345")
    _sell(env, part, number="219800345")
    _login(client, make_user)
    resp = client.get(reverse("actions_export"))
    assert resp.status_code == 200
    assert resp["Content-Type"] == XLSX_MIME
    assert ".xlsx" in resp["Content-Disposition"]
    assert resp["Content-Disposition"].isascii()  # безопасное имя файла
    workbook = openpyxl.load_workbook(BytesIO(resp.content))
    assert SHEET in workbook.sheetnames


def test_headers_preserved_in_order(client, make_user, env):
    part, _ = _brp(env, material="219800345")
    _sell(env, part, number="219800345")
    _login(client, make_user)
    sheet = _sheet(client.get(reverse("actions_export")).content)
    headers = [sheet[f"{c}6"].value for c in "ABCDEFGHIJKLM"]
    assert headers[1] and "АРТИКУЛ" in str(headers[1]).upper()
    assert headers[9] and "КОЛИЧЕСТВО" in str(headers[9]).upper()
    assert sheet["I10"].value == "=J10*G10"  # формулы шаблона сохранены
    assert sheet["L10"].value == "=K10*J10"


def test_empty_selection_does_not_500(client, make_user, env):
    _login(client, make_user)
    resp = client.get(reverse("actions_export") + "?part_number=НЕТ-ТАКОГО")
    assert resp.status_code == 200
    sheet = _sheet(resp.content)
    assert sheet[f"B{DATA_ROW}"].value is None  # пример шаблона очищен, данных нет


# --- Exact identity --------------------------------------------------------------------


def test_brp_exports_exact_material_no(client, make_user, env):
    part, _ = _brp(env, material="219800345")
    _sell(env, part, number="219800345")
    _login(client, make_user)
    assert "219800345" in _b_column(_sheet(client.get(reverse("actions_export")).content))


def test_brp_replacement_does_not_replace_number(client, make_user, env):
    # У exact 420931285 нет оптовой; оптовая берётся из replacement 420931284.
    BrpCatalogPart.objects.create(
        material_no="420931284", part_desc="OLD", retail_price_usd=Decimal("4"),
        wholesale_price_usd=Decimal("3"), replacement_no_1="420931285",
    )
    part, _ = _brp(env, material="420931285", retail="0", wholesale="0",
                   replacement="420931284")
    _sell(env, part, number="420931285")
    _login(client, make_user)
    sheet = _sheet(client.get(reverse("actions_export")).content)
    numbers = _b_column(sheet)
    assert "420931285" in numbers  # exact identity
    assert "420931284" not in numbers  # replacement — только источник цены
    assert sheet[f"K{DATA_ROW}"].value == Decimal("3")  # ОПТОВАЯ от источника
    assert sheet[f"K{DATA_ROW}"].value != Decimal("4")  # не розница источника


def test_polaris_exports_exact_part_number(client, make_user, env):
    part, _ = _polaris(env, number="3610075")
    _sell(env, part, number="3610075")
    _login(client, make_user)
    sheet = _sheet(client.get(reverse("actions_export")).content)
    assert "3610075" in _b_column(sheet)
    assert sheet[f"E{DATA_ROW}"].value == "POLARIS"


def test_polaris_superseded_does_not_replace_number(client, make_user, env):
    PolarisCatalogPart.objects.create(
        part_number="1111111", part_name="OLD", retail_price_usd=Decimal("15"),
        wholesale_price_usd=Decimal("8"), superseded_number="2222222",
    )
    part, _ = _polaris(env, number="2222222", retail="0", wholesale="0", superseded="1111111")
    _sell(env, part, number="2222222")
    _login(client, make_user)
    numbers = _b_column(_sheet(client.get(reverse("actions_export")).content))
    assert "2222222" in numbers
    assert "1111111" not in numbers


def test_warehouse_only_uses_snapshot_number(client, make_user, env):
    part = _warehouse_only(env, number="WH-500")
    action = _sell(env, part, number="WH-500")
    assert action.part_number == "WH-500"
    _login(client, make_user)
    assert "WH-500" in _b_column(_sheet(client.get(reverse("actions_export")).content))


# --- Группировка -----------------------------------------------------------------------


def test_same_exact_number_groups_and_sums_quantity(client, make_user, env):
    part, _ = _brp(env, material="219800345", qty=10)
    _sell(env, part, qty="2", number="219800345")
    _sell(env, part, qty="3", number="219800345")
    _login(client, make_user)
    sheet = _sheet(client.get(reverse("actions_export")).content)
    assert _b_column(sheet, 3)[:2] == ["219800345", None]  # одна строка
    assert Decimal(str(sheet[f"J{DATA_ROW}"].value)) == Decimal("5")


def test_different_exact_numbers_of_same_part_not_merged(client, make_user, env):
    part = _warehouse_only(env, number="WH-100", qty=10)
    PartNumber.objects.create(part=part, value="WH-200", kind=PartNumber.Kind.ARTICLE)
    _sell(env, part, number="WH-100")
    _sell(env, part, number="WH-200")
    _login(client, make_user)
    numbers = _b_column(_sheet(client.get(reverse("actions_export")).content))
    assert "WH-100" in numbers and "WH-200" in numbers  # не слиты в одну строку


def test_same_number_brp_and_polaris_not_merged(client, make_user, env):
    brp_part, _ = _brp(env, material="5555555")
    pol_part, _ = _polaris(env, number="5555555")
    _sell(env, brp_part, number="5555555")
    _sell(env, pol_part, number="5555555")
    _login(client, make_user)
    sheet = _sheet(client.get(reverse("actions_export")).content)
    rows = [
        (sheet[f"B{DATA_ROW + i}"].value, sheet[f"E{DATA_ROW + i}"].value) for i in range(2)
    ]
    assert sorted(rows) == [("5555555", "BRP"), ("5555555", "POLARIS")]


# --- Фильтры отчёта --------------------------------------------------------------------


def test_cancelled_excluded_by_default(client, make_user, env):
    part, _ = _brp(env, material="219800345", qty=10)
    bad = _sell(env, part, qty="2", number="219800345")
    _sell(env, part, qty="3", number="219800345")
    cancel_warehouse_action(bad, by=env["admin"], reason="Дубль")
    _login(client, make_user)
    sheet = _sheet(client.get(reverse("actions_export")).content)
    assert Decimal(str(sheet[f"J{DATA_ROW}"].value)) == Decimal("3")  # без отменённых


def test_cancelled_flag_does_not_leak_into_customs(client, make_user, env):
    """Контракт: даже с cancelled=1 в отчёте таможня считает только активные."""
    part, _ = _brp(env, material="219800345", qty=10)
    bad = _sell(env, part, qty="2", number="219800345")
    _sell(env, part, qty="3", number="219800345")
    cancel_warehouse_action(bad, by=env["admin"], reason="Дубль")
    _login(client, make_user)
    sheet = _sheet(client.get(reverse("actions_export") + "?cancelled=1").content)
    assert Decimal(str(sheet[f"J{DATA_ROW}"].value)) == Decimal("3")


def test_date_filter_applied(client, make_user, env):
    part, _ = _brp(env, material="219800345")
    _sell(env, part, number="219800345")
    _login(client, make_user)
    tomorrow = datetime.date.today() + datetime.timedelta(days=1)
    sheet = _sheet(client.get(reverse("actions_export") + f"?date_from={tomorrow}").content)
    assert sheet[f"B{DATA_ROW}"].value is None  # всё отфильтровано


def test_type_filter_applied(client, make_user, env):
    part, _ = _brp(env, material="219800345", qty=10)
    _sell(env, part, number="219800345")
    perform_action(part=part, location=env["loc"], action_type="reserve", quantity="1",
                   customer_comment="Петров", scanned_number="219800345", by=env["admin"])
    _login(client, make_user)
    sheet = _sheet(client.get(reverse("actions_export") + "?action_type=reserve").content)
    assert Decimal(str(sheet[f"J{DATA_ROW}"].value)) == Decimal("1")  # только резерв


def test_part_number_filter_applied(client, make_user, env):
    a, _ = _brp(env, material="219800345")
    b, _ = _brp(env, material="700700700")
    _sell(env, a, number="219800345")
    _sell(env, b, number="700700700")
    _login(client, make_user)
    resp = client.get(reverse("actions_export") + "?part_number=700700700")
    numbers = _b_column(_sheet(resp.content))
    assert "700700700" in numbers and "219800345" not in numbers


def test_location_filter_applied(client, make_user, env):
    part, _ = _brp(env, material="219800345", qty=10)
    _stock(part, env["loc2"], 5, env["sup"], env["admin"])
    _sell(env, part, number="219800345", location=env["loc2"])
    _login(client, make_user)
    sheet = _sheet(client.get(reverse("actions_export") + "?location_code=S04-L03-D01-C03").content)
    assert "219800345" in _b_column(sheet)
    sheet = _sheet(client.get(reverse("actions_export") + "?location_code=НЕТ").content)
    assert sheet[f"B{DATA_ROW}"].value is None


# --- Неполные данные и Decimal ----------------------------------------------------------


def test_missing_price_weights_and_customs_do_not_500(client, make_user, env):
    # нет оптовой цены и весов
    part, _ = _brp(env, material="777000111", retail="0", wholesale="0")
    _sell(env, part, number="777000111")
    assert not PartCustomsInfo.objects.filter(part_type=part).exists()
    _login(client, make_user)
    resp = client.get(reverse("actions_export"))
    assert resp.status_code == 200
    sheet = _sheet(resp.content)
    assert sheet[f"K{DATA_ROW}"].value is None  # цена пустая, не выдумана
    assert sheet[f"G{DATA_ROW}"].value is None and sheet[f"H{DATA_ROW}"].value is None


def test_decimal_quantity_written_as_number(client, make_user, env):
    part, _ = _brp(env, material="219800345", retail="10", wholesale="7", qty=10)
    _sell(env, part, qty="2", number="219800345")
    _login(client, make_user)
    sheet = _sheet(client.get(reverse("actions_export")).content)
    assert Decimal(str(sheet[f"J{DATA_ROW}"].value)) == Decimal("2")
    assert Decimal(str(sheet[f"K{DATA_ROW}"].value)) == Decimal("7")  # оптовая, не 10


# --- Безопасность Excel ------------------------------------------------------------------


def test_excel_safe_text_strips_control_chars_and_neutralises_formulas():
    assert excel_safe_text("BAD\x0bTEXT") == "BADTEXT"
    assert excel_safe_text("=SUM(A1)") == "'=SUM(A1)"
    assert excel_safe_text("+1") == "'+1"
    assert excel_safe_text("-1") == "'-1"
    assert excel_safe_text("@cmd") == "'@cmd"
    assert excel_safe_text("РОЛИК-ШКИВ 420931285") == "РОЛИК-ШКИВ 420931285"  # не искажён
    assert excel_safe_text(None) is None
    assert excel_safe_text("") is None
    assert len(excel_safe_text("я" * 40000)) == 32767


def test_control_character_in_name_does_not_break_workbook(client, make_user, env):
    part, _ = _brp(env, material="219800345", desc="BELT\x0bDRIVE")
    _sell(env, part, number="219800345")
    _login(client, make_user)
    resp = client.get(reverse("actions_export"))
    assert resp.status_code == 200  # раньше — IllegalCharacterError -> 500
    sheet = _sheet(resp.content)
    assert "\x0b" not in str(sheet[f"D{DATA_ROW}"].value)


def test_formula_like_name_is_not_a_formula(client, make_user, env):
    part, _ = _brp(env, material="219800345", desc="=HYPERLINK(1)")
    _sell(env, part, number="219800345")
    _login(client, make_user)
    sheet = _sheet(client.get(reverse("actions_export")).content)
    value = str(sheet[f"D{DATA_ROW}"].value)
    assert value.startswith("'=")  # текст, не формула
    assert sheet[f"D{DATA_ROW}"].data_type != "f"


# --- Read-only -----------------------------------------------------------------------------


def test_export_does_not_mutate_database(client, make_user, env):
    part, _ = _brp(env, material="219800345")
    _sell(env, part, number="219800345")
    before = (
        StockMovement.objects.count(),
        WarehouseAction.objects.count(),
        StockLot.objects.count(),
        PartCustomsInfo.objects.count(),
    )
    _login(client, make_user)
    assert client.get(reverse("actions_export")).status_code == 200
    after = (
        StockMovement.objects.count(),
        WarehouseAction.objects.count(),
        StockLot.objects.count(),
        PartCustomsInfo.objects.count(),
    )
    assert before == after  # GET-экспорт ничего не пишет


def test_historical_snapshot_survives_rate_change(client, make_user, env):
    """Смена курса не переписывает исторические суммы действия."""
    from apps.warehouse.models import ValuationSettings

    part, _ = _brp(env, material="219800345")
    action = _sell(env, part, number="219800345")
    snapshot = (action.part_number, action.unit_price_rub, action.total_price_rub)
    settings_row = ValuationSettings.get()
    settings_row.current_usd_rate = Decimal("200")
    settings_row.save()
    action.refresh_from_db()
    assert (action.part_number, action.unit_price_rub, action.total_price_rub) == snapshot
    _login(client, make_user)
    assert client.get(reverse("actions_export")).status_code == 200


# --- Оформление фактических строк (§3) -------------------------------------------------


def _four_rows(env):
    """Четыре реальные позиции: строки 10..13."""
    for i, material in enumerate(["111000001", "222000002", "333000003", "444000004"]):
        part, _ = _brp(env, material=material, desc=f"PART NUMBER {i}", qty=5)
        _sell(env, part, number=material)


def _data_values(sheet, last_row=60):
    """Значения товарного диапазона (со строки 10): строки 1-9 — шапка шаблона."""
    return [
        str(sheet.cell(r, c).value)
        for r in range(DATA_ROW, last_row)
        for c in range(1, 14)
        if sheet.cell(r, c).value is not None
    ]


def test_all_data_cells_centered_including_last_row(client, make_user, env):
    _four_rows(env)
    _login(client, make_user)
    sheet = _sheet(client.get(reverse("actions_export")).content)
    for offset in range(4):  # последняя (13-я) строка тоже
        row = DATA_ROW + offset
        for column in "ABCDEFGHIJKLM":
            cell = sheet[f"{column}{row}"]
            assert cell.alignment.horizontal == "center", f"{column}{row}"
            assert cell.alignment.vertical == "center", f"{column}{row}"
            assert cell.alignment.wrap_text is True, f"{column}{row}"


def test_last_row_formatted_like_first(client, make_user, env):
    _four_rows(env)
    _login(client, make_user)
    sheet = _sheet(client.get(reverse("actions_export")).content)
    first, last = sheet[f"C{DATA_ROW}"], sheet[f"C{DATA_ROW + 3}"]
    assert (
        first.alignment.horizontal,
        first.alignment.vertical,
        first.alignment.wrap_text,
    ) == (last.alignment.horizontal, last.alignment.vertical, last.alignment.wrap_text)
    # Первопричина «съехавшей» строки: в шаблоне у части строк жёстко задана
    # высота (15/18). Сброс в авто выравнивает все заполненные строки.
    heights = {sheet.row_dimensions[DATA_ROW + i].height for i in range(4)}
    assert heights == {None}


def test_long_name_wraps(client, make_user, env):
    part, _ = _brp(env, material="219800345", desc="VERY LONG DRIVE BELT NAME " * 5)
    _sell(env, part, number="219800345")
    _login(client, make_user)
    sheet = _sheet(client.get(reverse("actions_export")).content)
    assert sheet[f"D{DATA_ROW}"].alignment.wrap_text is True


# --- Страна производства (§4) ----------------------------------------------------------


def test_country_is_latin_canada_for_every_row(client, make_user, env):
    _four_rows(env)
    pol, _ = _polaris(env, number="3610075")
    _sell(env, pol, number="3610075")
    _login(client, make_user)
    sheet = _sheet(client.get(reverse("actions_export")).content)
    for offset in range(5):
        assert sheet[f"F{DATA_ROW + offset}"].value == "CANADA"


def test_no_cyrillic_kanada_anywhere_in_file(client, make_user, env):
    _four_rows(env)
    _login(client, make_user)
    sheet = _sheet(client.get(reverse("actions_export")).content)
    values = [
        str(sheet.cell(r, c).value)
        for r in range(1, 60)
        for c in range(1, 14)
        if sheet.cell(r, c).value is not None
    ]
    assert "КАНАДА" not in values


# --- Оптовая цена прайса (§5) ----------------------------------------------------------


def _price(sheet, row=DATA_ROW):
    """Excel хранит числа как float — сравниваем через Decimal(str(...))."""
    value = sheet[f"K{row}"].value
    return None if value is None else Decimal(str(value))


def test_brp_uses_wholesale_not_retail(client, make_user, env):
    part, _ = _brp(env, material="219800345", retail="35.99", wholesale="28.15")
    _sell(env, part, number="219800345")
    _login(client, make_user)
    sheet = _sheet(client.get(reverse("actions_export")).content)
    assert _price(sheet) == Decimal("28.15")
    assert _price(sheet) != Decimal("35.99")  # не розница


def test_polaris_uses_wholesale_price(client, make_user, env):
    part, _ = _polaris(env, number="3610075", wholesale="6", retail="20")
    _sell(env, part, number="3610075")
    _login(client, make_user)
    sheet = _sheet(client.get(reverse("actions_export")).content)
    assert _price(sheet) == Decimal("6")


def test_polaris_superseded_only_supplies_wholesale(client, make_user, env):
    PolarisCatalogPart.objects.create(
        part_number="1111111", part_name="OLD", retail_price_usd=Decimal("15"),
        wholesale_price_usd=Decimal("9"), superseded_number="2222222",
    )
    part, _ = _polaris(env, number="2222222", retail="0", wholesale="0", superseded="1111111")
    _sell(env, part, number="2222222")
    _login(client, make_user)
    sheet = _sheet(client.get(reverse("actions_export")).content)
    assert "2222222" in _b_column(sheet)  # exact part_number не подменён
    assert "1111111" not in _b_column(sheet)
    assert _price(sheet) == Decimal("9")  # связь дала только цену


def test_wholesale_ignores_rate_and_markup(client, make_user, env):
    from apps.brp.models import BrpPricingSettings
    from apps.warehouse.models import ValuationSettings

    part, _ = _brp(env, material="219800345", retail="35.99", wholesale="28.15")
    _sell(env, part, number="219800345")
    rate = ValuationSettings.get()
    rate.current_usd_rate = Decimal("500")
    rate.save()
    markup = BrpPricingSettings.get()
    markup.brp_markup_percent = Decimal("70")
    markup.save()
    _login(client, make_user)
    sheet = _sheet(client.get(reverse("actions_export")).content)
    assert _price(sheet) == Decimal("28.15")  # чистый USD прайса


def test_missing_wholesale_leaves_cell_empty(client, make_user, env):
    part, _ = _brp(env, material="219800345", retail="35.99", wholesale="0")
    _sell(env, part, number="219800345")
    _login(client, make_user)
    sheet = _sheet(client.get(reverse("actions_export")).content)
    assert _price(sheet) is None  # клиентскую цену не подставляем


# --- Область применения (§6-§7) --------------------------------------------------------


def _compat(part, make_name, vehicle_type_name, model_name="MODEL"):
    from apps.catalog.models import PartCompatibility, VehicleMake, VehicleModel, VehicleType

    vtype, _ = VehicleType.objects.get_or_create(name=vehicle_type_name)
    make, _ = VehicleMake.objects.get_or_create(name=make_name, vehicle_type=vtype)
    model, _ = VehicleModel.objects.get_or_create(vehicle_make=make, name=model_name)
    PartCompatibility.objects.create(part=part, vehicle_model=model)


def _application_for(client, make_user, env, part, number):
    _sell(env, part, number=number)
    _login(client, make_user)
    sheet = _sheet(client.get(reverse("actions_export")).content)
    return sheet[f"M{DATA_ROW}"].value


def test_sea_doo_is_watercraft(client, make_user, env):
    part, _ = _brp(env, material="219800345")
    _compat(part, "Sea-Doo", "Гидроцикл", "GTX")
    assert _application_for(client, make_user, env, part, "219800345") == "ГИДРОЦИКЛ"


def test_ski_doo_and_lynx_are_snowmobile(client, make_user, env):
    part, _ = _brp(env, material="219800345")
    _compat(part, "Ski-Doo", "Снегоход", "SUMMIT")
    _compat(part, "Lynx", "Снегоход", "RAVE")
    assert _application_for(client, make_user, env, part, "219800345") == "СНЕГОХОД"


def test_can_am_is_atv(client, make_user, env):
    part, _ = _brp(env, material="219800345")
    _compat(part, "Can-Am", "Квадроцикл", "OUTLANDER")
    assert _application_for(client, make_user, env, part, "219800345") == "КВАДРОЦИКЛ"


def test_marine_is_boat(client, make_user, env):
    part, _ = _brp(env, material="219800345")
    _compat(part, "Rotax Marine", "Катер", "BOAT")
    assert _application_for(client, make_user, env, part, "219800345") == "КАТЕР / ЛОДКА"


def test_motorcycle_never_becomes_moto_zapchasti(client, make_user, env):
    part, _ = _brp(env, material="219800345")
    _compat(part, "Indian", "Мотоцикл", "SCOUT")  # компания мотоциклы не обслуживает
    assert _application_for(client, make_user, env, part, "219800345") is None


def test_unknown_applicability_stays_empty(client, make_user, env):
    part, _ = _brp(env, material="219800345")  # совместимостей нет вовсе
    assert _application_for(client, make_user, env, part, "219800345") is None


def test_application_not_guessed_from_part_name(client, make_user, env):
    """«OIL SEAL» без применимости не превращается в категорию по названию."""
    part, _ = _brp(env, material="219800345", desc="OIL SEAL")
    assert _application_for(client, make_user, env, part, "219800345") is None


def test_multiple_applications_use_single_rule(client, make_user, env):
    part, _ = _brp(env, material="219800345")
    _compat(part, "Ski-Doo", "Снегоход", "SUMMIT")
    _compat(part, "Can-Am", "Квадроцикл", "OUTLANDER")
    assert _application_for(client, make_user, env, part, "219800345") == "УНИВЕРСАЛЬНЫЕ ЗАПЧАСТИ"


def test_legacy_default_does_not_override_real_applicability(client, make_user, env):
    part, _ = _brp(env, material="219800345")
    _compat(part, "Ski-Doo", "Снегоход", "SUMMIT")
    PartCustomsInfo.objects.create(part_type=part, application_area="МОТО ЗАПЧАСТИ")
    # Легаси-значение модели считается «не заполнено».
    assert _application_for(client, make_user, env, part, "219800345") == "СНЕГОХОД"


def test_manual_application_wins_over_catalog(client, make_user, env):
    part, _ = _brp(env, material="219800345")
    _compat(part, "Ski-Doo", "Снегоход", "SUMMIT")
    PartCustomsInfo.objects.create(part_type=part, application_area="Катер / лодка")
    assert _application_for(client, make_user, env, part, "219800345") == "КАТЕР / ЛОДКА"


def test_no_moto_zapchasti_in_exported_rows(client, make_user, env):
    _four_rows(env)
    _login(client, make_user)
    sheet = _sheet(client.get(reverse("actions_export")).content)
    assert "МОТО ЗАПЧАСТИ" not in _data_values(sheet)


# --- Заготовка шаблона и формулы (§8-§9) -----------------------------------------------


def test_no_leftover_template_values_after_last_row(client, make_user, env):
    _four_rows(env)  # реальные строки 10..13
    _login(client, make_user)
    sheet = _sheet(client.get(reverse("actions_export")).content)
    for row in range(DATA_ROW + 4, DATA_ROW + 20):
        for column in "ABCDEFGHIJKLM":
            value = sheet[f"{column}{row}"].value
            assert value is None, f"{column}{row} = {value!r} — заготовка шаблона"


def test_formulas_reference_their_own_row(client, make_user, env):
    _four_rows(env)
    _login(client, make_user)
    sheet = _sheet(client.get(reverse("actions_export")).content)
    for offset in range(4):
        row = DATA_ROW + offset
        assert sheet[f"I{row}"].value == f"=J{row}*G{row}"
        assert sheet[f"L{row}"].value == f"=K{row}*J{row}"
    last = DATA_ROW + 3
    assert sheet[f"I{last}"].value == f"=J{last}*G{last}"  # не ссылка на соседнюю строку


def test_template_header_and_hint_rows_preserved(client, make_user, env):
    """Строки 1-9 — шапка и подсказки брокера, их структуру не трогаем."""
    _four_rows(env)
    _login(client, make_user)
    sheet = _sheet(client.get(reverse("actions_export")).content)
    assert sheet["F6"].value == "СТРАНА ПРОИЗВОДСТВА"
    assert sheet["B9"].value == "БОЛЬШИМИ БУКВАМИ"
    assert sheet["M7"].value == "МОТО ЗАПЧАСТИ"  # подсказка шаблона, не выгруженная строка
