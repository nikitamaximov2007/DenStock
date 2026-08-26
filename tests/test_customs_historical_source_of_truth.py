"""Таможенные данные как исторический источник истины.

Правило, которое здесь закрепляется целиком:

  пользователь вводит таможенные данные -> база сохраняет введённое ТОЧНО ->
  историческое списание ссылается на ту версию, что действовала в его момент ->
  предпросмотр показывает именно её -> Excel выгружает именно её.

Ни каталог поставщика, ни цена прайса, ни производитель по умолчанию не имеют
права подставиться вместо ввода пользователя. Незаполненное поле остаётся
пустым: выдуманный ноль в таможенной декларации хуже пустой клетки.

Источник строк - складской журнал (StockMovement), а не журнал действий:
декларировать нужно всё, что физически покинуло склад, включая списания и
выдачи в ремонт, проведённые не сканером.
"""
import datetime
from decimal import Decimal
from io import BytesIO

import openpyxl
import pytest
from django.contrib.auth.models import Group
from django.urls import reverse
from django.utils import timezone

from apps.accounts import roles
from apps.actions.models import PartCustomsDataVersion, PartCustomsInfo
from apps.actions.services import (
    customs_data_version_for,
    export_customs_xlsx,
    historical_customs_rows,
    parse_customs_usd,
    perform_action,
)
from apps.catalog.models import Category, PartNumber, PartType, Unit
from apps.inventory.models import StockLot, StockMovement
from apps.inventory.services import (
    create_stock_lot,
    receive_stock_lot,
    return_stock_lot_quantity,
    write_off_stock_lot_quantity,
)
from apps.procurement.models import Batch, BatchLine
from apps.procurement.services import finalize_cost
from apps.suppliers.models import Supplier
from apps.warehouse.models import StorageLocation

PASSWORD = "parol-12345"
SHEET = "Лист1"
DATA_ROW = 10
TOTALS_ROW = 150
ApplicationArea = PartCustomsInfo.ApplicationArea


# --- Обстановка ------------------------------------------------------------


@pytest.fixture
def make_user(db, django_user_model):
    def _make(username, *, role=None, is_superuser=False):
        if is_superuser:
            return django_user_model.objects.create_superuser(username=username,
                                                              password=PASSWORD)
        user = django_user_model.objects.create_user(username=username, password=PASSWORD)
        if role:
            user.groups.add(Group.objects.get(name=role))
        return user

    return _make


@pytest.fixture
def env(db, make_user):
    admin = make_user("boss", is_superuser=True)
    supplier, _ = Supplier.objects.get_or_create(name="ООО Поставка")
    location, _ = StorageLocation.objects.get_or_create(
        code="S01-D01-C01",
        defaults={"name": "Ячейка", "storage_allowed": True, "is_active": True},
    )
    other, _ = StorageLocation.objects.get_or_create(
        code="S02-D02-C02",
        defaults={"name": "Вторая", "storage_allowed": True, "is_active": True},
    )
    category, _ = Category.objects.get_or_create(name="Двигатель", parent=None)
    return {
        "admin": admin, "sup": supplier, "loc": location,
        "loc2": other, "cat": category,
    }


def _part(env, *, number, name="ДЕТАЛЬ"):
    part = PartType.objects.create(
        name=name, category=env["cat"], unit=Unit.objects.get(name="Штука"),
        tracking_mode=PartType.TrackingMode.BULK, recommended_price=Decimal("1000"),
    )
    PartNumber.objects.create(
        part=part, value=number, kind=PartNumber.Kind.OEM, is_primary=True
    )
    return part


def _receive(env, part, quantity="10", unit_cost="100", location=None):
    batch = Batch.objects.create(supplier=env["sup"], shipping_cost=Decimal("0"))
    line = BatchLine.objects.create(
        batch=batch, part_type=part,
        quantity=Decimal(quantity), unit_cost_currency=Decimal(unit_cost),
    )
    batch.status = Batch.Status.ACCEPTED
    batch.save(update_fields=["status"])
    finalize_cost(batch, env["admin"])
    line.refresh_from_db()
    lot = create_stock_lot(line, location or env["loc"], Decimal(quantity))
    receive_stock_lot(lot, by=env["admin"])
    return lot


def _card(part, **overrides):
    values = {
        "customs_name_ru": "РЕМЕНЬ ПРИВОДНОЙ",
        "customs_name_en": "BELT DRIVE",
        "manufacturer": "BRP",
        "country_of_origin": "CANADA",
        "gross_weight_kg": Decimal("0.350"),
        "net_weight_kg": Decimal("0.300"),
        "customs_unit_price_usd": Decimal("12.50"),
        "application_area": ApplicationArea.SNOWMOBILE,
    }
    values.update(overrides)
    card = PartCustomsInfo.objects.filter(part_type=part).first()
    if card is None:
        return PartCustomsInfo.objects.create(part_type=part, **values)
    return _edit(card, **values)  # уже открытую карточку правим, как это делает форма


def _edit(card, **changes):
    """Правка карточки тем же путём, каким её правит пользователь."""
    for field, value in changes.items():
        setattr(card, field, value)
    card.save()
    card.refresh_from_db()
    return card


def _sell(env, part, *, quantity="1", number=""):
    return perform_action(
        part=part, location=env["loc"], action_type="sale", quantity=quantity,
        customer_comment="Иванов", scanned_number=number, by=env["admin"],
    )


def _row_for(rows, number, version=None):
    found = [r for r in rows if r["number"] == number
             and (version is None or r["version_number"] == version)]
    assert len(found) == 1, [(r["number"], r["version_number"]) for r in rows]
    return found[0]


def _login(client, make_user, *, role=None, name="boss"):
    if name != "boss":
        make_user(name, role=role)
    client.login(username=name, password=PASSWORD)


# --- 1-5. Ввод сохраняется точно и версионируется --------------------------


def test_entered_values_persist_exactly(env):
    part = _part(env, number="219800345")
    card = _card(
        part, country_of_origin="AUSTRIA", customs_name_en="DRIVE BELT",
        customs_unit_price_usd=Decimal("34.07"), source_reference="Инвойс 12/44",
    )
    card.refresh_from_db()
    assert card.country_of_origin == "AUSTRIA"
    assert card.customs_name_en == "DRIVE BELT"
    assert card.customs_unit_price_usd == Decimal("34.07")  # не 34.069999
    assert card.source_reference == "Инвойс 12/44"


def test_saving_card_records_first_version(env):
    part = _part(env, number="219800345")
    _card(part, country_of_origin="AUSTRIA")
    version = PartCustomsDataVersion.objects.get(part_type=part)
    assert version.version == 1
    assert version.country_of_origin == "AUSTRIA"


def test_unchanged_save_does_not_create_a_version(env):
    part = _part(env, number="219800345")
    card = _card(part)
    card.save()
    card.save()
    assert PartCustomsDataVersion.objects.filter(part_type=part).count() == 1


def test_opening_the_form_does_not_record_a_version(client, env, make_user):
    """Открытая, но не заполненная форма истории не создаёт.

    Страница правки заводит карточку со значениями по умолчанию уже на GET.
    Если считать её версией, она станет самой ранней и перехватит все прошлые
    списания - декларация уйдёт пустой при заполненной карточке.
    """
    part = _part(env, number="219800345")
    _receive(env, part, quantity="10")
    _sell(env, part, quantity="2", number="219800345")
    _login(client, make_user)
    client.get(reverse("actions_customs_edit", args=[part.pk]))
    assert PartCustomsInfo.objects.filter(part_type=part).exists()  # карточка есть
    assert PartCustomsDataVersion.objects.filter(part_type=part).count() == 0

    _card(part, customs_unit_price_usd=Decimal("12.50"))
    versions = PartCustomsDataVersion.objects.filter(part_type=part)
    assert [v.version for v in versions] == [1]  # первая версия - настоящий ввод
    row = _row_for(historical_customs_rows(), "219800345")
    assert row["version_number"] == 1
    assert row["usd_price"] == Decimal("12.50")  # история не перехвачена пустышкой


def test_default_manufacturer_alone_is_not_an_entry(env):
    """Одно лишь «BRP» из умолчания модели заявлением не считается."""
    part = _part(env, number="219800345")
    PartCustomsInfo.objects.create(part_type=part)  # ровно умолчания
    assert PartCustomsDataVersion.objects.filter(part_type=part).count() == 0
    card = PartCustomsInfo.objects.get(part_type=part)
    assert card.manufacturer == "BRP"  # умолчание на месте
    _edit(card, country_of_origin="CANADA")  # первый настоящий факт
    assert PartCustomsDataVersion.objects.filter(part_type=part).count() == 1


def test_real_correction_creates_next_version(env):
    part = _part(env, number="219800345")
    card = _card(part, customs_unit_price_usd=Decimal("10"))
    _edit(card, customs_unit_price_usd=Decimal("11"))
    versions = list(PartCustomsDataVersion.objects.filter(part_type=part).order_by("version"))
    assert [v.version for v in versions] == [1, 2]
    assert [v.customs_unit_price_usd for v in versions] == [Decimal("10"), Decimal("11")]


def test_earlier_version_is_immutable(env):
    part = _part(env, number="219800345")
    card = _card(part, country_of_origin="CANADA")
    first = PartCustomsDataVersion.objects.get(part_type=part, version=1)
    _edit(card, country_of_origin="AUSTRIA")
    first.refresh_from_db()
    assert first.country_of_origin == "CANADA"  # прошлое не переписано


# --- 6-8. Историческая привязка списаний -----------------------------------


def test_first_version_covers_movements_made_before_it(env):
    """Карточку заводят ПОСЛЕ того, как деталь начала расходоваться."""
    part = _part(env, number="219800345")
    _receive(env, part)
    _sell(env, part, quantity="2", number="219800345")
    _card(part, customs_unit_price_usd=Decimal("12.50"))  # заведена уже потом
    row = _row_for(historical_customs_rows(), "219800345")
    assert row["version_number"] == 1
    assert row["usd_price"] == Decimal("12.50")
    assert row["quantity"] == Decimal("2")


def test_later_version_does_not_rewrite_an_earlier_write_off(env):
    part = _part(env, number="219800345")
    _receive(env, part)
    card = _card(part, customs_unit_price_usd=Decimal("10"))
    _sell(env, part, quantity="2", number="219800345")
    _edit(card, customs_unit_price_usd=Decimal("99"))
    row = _row_for(historical_customs_rows(), "219800345")
    assert row["version_number"] == 1
    assert row["usd_price"] == Decimal("10")  # не 99


def test_write_offs_under_two_versions_stay_separate_rows(env):
    part = _part(env, number="219800345")
    _receive(env, part, quantity="20")
    card = _card(part, customs_unit_price_usd=Decimal("10"))
    _sell(env, part, quantity="2", number="219800345")
    _edit(card, customs_unit_price_usd=Decimal("20"))
    _sell(env, part, quantity="3", number="219800345")

    rows = [r for r in historical_customs_rows() if r["number"] == "219800345"]
    assert len(rows) == 2
    by_version = {r["version_number"]: r for r in rows}
    assert by_version[1]["quantity"] == Decimal("2")
    assert by_version[1]["usd_price"] == Decimal("10")
    assert by_version[2]["quantity"] == Decimal("3")
    assert by_version[2]["usd_price"] == Decimal("20")


def test_version_lookup_returns_the_one_in_force(env):
    part = _part(env, number="219800345")
    card = _card(part, customs_unit_price_usd=Decimal("10"))
    first_saved = timezone.now()
    _edit(card, customs_unit_price_usd=Decimal("20"))
    assert customs_data_version_for(part, first_saved).version == 1
    assert customs_data_version_for(part, timezone.now()).version == 2
    long_ago = first_saved - datetime.timedelta(days=365)
    assert customs_data_version_for(part, long_ago).version == 1  # история покрыта


# --- 9-12. Родословная возвратов -------------------------------------------


def test_return_reduces_the_version_the_goods_left_under(env):
    """Возврат гасит свою выдачу, а не ту версию, что действует в день возврата."""
    part = _part(env, number="219800345")
    lot = _receive(env, part, quantity="20")
    card = _card(part, customs_unit_price_usd=Decimal("10"))
    action = _sell(env, part, quantity="5", number="219800345")
    _edit(card, customs_unit_price_usd=Decimal("20"))  # правка ПОСЛЕ продажи
    return_stock_lot_quantity(
        lot.batch_line, env["loc"], Decimal("2"),
        unit_cost_rub=lot.landed_unit_cost_rub,
        restock_status=StockLot.Status.AVAILABLE, by=env["admin"],
        document_id=action.sale_id, comment="Возврат",
    )
    rows = [r for r in historical_customs_rows() if r["number"] == "219800345"]
    assert len(rows) == 1  # фантомной строки по второй версии не появилось
    assert rows[0]["version_number"] == 1
    assert rows[0]["quantity"] == Decimal("3")  # 5 - 2, по своей версии


def test_partial_return_leaves_the_remainder(env):
    part = _part(env, number="219800345")
    lot = _receive(env, part, quantity="20")
    _card(part)
    action = _sell(env, part, quantity="5", number="219800345")
    return_stock_lot_quantity(
        lot.batch_line, env["loc"], Decimal("2"),
        unit_cost_rub=lot.landed_unit_cost_rub,
        restock_status=StockLot.Status.AVAILABLE, by=env["admin"],
        document_id=action.sale_id, comment="Возврат",
    )
    assert _row_for(historical_customs_rows(), "219800345")["quantity"] == Decimal("3")


def test_full_return_removes_the_row(env):
    part = _part(env, number="219800345")
    lot = _receive(env, part, quantity="20")
    _card(part)
    action = _sell(env, part, quantity="4", number="219800345")
    return_stock_lot_quantity(
        lot.batch_line, env["loc"], Decimal("4"),
        unit_cost_rub=lot.landed_unit_cost_rub,
        restock_status=StockLot.Status.AVAILABLE, by=env["admin"],
        document_id=action.sale_id, comment="Возврат",
    )
    assert [r for r in historical_customs_rows() if r["number"] == "219800345"] == []


def test_return_does_not_reduce_another_part(env):
    kept = _part(env, number="111000111", name="ПЕРВАЯ")
    given_back = _part(env, number="222000222", name="ВТОРАЯ")
    _card(kept)
    _card(given_back)
    _receive(env, kept, quantity="10")
    lot = _receive(env, given_back, quantity="10")
    _sell(env, kept, quantity="3", number="111000111")
    action = _sell(env, given_back, quantity="3", number="222000222")
    return_stock_lot_quantity(
        lot.batch_line, env["loc"], Decimal("3"),
        unit_cost_rub=lot.landed_unit_cost_rub,
        restock_status=StockLot.Status.AVAILABLE, by=env["admin"],
        document_id=action.sale_id, comment="Возврат",
    )
    rows = historical_customs_rows()
    assert _row_for(rows, "111000111")["quantity"] == Decimal("3")  # чужой расход цел
    assert [r for r in rows if r["number"] == "222000222"] == []


def test_cancelled_sale_leaves_no_customs_consumption(env):
    from apps.actions.services import cancel_warehouse_action

    part = _part(env, number="219800345")
    _receive(env, part, quantity="20")
    _card(part)
    bad = _sell(env, part, quantity="2", number="219800345")
    _sell(env, part, quantity="3", number="219800345")
    cancel_warehouse_action(bad, by=env["admin"], reason="Дубль")
    assert _row_for(historical_customs_rows(), "219800345")["quantity"] == Decimal("3")


# --- 13-17. Канонический журнал: что считается выбытием ---------------------


def test_receipt_is_not_customs_consumption(env):
    part = _part(env, number="219800345")
    _card(part)
    _receive(env, part, quantity="10")
    assert historical_customs_rows() == []


def test_relocation_is_not_customs_consumption(env):
    from apps.inventory.services import move_stock_lot

    part = _part(env, number="219800345")
    _card(part)
    lot = _receive(env, part, quantity="10")
    move_stock_lot(lot, env["loc2"], by=env["admin"])
    assert historical_customs_rows() == []


def test_stocktaking_adjustment_is_not_customs_consumption(env):
    part = _part(env, number="219800345")
    _card(part)
    lot = _receive(env, part, quantity="10")
    StockMovement.objects.create(
        movement_type=StockMovement.MovementType.ADJUST_OUT,
        part_type=part, stock_lot=lot, quantity=Decimal("3"),
        from_location=env["loc"], created_by=env["admin"],
    )
    assert historical_customs_rows() == []


def test_write_off_is_customs_consumption(env):
    part = _part(env, number="219800345")
    _card(part)
    lot = _receive(env, part, quantity="10")
    write_off_stock_lot_quantity(lot, Decimal("2"), by=env["admin"], comment="Брак")
    assert _row_for(historical_customs_rows(), "219800345")["quantity"] == Decimal("2")


def test_repair_issue_is_customs_consumption(env):
    part = _part(env, number="219800345")
    _receive(env, part, quantity="10")
    _card(part)
    perform_action(
        part=part, location=env["loc"], action_type="repair", quantity="2",
        customer_comment="Сидоров", scanned_number="219800345", by=env["admin"],
    )
    assert _row_for(historical_customs_rows(), "219800345")["quantity"] == Decimal("2")


# --- 18-21. Никаких подстановок --------------------------------------------


def test_catalog_wholesale_price_never_substitutes(env):
    from apps.brp.models import BrpCatalogPart
    from apps.brp.services import promote_to_warehouse

    brp = BrpCatalogPart.objects.create(
        material_no="219800345", part_desc="BELT DRIVE",
        retail_price_usd=Decimal("35.99"), wholesale_price_usd=Decimal("28.15"),
    )
    part = promote_to_warehouse(brp, by=env["admin"])
    _receive(env, part, quantity="10")
    _card(part, customs_unit_price_usd=None)
    _sell(env, part, quantity="1", number="219800345")
    row = _row_for(historical_customs_rows(), "219800345")
    assert row["usd_price"] is None  # ни оптовой, ни розничной, ни нуля
    assert "нет таможенной цены в USD" in row["warnings"]


def test_country_is_not_hardcoded_to_canada(env):
    part = _part(env, number="219800345")
    _receive(env, part, quantity="10")
    _card(part, country_of_origin="")
    _sell(env, part, quantity="1", number="219800345")
    row = _row_for(historical_customs_rows(), "219800345")
    assert row["country"] == ""  # прежний хардкод «CANADA» исчез
    assert "не заполнена страна производства" in row["warnings"]

    other = _part(env, number="700100700", name="ВТОРАЯ")
    _receive(env, other, quantity="10")
    _card(other, country_of_origin="AUSTRIA")
    _sell(env, other, quantity="1", number="700100700")
    assert _row_for(historical_customs_rows(), "700100700")["country"] == "AUSTRIA"


def test_catalog_name_never_substitutes(env):
    part = _part(env, number="219800345", name="НАЗВАНИЕ ИЗ КАТАЛОГА")
    _receive(env, part, quantity="10")
    _card(part, customs_name_ru="", customs_name_en="")
    _sell(env, part, quantity="1", number="219800345")
    row = _row_for(historical_customs_rows(), "219800345")
    assert row["name_ru"] == "" and row["name_en"] == ""
    assert "НАЗВАНИЕ ИЗ КАТАЛОГА" not in (row["name_ru"], row["name_en"])


def test_customs_price_input_rejects_nonsense():
    assert parse_customs_usd("") is None
    assert parse_customs_usd("12,50") == Decimal("12.50")  # запятая как разделитель
    with pytest.raises(ValueError, match="больше нуля"):
        parse_customs_usd("0")
    with pytest.raises(ValueError, match="больше нуля"):
        parse_customs_usd("-5")
    with pytest.raises(ValueError, match="числом"):
        parse_customs_usd("абв")
    with pytest.raises(ValueError, match="2 знак"):
        parse_customs_usd("12.505")


# --- 22-26. Excel за пределами 140 строк -----------------------------------


def _rows(count):
    """Готовые строки экспорта без обращения к базе: проверяется сам писатель."""
    return [
        {
            "number": f"{700000 + index}", "name_ru": "ДЕТАЛЬ", "name_en": "PART",
            "manufacturer": "BRP", "country": "CANADA",
            "gross_weight_kg": Decimal("0.350"), "net_weight_kg": Decimal("0.300"),
            "usd_price": Decimal("12.50"), "quantity": Decimal("2"),
            "application_area": "СНЕГОХОД",
        }
        for index in range(count)
    ]


@pytest.mark.parametrize("count", [141, 200, 500])
def test_all_rows_are_written_beyond_the_template_limit(count):
    sheet = openpyxl.load_workbook(export_customs_xlsx(rows=_rows(count)))[SHEET]
    last = DATA_ROW + count - 1
    assert str(sheet[f"B{DATA_ROW}"].value) == "700000"
    assert str(sheet[f"B{last}"].value) == str(700000 + count - 1)  # последняя на месте
    assert sheet[f"B{last + 1}"].value is None  # и ничего лишнего за ней


@pytest.mark.parametrize("count", [141, 200, 500])
def test_totals_row_survives_and_sums_every_row(count):
    sheet = openpyxl.load_workbook(export_customs_xlsx(rows=_rows(count)))[SHEET]
    last = DATA_ROW + count - 1
    assert sheet[f"I{last + 1}"].value == f"=SUM(I7:I{last})"
    assert "вес" in str(sheet[f"F{last + 1}"].value).lower()  # подпись итога уехала вниз


@pytest.mark.parametrize("count", [141, 200, 500])
def test_totals_merge_moves_off_the_data_rows(count):
    """Объединение подписи итога не должно склеить колонки строки данных."""
    sheet = openpyxl.load_workbook(export_customs_xlsx(rows=_rows(count)))[SHEET]
    last = DATA_ROW + count - 1
    merged = [m for m in sheet.merged_cells.ranges if m.min_row >= DATA_ROW]
    assert [str(m) for m in merged] == [f"F{last + 1}:H{last + 1}"]
    assert Decimal(str(sheet[f"H{last}"].value)) == Decimal("0.3")  # нетто не проглочен


@pytest.mark.parametrize("count", [141, 200])
def test_formulas_reference_their_own_row_beyond_the_limit(count):
    sheet = openpyxl.load_workbook(export_customs_xlsx(rows=_rows(count)))[SHEET]
    for row in (DATA_ROW, TOTALS_ROW - 1, TOTALS_ROW, DATA_ROW + count - 1):
        assert sheet[f"I{row}"].value == f"=J{row}*G{row}"
        assert sheet[f"L{row}"].value == f"=K{row}*J{row}"


def test_empty_values_stay_empty_and_never_become_zero():
    rows = _rows(141)
    rows[140] = {**rows[140], "gross_weight_kg": None, "net_weight_kg": None,
                 "usd_price": None, "application_area": ""}
    sheet = openpyxl.load_workbook(export_customs_xlsx(rows=rows))[SHEET]
    last = DATA_ROW + 140
    assert sheet[f"G{last}"].value is None
    assert sheet[f"H{last}"].value is None
    assert sheet[f"K{last}"].value is None
    assert sheet[f"M{last}"].value is None


# --- 27-30. Экран и права --------------------------------------------------


def test_preview_shows_the_saved_version_not_the_current_card(client, env, make_user):
    part = _part(env, number="219800345")
    _receive(env, part, quantity="20")
    card = _card(part, customs_unit_price_usd=Decimal("10"), country_of_origin="CANADA")
    _sell(env, part, quantity="2", number="219800345")
    _edit(card, customs_unit_price_usd=Decimal("99"), country_of_origin="AUSTRIA")
    _login(client, make_user)
    html = client.get(reverse("actions_report")).content.decode()
    assert "CANADA" in html  # то, что действовало в момент списания
    assert "AUSTRIA" not in html
    assert "99" not in html.split("Таможенные данные")[-1][:4000]


def test_excel_carries_the_same_version_the_preview_showed(client, env, make_user):
    part = _part(env, number="219800345")
    _receive(env, part, quantity="20")
    card = _card(part, customs_unit_price_usd=Decimal("10"), country_of_origin="CANADA")
    _sell(env, part, quantity="2", number="219800345")
    _edit(card, customs_unit_price_usd=Decimal("99"), country_of_origin="AUSTRIA")
    _login(client, make_user)
    sheet = openpyxl.load_workbook(
        BytesIO(client.get(reverse("actions_export")).content)
    )[SHEET]
    assert sheet[f"F{DATA_ROW}"].value == "CANADA"
    assert Decimal(str(sheet[f"K{DATA_ROW}"].value)) == Decimal("10")


def test_export_refuses_when_nothing_was_ever_entered(client, env, make_user):
    part = _part(env, number="219800345")
    _receive(env, part, quantity="10")
    _sell(env, part, quantity="1", number="219800345")
    _login(client, make_user)
    resp = client.get(reverse("actions_export"))
    assert resp.status_code == 302
    assert "не заведены таможенные данные" in client.get(resp.url).content.decode()


def test_viewer_cannot_export_customs_xlsx(client, env, make_user):
    part = _part(env, number="219800345")
    _receive(env, part, quantity="10")
    _card(part)
    _sell(env, part, quantity="1", number="219800345")
    _login(client, make_user, role=roles.VIEWER, name="viewer")
    assert client.get(reverse("actions_export")).status_code == 403
    assert client.get(reverse("actions_report")).status_code == 403


def test_unauthenticated_cannot_export_customs_xlsx(client, env):
    resp = client.get(reverse("actions_export"))
    assert resp.status_code == 302 and "login" in resp.url


def test_export_does_not_write_to_the_database(client, env, make_user):
    part = _part(env, number="219800345")
    _receive(env, part, quantity="10")
    _card(part)
    _sell(env, part, quantity="1", number="219800345")
    before = (
        StockMovement.objects.count(),
        PartCustomsDataVersion.objects.count(),
        PartCustomsInfo.objects.count(),
    )
    _login(client, make_user)
    client.get(reverse("actions_report"))
    client.get(reverse("actions_export"))
    assert (
        StockMovement.objects.count(),
        PartCustomsDataVersion.objects.count(),
        PartCustomsInfo.objects.count(),
    ) == before
