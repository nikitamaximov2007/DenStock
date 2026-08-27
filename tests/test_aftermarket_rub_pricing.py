"""Дилерская цена аналога становится рублёвой ценой клиента.

Дилерская цена поставщика в каталоге аналогов - это та же оптовая цена в
долларах, что и у BRP. Значит и цена клиента считается той же единственной
формулой проекта, а не второй её копией: одна наценка магазина, одно
округление, один источник истины.

Здесь это закреплено не пересчётом формулы в тесте (иначе обе реализации
могли бы ошибаться одинаково), а прямым сравнением двух путей с каноническим
`apps.brp.pricing.customer_price_rub`.
"""
from decimal import Decimal
from io import BytesIO

import pytest
from django.core.files.uploadedfile import SimpleUploadedFile
from django.urls import reverse
from openpyxl import Workbook

from apps.brp.models import BrpCatalogPart, BrpPricingSettings
from apps.brp.pricing import customer_price_rub
from apps.brp.services import promote_to_warehouse
from apps.catalog.models import PartAnalog, PartType
from apps.catalog.services import (
    get_current_price_settings,
    plan_linked_part_price_refresh,
    refresh_linked_part_prices,
)
from apps.catalog_import.models import AftermarketCatalogPart, CatalogImportBatch
from apps.inventory.models import PartItem, StockBalance, StockLot, StockMovement
from apps.warehouse.models import ValuationSettings

HEADERS = ["Manufacturer", "Item SKU", "Manufacturer Number", "Description", "Dlr Cost"]
PASSWORD = "parol-12345"

# Разные производители и разные порядки цены: копеечная мелочь, обычная
# деталь, дробный цент и дорогая позиция.
REAL_SHAPED_ROWS = [
    ["WOODYS", "AA6-9750", "AA6-9750", "ACE 6 X 60 TURNING", "107.6"],
    ["ALL BALLS RACING INC", "AB141001", "14-1001", "CHAIN CASE BEARING KIT", "32.74"],
    ["NGK", "AB6", "2910", "NICKEL SPARK PLUG", "6.30"],
    ["PROX", "PX010", "01.1395.100", "PISTON KIT", "89.08"],
    ["WISECO", "WS4820", "4820M09400", "FORGED PISTON", "1274.99"],
]


@pytest.fixture
def rates(db):
    valuation = ValuationSettings.get()
    valuation.current_usd_rate = Decimal("105")
    valuation.save()
    markup = BrpPricingSettings.get()
    markup.brp_markup_percent = Decimal("40")
    markup.save()
    return valuation, markup


@pytest.fixture
def boss(db, django_user_model, client, rates):
    django_user_model.objects.create_superuser(username="boss", password=PASSWORD)
    client.login(username="boss", password=PASSWORD)
    return client


def _book(rows, sheet_name="diorlight priceupdate"):
    book = Workbook()
    book.remove(book.active)
    sheet = book.create_sheet(sheet_name)
    sheet.append(HEADERS)
    sheet.append([""] * len(HEADERS))
    for row in rows:
        sheet.append(row)
    output = BytesIO()
    book.save(output)
    return output.getvalue()


def _import(client, rows):
    """Загрузить и применить книгу через настоящий экран импорта."""
    client.post(
        reverse("catalog_import_upload"),
        {"catalog": "analogs", "workbook": SimpleUploadedFile("dealer.xlsx", _book(rows))},
        follow=True,
    )
    batch = CatalogImportBatch.objects.order_by("-pk").first()
    client.post(reverse("catalog_import_apply", args=[batch.pk]), follow=True)
    batch.refresh_from_db()
    assert batch.status == CatalogImportBatch.Status.APPLIED, batch.error_text
    return batch


def _canonical(usd) -> Decimal:
    """Цена клиента канонической службой - без повторения арифметики в тесте."""
    settings = get_current_price_settings(create=False)
    return customer_price_rub(usd, settings.current_usd_rate, settings.brp_markup_percent)


def _stock():
    return (
        StockLot.objects.count(), StockMovement.objects.count(),
        StockBalance.objects.count(), PartItem.objects.count(),
    )


# --- Тождество формулы: главный тест ----------------------------------------


def test_aftermarket_and_brp_agree_on_the_same_wholesale_input(boss, rates):
    """Одна и та же оптовая цена даёт одну и ту же цену клиента обоими путями."""
    wholesale = Decimal("89.08")

    brp = BrpCatalogPart.objects.create(
        material_no="TEST-89080", material_no_norm="TEST89080",
        part_desc="SAME WHOLESALE", wholesale_price_usd=wholesale,
        retail_price_usd=Decimal("199.99"), is_current=True,
    )
    brp_part = promote_to_warehouse(brp, by=None)

    _import(boss, [["PROX", "SKU-89080", "PX-89080", "SAME WHOLESALE", str(wholesale)]])
    analog = AftermarketCatalogPart.objects.get(manufacturer_number="PX-89080").part

    assert brp_part.recommended_price is not None
    assert analog.recommended_price == brp_part.recommended_price
    assert analog.recommended_price == _canonical(wholesale)


@pytest.mark.parametrize(
    "wholesale",
    ["0.01", "6.30", "32.74", "89.08", "107.6", "107.605", "1274.99", "99999.99"],
)
def test_rounding_matches_the_canonical_service_exactly(boss, rates, wholesale):
    """Округление - каноническое, включая дробный цент и крупные суммы.

    Считается от СОХРАНЁННОЙ дилерской цены, а не от строки в книге: каталог
    хранит цену поставщика с точностью до цента, и пересчёт по базе обязан
    давать ровно то же число, что и импорт.
    """
    _import(boss, [["PROX", f"SKU-{wholesale}", f"NUM-{wholesale}", "ROUNDING", wholesale]])
    entry = AftermarketCatalogPart.objects.get(manufacturer_number=f"NUM-{wholesale}")
    assert entry.part.recommended_price == _canonical(entry.dealer_cost_usd)

    # Пересчёт по базе повторяет результат импорта до копейки.
    imported = entry.part.recommended_price
    PartType.objects.filter(pk=entry.part_id).update(recommended_price=None)
    refresh_linked_part_prices(
        usd_rate=Decimal("105"), brp_markup=Decimal("40"), polaris_markup=Decimal("40"),
        catalogs=frozenset({"aftermarket"}),
    )
    entry.part.refresh_from_db()
    assert entry.part.recommended_price == imported


def test_the_shop_markup_is_not_duplicated_for_analogs(boss, rates):
    """Смена наценки двигает обе цены одинаково: таблицы наценок у аналогов нет."""
    _import(boss, [["PROX", "SKU-M", "NUM-M", "MARKUP", "100.00"]])
    part = AftermarketCatalogPart.objects.get(manufacturer_number="NUM-M").part
    assert part.recommended_price == _canonical(Decimal("100.00"))

    markup = BrpPricingSettings.get()
    markup.brp_markup_percent = Decimal("55")
    markup.save()
    refresh_linked_part_prices(
        usd_rate=Decimal("105"), brp_markup=Decimal("55"), polaris_markup=Decimal("40"),
        catalogs=frozenset({"aftermarket"}),
    )
    part.refresh_from_db()
    assert part.recommended_price == _canonical(Decimal("100.00"))


# --- Отсутствующая цена -----------------------------------------------------


@pytest.mark.parametrize("dealer_cost", ["", "0", "не число", "-5"])
def test_missing_or_impossible_dealer_cost_invents_no_price(boss, rates, dealer_cost):
    _import(boss, [
        ["PROX", "SKU-OK", "NUM-OK", "С ЦЕНОЙ", "10.00"],
        ["PROX", "SKU-BAD", "NUM-BAD", "БЕЗ ЦЕНЫ", dealer_cost],
    ])
    good = AftermarketCatalogPart.objects.get(manufacturer_number="NUM-OK")
    assert good.part.recommended_price == _canonical(Decimal("10.00"))
    bad = AftermarketCatalogPart.objects.filter(manufacturer_number="NUM-BAD").first()
    if bad is not None:  # строка могла быть отброшена разбором - это тоже верно
        assert bad.part.recommended_price is None


def test_msrp_is_never_used_as_the_price_source(boss, rates):
    """Даже когда MSRP в книге есть, цена считается из дилерской."""
    book = Workbook()
    book.remove(book.active)
    sheet = book.create_sheet("priceupdate")
    sheet.append(["Manufacturer", "Item SKU", "Manufacturer Number",
                  "Description", "MSRP", "Dlr Cost"])
    sheet.append([""] * 6)
    sheet.append(["PROX", "SKU-MSRP", "NUM-MSRP", "MSRP ROW", "999.00", "50.00"])
    output = BytesIO()
    book.save(output)
    boss.post(
        reverse("catalog_import_upload"),
        {"catalog": "analogs",
         "workbook": SimpleUploadedFile("dealer.xlsx", output.getvalue())},
        follow=True,
    )
    batch = CatalogImportBatch.objects.order_by("-pk").first()
    boss.post(reverse("catalog_import_apply", args=[batch.pk]), follow=True)

    entry = AftermarketCatalogPart.objects.get(manufacturer_number="NUM-MSRP")
    assert entry.msrp_usd == Decimal("999.00")  # исходные данные сохранены
    assert entry.dealer_cost_usd == Decimal("50.00")
    assert entry.part.recommended_price == _canonical(Decimal("50.00"))
    assert entry.part.recommended_price != _canonical(Decimal("999.00"))


# --- Обновление и повтор ----------------------------------------------------


def test_a_new_dealer_cost_moves_the_customer_price(boss, rates):
    _import(boss, [["PROX", "SKU-U", "NUM-U", "ПОРШЕНЬ", "89.08"]])
    entry = AftermarketCatalogPart.objects.get(manufacturer_number="NUM-U")
    first = entry.part.recommended_price
    assert first == _canonical(Decimal("89.08"))
    parts_before = PartType.objects.count()

    _import(boss, [["PROX", "SKU-U", "NUM-U", "ПОРШЕНЬ", "95.50"]])
    entry.refresh_from_db()
    entry.part.refresh_from_db()
    assert entry.dealer_cost_usd == Decimal("95.50")  # исходная цена обновилась
    assert entry.part.recommended_price == _canonical(Decimal("95.50"))
    assert entry.part.recommended_price != first
    assert PartType.objects.count() == parts_before  # дубля не появилось


def test_reimporting_the_same_file_changes_nothing(boss, rates):
    _import(boss, REAL_SHAPED_ROWS)
    before = {
        e.manufacturer_number: (e.dealer_cost_usd, e.part.recommended_price)
        for e in AftermarketCatalogPart.objects.select_related("part")
    }
    parts_before = PartType.objects.count()
    stock_before = _stock()

    _import(boss, REAL_SHAPED_ROWS)
    after = {
        e.manufacturer_number: (e.dealer_cost_usd, e.part.recommended_price)
        for e in AftermarketCatalogPart.objects.select_related("part")
    }
    assert after == before
    assert PartType.objects.count() == parts_before
    assert _stock() == stock_before


def test_recalculation_is_idempotent(boss, rates):
    _import(boss, REAL_SHAPED_ROWS)
    kwargs = dict(usd_rate=Decimal("105"), brp_markup=Decimal("40"),
                  polaris_markup=Decimal("40"), catalogs=frozenset({"aftermarket"}))
    assert refresh_linked_part_prices(**kwargs) == 0  # импорт уже посчитал
    plan = plan_linked_part_price_refresh(**kwargs)
    assert plan.updated == 0
    assert plan.unchanged == len(REAL_SHAPED_ROWS)


# --- Пересчёт уже заведённого каталога --------------------------------------


def test_recalculation_prices_cards_that_have_none(boss, rates):
    """Карточки, заведённые до этой правки, получают цену пересчётом."""
    _import(boss, REAL_SHAPED_ROWS)
    PartType.objects.filter(aftermarket_catalog_entry__isnull=False).update(
        recommended_price=None
    )
    kwargs = dict(usd_rate=Decimal("105"), brp_markup=Decimal("40"),
                  polaris_markup=Decimal("40"), catalogs=frozenset({"aftermarket"}))

    plan = plan_linked_part_price_refresh(**kwargs)
    assert plan.aftermarket_links == len(REAL_SHAPED_ROWS)
    assert plan.updated == len(REAL_SHAPED_ROWS)
    assert PartType.objects.filter(
        aftermarket_catalog_entry__isnull=False, recommended_price__isnull=False
    ).count() == 0  # план ничего не записал

    assert refresh_linked_part_prices(**kwargs) == len(REAL_SHAPED_ROWS)
    for row in REAL_SHAPED_ROWS:
        entry = AftermarketCatalogPart.objects.get(manufacturer_number=row[2])
        assert entry.part.recommended_price == _canonical(Decimal(row[4]))


def test_recalculation_leaves_zero_priced_cards_without_a_price(boss, rates):
    _import(boss, [["PROX", "SKU-Z", "NUM-Z", "НУЛЕВАЯ", "0"],
                   ["PROX", "SKU-P", "NUM-P", "С ЦЕНОЙ", "12.00"]])
    kwargs = dict(usd_rate=Decimal("105"), brp_markup=Decimal("40"),
                  polaris_markup=Decimal("40"), catalogs=frozenset({"aftermarket"}))
    refresh_linked_part_prices(**kwargs)
    zero = AftermarketCatalogPart.objects.filter(manufacturer_number="NUM-Z").first()
    if zero is not None:
        assert zero.part.recommended_price is None
    assert AftermarketCatalogPart.objects.get(
        manufacturer_number="NUM-P"
    ).part.recommended_price == _canonical(Decimal("12.00"))


def test_recalculation_does_not_touch_brp_or_manual_parts(boss, rates):
    """Пересчёт аналогов трогает только аналоги."""
    from apps.catalog.services import create_manual_part

    brp = BrpCatalogPart.objects.create(
        material_no="BRP-KEEP", material_no_norm="BRPKEEP", part_desc="BRP",
        wholesale_price_usd=Decimal("20"), retail_price_usd=Decimal("40"), is_current=True,
    )
    brp_part = promote_to_warehouse(brp, by=None)
    manual = create_manual_part(
        name="Ручная деталь", article="MANUAL-1", price=Decimal("777"),
        manufacturer_name="Мастерская",
    )
    _import(boss, REAL_SHAPED_ROWS)
    brp_price, manual_price = brp_part.recommended_price, manual.recommended_price

    refresh_linked_part_prices(
        usd_rate=Decimal("105"), brp_markup=Decimal("40"), polaris_markup=Decimal("40"),
        catalogs=frozenset({"aftermarket"}),
    )
    brp_part.refresh_from_db()
    manual.refresh_from_db()
    assert brp_part.recommended_price == brp_price
    assert manual.recommended_price == manual_price == Decimal("777.00")


# --- Безопасность соседних областей -----------------------------------------


def test_pricing_never_touches_stock(boss, rates):
    before = _stock()
    _import(boss, REAL_SHAPED_ROWS)
    assert _stock() == before
    refresh_linked_part_prices(
        usd_rate=Decimal("105"), brp_markup=Decimal("40"), polaris_markup=Decimal("40"),
        catalogs=frozenset({"aftermarket"}),
    )
    assert _stock() == before


def test_dealer_cost_never_becomes_warehouse_cost(boss, rates):
    _import(boss, REAL_SHAPED_ROWS)
    entry = AftermarketCatalogPart.objects.get(manufacturer_number="14-1001")
    assert entry.dealer_cost_usd == Decimal("32.74")  # исходная цена на месте
    assert StockLot.objects.filter(part_type=entry.part).count() == 0
    assert StockMovement.objects.filter(part_type=entry.part).count() == 0


def test_pricing_never_touches_customs(boss, rates):
    from apps.actions.models import PartCustomsDataVersion, PartCustomsInfo

    _import(boss, REAL_SHAPED_ROWS)
    part = AftermarketCatalogPart.objects.get(manufacturer_number="2910").part
    PartCustomsInfo.objects.create(
        part_type=part, country_of_origin="JAPAN", customs_unit_price_usd=Decimal("3.00")
    )
    versions_before = PartCustomsDataVersion.objects.count()
    customs_before = PartCustomsInfo.objects.get(part_type=part).customs_unit_price_usd

    refresh_linked_part_prices(
        usd_rate=Decimal("105"), brp_markup=Decimal("40"), polaris_markup=Decimal("40"),
        catalogs=frozenset({"aftermarket"}),
    )
    assert PartCustomsDataVersion.objects.count() == versions_before
    assert PartCustomsInfo.objects.get(part_type=part).customs_unit_price_usd == customs_before
    assert customs_before == Decimal("3.00")  # дилерская цена сюда не протекла


def test_pricing_creates_no_analog_relations(boss, rates):
    before = PartAnalog.objects.count()
    _import(boss, REAL_SHAPED_ROWS)
    refresh_linked_part_prices(
        usd_rate=Decimal("105"), brp_markup=Decimal("40"), polaris_markup=Decimal("40"),
        catalogs=frozenset({"aftermarket"}),
    )
    assert PartAnalog.objects.count() == before == 0


# --- Экран ------------------------------------------------------------------


def test_search_shows_the_calculated_rub_price(boss, rates):
    _import(boss, REAL_SHAPED_ROWS)
    html = boss.get(reverse("part_search") + "?q=14-1001").content.decode()
    expected = _canonical(Decimal("32.74"))
    assert "Цена:" in html  # операторский словарь: одна «Цена»
    assert f"{expected:,}".replace(",", " ") in html.replace(" ", " ")


def test_part_detail_shows_the_calculated_rub_price(boss, rates):
    _import(boss, REAL_SHAPED_ROWS)
    part = AftermarketCatalogPart.objects.get(manufacturer_number="14-1001").part
    html = boss.get(reverse("part_detail", args=[part.pk])).content.decode()
    expected = _canonical(Decimal("32.74"))
    assert f"{expected:,}".replace(",", " ") in html.replace(" ", " ")
