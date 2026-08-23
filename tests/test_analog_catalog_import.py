"""Импорт каталога аналогов из Excel.

Файл поставщика - недоверенные данные. Разбор ничего не исполняет, значения
берутся посчитанными, а решения принимаются консервативно: если по артикулу
нашлось несколько деталей, строка показывается человеку, а не решается за него
выбором первой попавшейся.

Главное правило домена действует и здесь: артикул не определяет деталь. У
аналога он часто совпадает с исходной, поэтому совпавшей считается только та
деталь, у которой сошлись все заполненные признаки.

Импорт трогает только справочник. Остатков он не создаёт вовсе.
"""
from decimal import Decimal

import pytest
from openpyxl import Workbook

from apps.catalog.models import PartAnalog, PartType
from apps.catalog.services import create_manual_part, link_analog
from apps.catalog_import.analog_catalog import (
    AnalogCatalogError,
    apply_file,
    build_plan,
)
from apps.inventory.models import PartItem, StockBalance, StockLot, StockMovement

HEADERS = [
    "Исходный артикул", "Артикул аналога", "Название аналога",
    "Цена", "Производитель аналога", "Штрихкод аналога",
]


def make_file(tmp_path, rows, *, headers=None, name="analogs.xlsx"):
    """Настоящий .xlsx: разбор должен проверяться на файле, а не на списке."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(headers if headers is not None else HEADERS)
    for row in rows:
        sheet.append(list(row))
    path = tmp_path / name
    workbook.save(path)
    return path


@pytest.fixture
def original(db):
    return create_manual_part(
        name="Поршень BRP", article="SAME-001", price=Decimal("10000"),
        manufacturer_name="BRP",
    )


def _stock_snapshot():
    return (
        StockMovement.objects.count(),
        StockLot.objects.count(),
        PartItem.objects.count(),
        StockBalance.objects.count(),
    )


# --- Обычный файл -------------------------------------------------------------------


def test_a_plain_file_is_planned_without_writing_anything(tmp_path, original):
    path = make_file(tmp_path, [
        ["SAME-001", "SAME-001", "Поршень XYZ", "4500", "XYZ", ""],
        ["SAME-001", "ALT-002", "Поршень ALT", "5000", "ALT", ""],
    ])
    before = PartType.objects.count(), PartAnalog.objects.count()

    plan = build_plan(path).as_summary()

    assert plan["rows_total"] == 2
    assert plan["will_create_parts"] == 2
    assert plan["will_create_links"] == 2
    assert plan["needs_attention"] == 0
    assert (PartType.objects.count(), PartAnalog.objects.count()) == before


def test_applying_creates_the_parts_and_the_links(tmp_path, original):
    path = make_file(tmp_path, [
        ["SAME-001", "SAME-001", "Поршень XYZ", "4500", "XYZ", "4600000000028"],
    ])

    result = apply_file(path)

    assert result["created_parts"] == 1
    assert result["created_links"] == 1
    analog = PartType.objects.get(name="Поршень XYZ")
    assert analog.recommended_price == Decimal("4500")
    assert analog.manufacturer.name == "XYZ"
    assert analog.numbers.get().value == "SAME-001"
    assert analog.barcodes.get().value == "4600000000028"
    assert PartAnalog.objects.filter(original=original, analog=analog).exists()


def test_the_import_never_creates_stock(tmp_path, original):
    """Каталог - это справочник. Остаток появляется только на приёмке."""
    path = make_file(tmp_path, [
        ["SAME-001", "SAME-001", "Поршень XYZ", "4500", "XYZ", ""],
        ["SAME-001", "ALT-002", "Поршень ALT", "5000", "ALT", ""],
    ])
    before = _stock_snapshot()

    apply_file(path)

    assert _stock_snapshot() == before


def test_the_analog_may_carry_the_same_article_as_the_original(tmp_path, original):
    path = make_file(tmp_path, [
        ["SAME-001", "SAME-001", "Поршень XYZ", "4500", "XYZ", ""],
    ])
    apply_file(path)

    analog = PartType.objects.get(name="Поршень XYZ")
    assert analog.numbers.get().value == original.numbers.get().value
    assert analog.pk != original.pk


# --- Повторный запуск ----------------------------------------------------------------


def test_the_same_file_twice_changes_nothing_the_second_time(tmp_path, original):
    """Тот же файл могут запустить дважды. Дубликатов быть не должно."""
    path = make_file(tmp_path, [
        ["SAME-001", "SAME-001", "Поршень XYZ", "4500", "XYZ", ""],
        ["SAME-001", "ALT-002", "Поршень ALT", "5000", "ALT", ""],
    ])
    first = apply_file(path)
    parts_after_first = PartType.objects.count()
    links_after_first = PartAnalog.objects.count()

    second = apply_file(path)

    assert first["created_parts"] == 2
    assert second["created_parts"] == 0
    assert second["reused_parts"] == 2
    assert second["already_linked"] == 2
    assert PartType.objects.count() == parts_after_first
    assert PartAnalog.objects.count() == links_after_first


def test_the_second_plan_says_everything_is_already_there(tmp_path, original):
    path = make_file(tmp_path, [
        ["SAME-001", "SAME-001", "Поршень XYZ", "4500", "XYZ", ""],
    ])
    apply_file(path)

    plan = build_plan(path).as_summary()
    assert plan["will_create_parts"] == 0
    assert plan["will_reuse_parts"] == 1
    assert plan["already_linked"] == 1
    assert plan["will_create_links"] == 0


def test_a_repeated_row_inside_one_file_does_not_double(tmp_path, original):
    path = make_file(tmp_path, [
        ["SAME-001", "SAME-001", "Поршень XYZ", "4500", "XYZ", ""],
        ["SAME-001", "SAME-001", "Поршень XYZ", "4500", "XYZ", ""],
    ])
    result = apply_file(path)

    assert PartType.objects.filter(name="Поршень XYZ").count() == 1
    assert PartAnalog.objects.count() == 1
    assert result["already_linked"] == 1


# --- Уже заведённые детали -------------------------------------------------------------


def test_an_existing_analog_is_reused_not_duplicated(tmp_path, original):
    """Совпали название, артикул и завод - это та же карточка."""
    existing = create_manual_part(
        name="Поршень XYZ", article="SAME-001", price=Decimal("4500"),
        manufacturer_name="XYZ",
    )
    path = make_file(tmp_path, [
        ["SAME-001", "SAME-001", "Поршень XYZ", "4500", "XYZ", ""],
    ])

    result = apply_file(path)

    assert result["created_parts"] == 0
    assert result["reused_parts"] == 1
    assert PartAnalog.objects.get().analog_id == existing.pk


def test_two_unrelated_parts_with_one_article_stay_ambiguous(tmp_path, original):
    """Ничем не связанные детали с одним номером файл различить не может.

    Обе карточки равноправны, и в файле нет ничего, что указывало бы на одну из
    них. Выбрать первую было бы тихой ошибкой, поэтому строка уходит человеку.
    """
    create_manual_part(name="Поршень XYZ", article="SAME-001", manufacturer_name="XYZ")
    path = make_file(tmp_path, [
        ["SAME-001", "SAME-001", "Поршень ABC", "3900", "ABC", ""],
    ])

    plan = build_plan(path).as_summary()

    assert plan["ambiguous"] == 1
    assert plan["will_create_links"] == 0
    assert "Производитель оригинала" in plan["problems"][0]["detail"], (
        "не сказано, чем именно различить"
    )


def test_naming_the_maker_resolves_it_and_a_third_same_article_appears(tmp_path, original):
    """Один артикул у разных деталей - норма, и склеивать их нельзя."""
    create_manual_part(name="Поршень XYZ", article="SAME-001", manufacturer_name="XYZ")
    path = make_file(
        tmp_path,
        [["SAME-001", "SAME-001", "Поршень ABC", "3900", "ABC", "", "BRP"]],
        headers=[*HEADERS, "Производитель оригинала"],
    )

    result = apply_file(path)

    assert result["created_parts"] == 1
    assert PartType.objects.filter(numbers__normalized_value="SAME001").count() == 3
    assert PartAnalog.objects.get().original_id == original.pk


def test_a_part_already_marked_as_an_analog_is_not_taken_for_the_original(tmp_path, original):
    """Сужение опирается на записанный факт, а не на догадку."""
    already = create_manual_part(
        name="Поршень XYZ", article="SAME-001", manufacturer_name="XYZ"
    )
    link_analog(original=original, analog=already)
    path = make_file(tmp_path, [
        ["SAME-001", "ALT-003", "Поршень третий", "3900", "ABC", ""],
    ])

    result = apply_file(path)

    assert result["created_links"] == 1
    assert PartAnalog.objects.filter(analog__name="Поршень третий").get().original_id == (
        original.pk
    )


def test_the_existing_stock_of_a_reused_part_is_untouched(tmp_path, original, django_user_model):
    """Переиспользование карточки не должно шевелить её остаток."""
    existing = create_manual_part(
        name="Поршень XYZ", article="SAME-001", manufacturer_name="XYZ"
    )
    StockBalance.objects.filter(part_type=existing)  # проверяем отсутствие записи
    before = _stock_snapshot()

    path = make_file(tmp_path, [
        ["SAME-001", "SAME-001", "Поршень XYZ", "9999", "XYZ", ""],
    ])
    apply_file(path)

    assert _stock_snapshot() == before
    existing.refresh_from_db()
    # Цена существующей карточки не переписывается файлом: за неё отвечает
    # человек, а импорт только связывает.
    assert existing.recommended_price is None


# --- Неоднозначность --------------------------------------------------------------------


def test_two_originals_with_one_article_are_shown_not_guessed(tmp_path, original):
    """Самое опасное место: выбрать первую было бы тихой ошибкой."""
    create_manual_part(name="Поршень BRP другой", article="SAME-001", manufacturer_name="BRP-2")
    path = make_file(tmp_path, [
        ["SAME-001", "ALT-002", "Поршень XYZ", "4500", "XYZ", ""],
    ])

    plan = build_plan(path).as_summary()

    assert plan["ambiguous"] == 1
    assert plan["will_create_links"] == 0
    assert "несколько" in plan["problems"][0]["reason"].lower()


def test_an_ambiguous_row_is_skipped_on_apply(tmp_path, original):
    create_manual_part(name="Поршень BRP другой", article="SAME-001", manufacturer_name="BRP-2")
    path = make_file(tmp_path, [
        ["SAME-001", "ALT-002", "Поршень XYZ", "4500", "XYZ", ""],
    ])

    result = apply_file(path)

    assert result["skipped"] == 1
    assert result["created_links"] == 0
    assert not PartAnalog.objects.exists()


def test_the_manufacturer_column_resolves_the_ambiguity(tmp_path, original):
    """Если завод указан, он сужает выбор - но не выбирает вслепую."""
    create_manual_part(name="Поршень BRP другой", article="SAME-001", manufacturer_name="ACME")
    path = make_file(
        tmp_path,
        [["SAME-001", "ALT-002", "Поршень XYZ", "4500", "XYZ", "", "BRP"]],
        headers=[*HEADERS, "Производитель оригинала"],
    )

    result = apply_file(path)

    assert result["created_links"] == 1
    assert PartAnalog.objects.get().original_id == original.pk


def test_one_row_may_be_ambiguous_while_the_rest_import(tmp_path, original):
    """Три спорных строки не должны срывать заведение остальной тысячи."""
    create_manual_part(name="Поршень BRP другой", article="DUP-001", manufacturer_name="BRP-2")
    create_manual_part(name="Поршень BRP третий", article="DUP-001", manufacturer_name="BRP-3")
    path = make_file(tmp_path, [
        ["SAME-001", "A-1", "Аналог первый", "100", "XYZ", ""],
        ["DUP-001", "A-2", "Аналог второй", "200", "XYZ", ""],
        ["SAME-001", "A-3", "Аналог третий", "300", "XYZ", ""],
    ])

    result = apply_file(path)

    assert result["created_links"] == 2
    assert result["skipped"] == 1
    assert PartAnalog.objects.count() == 2


# --- Плохие строки ------------------------------------------------------------------------


@pytest.mark.parametrize(
    ("row", "reason"),
    [
        (["", "A-1", "Аналог", "100", "", ""], "Не указан исходный артикул"),
        (["SAME-001", "A-1", "", "100", "", ""], "Не указано название аналога"),
        (["SAME-001", "A-1", "Аналог", "не число", "", ""], "Некорректная цена"),
        (["SAME-001", "A-1", "Аналог", "-5", "", ""], "Некорректная цена"),
        (["НЕТ-ТАКОГО", "A-1", "Аналог", "100", "", ""], "Не найдена исходная деталь"),
    ],
)
def test_a_bad_row_is_reported_with_its_number(tmp_path, original, row, reason):
    path = make_file(tmp_path, [row])

    plan = build_plan(path).as_summary()

    assert plan["invalid"] == 1
    assert plan["problems"][0]["row"] == 2, "номер строки должен указывать на файл"
    assert plan["problems"][0]["reason"] == reason


def test_a_bad_row_never_stops_the_file_from_being_read(tmp_path, original):
    path = make_file(tmp_path, [
        ["SAME-001", "A-1", "Хороший", "100", "", ""],
        ["", "A-2", "Плохой", "100", "", ""],
        ["SAME-001", "A-3", "Тоже хороший", "300", "", ""],
    ])

    result = apply_file(path)

    assert result["created_parts"] == 2
    assert result["skipped"] == 1


def test_a_file_without_the_needed_columns_says_which_ones(tmp_path, original):
    path = make_file(tmp_path, [["что-то"]], headers=["Совсем не то"])

    with pytest.raises(AnalogCatalogError) as failure:
        build_plan(path)
    assert "не хватает" in str(failure.value).lower()


def test_an_empty_file_is_refused_with_words(tmp_path, original):
    workbook = Workbook()
    path = tmp_path / "empty.xlsx"
    workbook.save(path)

    with pytest.raises(AnalogCatalogError):
        build_plan(path)


def test_a_file_that_is_not_excel_is_refused(tmp_path, original):
    path = tmp_path / "not-excel.xlsx"
    path.write_bytes(b"\x00\x01 not a workbook")

    with pytest.raises(AnalogCatalogError) as failure:
        build_plan(path)
    assert "не читается" in str(failure.value).lower()


def test_a_wrong_extension_is_refused(tmp_path, original):
    path = tmp_path / "analogs.csv"
    path.write_text("a,b,c", encoding="utf-8")

    with pytest.raises(AnalogCatalogError) as failure:
        build_plan(path)
    assert "xlsx" in str(failure.value).lower()


# --- Как оператор на самом деле заполняет файл ------------------------------------------------


@pytest.mark.parametrize(
    "written", ["SAME-001", "same-001", " SAME-001 ", "SAME 001", "SAME.001"]
)
def test_the_original_article_is_found_however_it_is_written(tmp_path, original, written):
    path = make_file(tmp_path, [[written, "A-1", "Аналог", "100", "", ""]])
    result = apply_file(path)
    assert result["created_links"] == 1


@pytest.mark.parametrize("header_style", ["original_article", "Оригинальный артикул"])
def test_header_synonyms_are_understood(tmp_path, original, header_style):
    path = make_file(
        tmp_path,
        [["SAME-001", "A-1", "Аналог", "100", "", ""]],
        headers=[header_style, "Артикул аналога", "Название аналога",
                 "Цена", "Производитель аналога", "Штрихкод аналога"],
    )
    assert build_plan(path).as_summary()["will_create_links"] == 1


def test_a_comma_decimal_price_is_accepted(tmp_path, original):
    path = make_file(tmp_path, [["SAME-001", "A-1", "Аналог", "4500,50", "", ""]])
    apply_file(path)
    assert PartType.objects.get(name="Аналог").recommended_price == Decimal("4500.50")


def test_an_empty_price_is_allowed(tmp_path, original):
    path = make_file(tmp_path, [["SAME-001", "A-1", "Аналог", "", "", ""]])
    apply_file(path)
    assert PartType.objects.get(name="Аналог").recommended_price is None


def test_blank_lines_in_the_middle_are_ignored(tmp_path, original):
    path = make_file(tmp_path, [
        ["SAME-001", "A-1", "Первый", "100", "", ""],
        ["", "", "", "", "", ""],
        ["SAME-001", "A-2", "Второй", "200", "", ""],
    ])
    plan = build_plan(path).as_summary()
    assert plan["rows_total"] == 2
    assert plan["invalid"] == 0


def test_a_number_typed_as_a_number_still_works(tmp_path, original):
    """Excel часто хранит артикул числом, а не текстом."""
    numeric = create_manual_part(name="Числовой", article="420123456")
    path = make_file(tmp_path, [[420123456, "A-9", "Аналог числового", 500, "", ""]])

    result = apply_file(path)

    assert result["created_links"] == 1
    assert PartAnalog.objects.get().original_id == numeric.pk


# --- Один аналог для нескольких исходных ---------------------------------------------------


def test_the_same_analog_may_be_linked_to_several_originals(tmp_path, original):
    second = create_manual_part(name="Поршень BRP 850", article="ORIGINAL-002")
    path = make_file(tmp_path, [
        ["SAME-001", "ALT-002", "Поршень ALT", "5000", "ALT", ""],
        ["ORIGINAL-002", "ALT-002", "Поршень ALT", "5000", "ALT", ""],
    ])

    result = apply_file(path)

    assert result["created_parts"] == 1, "одна и та же деталь заведена дважды"
    assert result["created_links"] == 2
    analog = PartType.objects.get(name="Поршень ALT")
    assert set(
        PartAnalog.objects.filter(analog=analog).values_list("original_id", flat=True)
    ) == {original.pk, second.pk}


def test_a_reverse_pair_in_the_file_is_reported(tmp_path, original):
    analog = create_manual_part(name="Поршень XYZ", article="ALT-002", manufacturer_name="XYZ")
    link_analog(original=original, analog=analog)
    path = make_file(tmp_path, [
        ["ALT-002", "SAME-001", "Поршень BRP", "10000", "BRP", ""],
    ])

    plan = build_plan(path).as_summary()

    assert plan["will_create_links"] == 0
    assert plan["problems"], "обратная пара должна попасть в список"


# --- Штрихкод ---------------------------------------------------------------------------------


def test_a_taken_barcode_is_reported_and_the_row_skipped(tmp_path, original):
    create_manual_part(name="Кто-то", article="Z-1", barcode="4600000000028")
    path = make_file(tmp_path, [
        ["SAME-001", "A-1", "Аналог", "100", "", "4600000000028"],
    ])

    result = apply_file(path)

    assert result["skipped"] == 1
    assert result["created_parts"] == 0
    assert "Штрихкод" in result["problems"][0]["detail"]


# --- Формулы и безопасность ---------------------------------------------------------------------


def test_a_formula_cell_is_never_executed_and_its_text_never_stored(tmp_path, original):
    """Книга открывается по значениям, и формулы не вычисляются.

    У файла, сохранённого не Excel-ом, посчитанного значения нет вовсе, поэтому
    ячейка приходит пустой. Важно здесь другое: текст формулы никуда не попадает
    и ничего не исполняет.
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(HEADERS)
    sheet.append(["SAME-001", "A-1", "Аналог", "=1+1", "=2+2", ""])
    path = tmp_path / "formula.xlsx"
    workbook.save(path)

    apply_file(path)

    made = PartType.objects.get(numbers__value="A-1")
    assert made.recommended_price is None, "формула превратилась в цену"
    assert made.manufacturer is None or "=" not in made.manufacturer.name
    assert "=" not in made.name


def test_text_from_the_file_is_stored_as_text_not_as_markup(tmp_path, original):
    path = make_file(tmp_path, [
        ["SAME-001", "A-1", "<b>Аналог</b>", "100", "", ""],
    ])
    apply_file(path)
    made = PartType.objects.get(numbers__value="A-1")
    assert made.name == "<b>Аналог</b>"


# --- Размер файла ------------------------------------------------------------------------------


@pytest.mark.slow
def test_a_large_file_is_planned_without_a_query_per_row(tmp_path, original):
    """Десять тысяч строк - это не повод для десяти тысяч обращений к базе."""
    from django.db import connection
    from django.test.utils import CaptureQueriesContext

    rows = [["SAME-001", f"A-{index}", f"Аналог {index}", "100", "", ""]
            for index in range(10_000)]
    path = make_file(tmp_path, rows, name="big.xlsx")

    with CaptureQueriesContext(connection) as captured:
        plan = build_plan(path).as_summary()

    assert plan["rows_total"] == 10_000
    assert plan["will_create_parts"] == 10_000
    # Всё нужное берётся из базы заранее, поэтому число запросов не зависит от
    # числа строк вовсе. Измерено: до указателя разбор стоил около двух запросов
    # на строку и 56 секунд, после - 6 секунд.
    assert len(captured) < 20, f"запросов {len(captured)} на 10 000 строк"
