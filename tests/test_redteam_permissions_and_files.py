"""Red-team: права доступа и приватность коммерческого прайса.

Проверяется не «страница открылась», а обратное: что закрытая страница НЕ
отдаёт данные. Отдельно проверяется, что при отказе персональные данные
клиента и содержимое прайса не утекают в тело ответа.
"""
import pytest
from django.contrib.auth.models import Group
from django.core.files.uploadedfile import SimpleUploadedFile
from django.urls import reverse
from openpyxl import Workbook

from apps.accounts import roles
from apps.catalog_import.models import CatalogImportBatch
from apps.catalog_import.services import save_upload
from apps.customers.models import Customer

PASSWORD = "parol-12345"
PHONE = "+7 912 123-45-67"

HEADERS = [
    "Material_No", "Part_Desc", "Last_Yr_Util", "Status",
    "РОЗНИЦА", "ОПТОВАЯ", "ЗАМЕНА НОМЕРА", "ЗАМЕНА НОМЕРА",
]


@pytest.fixture
def make_user(db, django_user_model):
    def _make(username, *, role=None, is_superuser=False):
        if is_superuser:
            user = django_user_model.objects.create_superuser(username=username, password=PASSWORD)
        else:
            user = django_user_model.objects.create_user(username=username, password=PASSWORD)
        if role:
            user.groups.add(Group.objects.get(name=role))
        return user

    return _make


@pytest.fixture
def customer(db):
    return Customer.objects.create(name="Иванов Иван", phone=PHONE)


@pytest.fixture
def batch(db, make_user, settings, tmp_path):
    settings.PRIVATE_MEDIA_ROOT = str(tmp_path / "private")
    admin = make_user("owner", is_superuser=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(HEADERS)
    sheet.append([""] * len(HEADERS))
    sheet.append(["420831955", "SECRET ROLLER", "", "", "25.99", "20.00", "", ""])
    path = tmp_path / "price.xlsx"
    workbook.save(path)
    upload = SimpleUploadedFile(
        "price.xlsx", path.read_bytes(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    return save_upload(upload, catalog="brp", by=admin)


CATALOG_ROUTES = ["catalog_import_list"]
CUSTOMER_ROUTES = ["customer_list", "customer_create"]
REPORT_ROUTES = [
    "reports_sales_by_client",
    "reports_repairs_by_client",
    "reports_clients_overview",
]


# --- Аноним --------------------------------------------------------------------------------


@pytest.mark.parametrize(
    "route", CATALOG_ROUTES + CUSTOMER_ROUTES + REPORT_ROUTES
)
def test_anonymous_is_never_served(client, db, route):
    response = client.get(reverse(route))
    assert response.status_code in (301, 302)
    assert "/login" in response["Location"]


def test_anonymous_cannot_download_price(client, batch):
    response = client.get(reverse("catalog_import_download", args=[batch.pk]))
    assert response.status_code in (301, 302)
    assert "/login" in response["Location"]


# --- Складской оператор ----------------------------------------------------------------------


@pytest.mark.parametrize("role", [roles.STOREKEEPER, roles.SELLER, roles.VIEWER])
def test_operator_roles_cannot_touch_catalog_import(client, make_user, db, role):
    make_user(f"user-{role}", role=role)
    client.login(username=f"user-{role}", password=PASSWORD)
    assert client.get(reverse("catalog_import_list")).status_code == 403
    assert client.post(reverse("catalog_import_upload"), {}).status_code == 403


@pytest.mark.parametrize("role", [roles.STOREKEEPER, roles.SELLER, roles.VIEWER])
def test_operator_roles_cannot_apply_or_download(client, make_user, db, role, batch):
    make_user(f"user-{role}", role=role)
    client.login(username=f"user-{role}", password=PASSWORD)
    assert client.post(reverse("catalog_import_apply", args=[batch.pk])).status_code == 403
    assert client.get(reverse("catalog_import_download", args=[batch.pk])).status_code == 403
    assert client.get(reverse("catalog_import_inspect", args=[batch.pk])).status_code == 403


def test_denied_catalog_response_leaks_no_price_data(client, make_user, db, batch):
    make_user("sklad", role=roles.STOREKEEPER)
    client.login(username="sklad", password=PASSWORD)
    for route, args in (
        ("catalog_import_detail", [batch.pk]),
        ("catalog_import_download", [batch.pk]),
        ("catalog_import_inspect", [batch.pk]),
    ):
        response = client.get(reverse(route, args=args))
        assert response.status_code == 403
        body = response.content.decode(errors="replace")
        assert "SECRET ROLLER" not in body
        assert batch.source_sha256 not in body


# --- Персональные данные клиента ---------------------------------------------------------------


def test_user_without_any_client_right_gets_no_phone(client, make_user, customer):
    """Роль без прав на документы клиента не получает ни страницу, ни телефон."""
    make_user("nobody")
    client.login(username="nobody", password=PASSWORD)
    for route, args in (
        ("customer_list", []),
        ("customer_detail", [customer.pk]),
        ("customer_edit", [customer.pk]),
    ):
        response = client.get(reverse(route, args=args))
        assert response.status_code == 403
        assert PHONE not in response.content.decode(errors="replace")


def test_viewer_cannot_edit_customer(client, make_user, customer):
    make_user("viewer", role=roles.VIEWER)
    client.login(username="viewer", password=PASSWORD)
    assert client.get(reverse("customer_edit", args=[customer.pk])).status_code == 403
    response = client.post(
        reverse("customer_edit", args=[customer.pk]),
        {"name": "Взломан", "phone": "", "comment": ""},
    )
    assert response.status_code == 403
    customer.refresh_from_db()
    assert customer.name == "Иванов Иван"


# --- Приватность файла --------------------------------------------------------------------------


def test_owner_download_has_private_headers(client, make_user, batch):
    make_user("boss", is_superuser=True)
    client.login(username="boss", password=PASSWORD)
    response = client.get(reverse("catalog_import_download", args=[batch.pk]))
    assert response.status_code == 200
    assert response["Cache-Control"] == "private, no-store"
    assert response["X-Content-Type-Options"] == "nosniff"
    response.close()


def test_unknown_batch_id_is_404_not_500(client, make_user, db):
    make_user("boss", is_superuser=True)
    client.login(username="boss", password=PASSWORD)
    assert client.get(reverse("catalog_import_download", args=[999999])).status_code == 404


def test_missing_file_is_404(client, make_user, batch):
    from apps.catalog_import.services import stored_file_path

    make_user("boss", is_superuser=True)
    client.login(username="boss", password=PASSWORD)
    stored_file_path(batch).unlink()
    assert client.get(reverse("catalog_import_download", args=[batch.pk])).status_code == 404


def test_stored_path_is_generated_not_user_supplied(batch):
    """Имя файла в хранилище не берётся из загрузки: обход пути невозможен."""
    assert batch.source_filename == "price.xlsx"
    assert batch.stored_path != batch.source_filename
    assert ".." not in batch.stored_path
    assert "/" not in batch.stored_path and "\\" not in batch.stored_path
    assert batch.stored_path.endswith(".xlsx")


def test_traversal_filename_cannot_escape_storage(db, make_user, settings, tmp_path):
    """Имя с обходом каталога не влияет на то, куда ляжет файл."""
    from apps.catalog_import.services import stored_file_path

    settings.PRIVATE_MEDIA_ROOT = str(tmp_path / "private")
    admin = make_user("owner2", is_superuser=True)
    workbook = Workbook()
    workbook.active.append(HEADERS)
    path = tmp_path / "evil.xlsx"
    workbook.save(path)
    upload = SimpleUploadedFile(
        "../../../../evil.xlsx", path.read_bytes(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    created = save_upload(upload, catalog="brp", by=admin)
    resolved = stored_file_path(created)
    root = (tmp_path / "private" / "catalog-imports").resolve()
    assert resolved.parent == root


def test_batch_of_unknown_catalog_cannot_be_created(db, make_user, settings, tmp_path):
    from apps.catalog_import.adapters import CatalogAdapterError

    settings.PRIVATE_MEDIA_ROOT = str(tmp_path / "private")
    admin = make_user("owner3", is_superuser=True)
    workbook = Workbook()
    workbook.active.append(HEADERS)
    path = tmp_path / "x.xlsx"
    workbook.save(path)
    upload = SimpleUploadedFile(
        "x.xlsx", path.read_bytes(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    with pytest.raises(CatalogAdapterError):
        save_upload(upload, catalog="polaris", by=admin)
    assert not CatalogImportBatch.objects.exists()
