from decimal import Decimal
from io import BytesIO

import pytest
from django.core.files.uploadedfile import SimpleUploadedFile
from django.urls import reverse
from openpyxl import Workbook

from apps.catalog.models import PartAnalog, PartNumber, PartType
from apps.catalog.services import create_manual_part, link_analog
from apps.catalog_import.models import AftermarketCatalogPart, CatalogImportBatch
from apps.inventory.models import StockBalance, StockLot, StockMovement

HEADERS = ["Manufacturer", "Item SKU", "Manufacturer Number", "Description", "MSRP", "Dlr Cost"]


def workbook(rows):
    book = Workbook()
    sheet = book.active
    sheet.title = "priceupdate"
    sheet.append(HEADERS)
    sheet.append(["", "", "", "", "", ""])
    for row in rows:
        sheet.append(row)
    output = BytesIO()
    book.save(output)
    return output.getvalue()


def upload(rows):
    return SimpleUploadedFile(
        "dealer.xlsx",
        workbook(rows),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def stock_snapshot():
    return (StockBalance.objects.count(), StockLot.objects.count(), StockMovement.objects.count())


@pytest.fixture
def boss(client, django_user_model):
    user = django_user_model.objects.create_superuser("boss", password="password")
    client.force_login(user)
    return client


def send(client, rows):
    return client.post(
        reverse("catalog_import_upload"),
        {"catalog": "analogs", "workbook": upload(rows)},
        follow=True,
    )


def apply(client, rows):
    send(client, rows)
    batch = CatalogImportBatch.objects.latest("pk")
    client.post(reverse("catalog_import_apply", args=[batch.pk]), follow=True)
    batch.refresh_from_db()
    return batch


def test_known_dealer_format_is_detected_previewed_and_applied_without_stock(boss):
    rows = [[" 40 BELOW ", "00128006", "SM-12750", "40 BELOW, TUNNEL PACK", "101.95", "63.31"]]
    before = stock_snapshot()

    response = send(boss, rows)

    assert response.status_code == 200
    batch = CatalogImportBatch.objects.get()
    assert batch.catalog == CatalogImportBatch.Catalog.AFTERMARKET
    assert batch.status == CatalogImportBatch.Status.CHECKED
    assert batch.summary["format"] == "AFTERMARKET_SUPPLIER_CATALOG"
    assert batch.summary["blank_rows"] == 1
    assert batch.summary["currency"] == "USD"
    assert "Каталог аналогов / aftermarket" in response.content.decode()
    assert "не хватает обязательных колонок" not in response.content.decode().lower()
    assert PartType.objects.count() == 0

    boss.post(reverse("catalog_import_apply", args=[batch.pk]), follow=True)
    entry = AftermarketCatalogPart.objects.select_related("part", "manufacturer").get()
    assert entry.manufacturer.name == "40 BELOW"
    assert entry.manufacturer_number == "SM-12750"
    assert entry.supplier_sku == "00128006"
    assert entry.msrp_usd == Decimal("101.95")
    assert entry.dealer_cost_usd == Decimal("63.31")
    assert entry.part.recommended_price is None
    assert entry.part.numbers.get(kind=PartNumber.Kind.ARTICLE).value == "SM-12750"
    assert entry.part.numbers.get(kind=PartNumber.Kind.INTERNAL_REF).value == "00128006"
    assert stock_snapshot() == before
    assert not PartAnalog.objects.exists()


def test_reimport_updates_source_usd_values_without_zeroing_known_price_or_creating_duplicates(
    boss,
):
    apply(boss, [["40 BELOW", "128006", "SM-12750", "Old", "101.95", "63.31"]])
    batch = apply(boss, [["40 below", "128006", "SM-12750", "New", "0", "105.54"]])

    entry = AftermarketCatalogPart.objects.select_related("part").get()
    assert batch.apply_summary["updated_parts"] == 1
    assert AftermarketCatalogPart.objects.count() == 1
    assert PartType.objects.count() == 1
    assert entry.part.name == "New"
    assert entry.msrp_usd == Decimal("101.95")
    assert entry.dealer_cost_usd == Decimal("105.54")


def test_manual_same_manufacturer_and_article_is_ambiguous_not_silently_merged(boss):
    create_manual_part(name="Manual", article="SM-12750", manufacturer_name="40 BELOW")
    response = send(boss, [["40 BELOW", "128006", "SM-12750", "Imported", "101.95", "63.31"]])

    batch = CatalogImportBatch.objects.get()
    assert response.status_code == 200
    assert batch.status == CatalogImportBatch.Status.CHECK_FAILED
    assert batch.summary["ambiguous"] == 1
    assert AftermarketCatalogPart.objects.count() == 0


def test_imported_part_can_be_linked_later_without_import_creating_relation(boss):
    original = create_manual_part(name="Original", article="OEM-1", manufacturer_name="BRP")
    apply(boss, [["40 BELOW", "128006", "SM-12750", "Independent", "101.95", "63.31"]])
    imported = AftermarketCatalogPart.objects.get().part

    link_analog(original=original, analog=imported)

    assert PartAnalog.objects.get().analog_id == imported.pk


def test_imported_article_and_supplier_sku_are_available_to_ordinary_search(boss):
    apply(boss, [["40 BELOW", "00128006", "SM-12750", "Independent", "101.95", "63.31"]])

    by_article = boss.get(reverse("part_search"), {"q": "SM-12750"})
    by_sku = boss.get(reverse("part_search"), {"q": "00128006"})

    assert "Independent" in by_article.content.decode()
    assert "Independent" in by_sku.content.decode()


def test_numeric_sku_with_a_zero_format_keeps_leading_zeroes(boss):
    book = Workbook()
    sheet = book.active
    sheet.title = "priceupdate"
    sheet.append(HEADERS)
    sheet.append(["", "", "", "", "", ""])
    sheet.append(["40 BELOW", 128006, "SM-12750", "Independent", "101.95", "63.31"])
    sheet.cell(3, 2).number_format = "00000000"
    output = BytesIO()
    book.save(output)

    boss.post(
        reverse("catalog_import_upload"),
        {
            "catalog": "analogs",
            "workbook": SimpleUploadedFile("dealer.xlsx", output.getvalue()),
        },
        follow=True,
    )
    batch = CatalogImportBatch.objects.get()
    boss.post(reverse("catalog_import_apply", args=[batch.pk]), follow=True)

    assert AftermarketCatalogPart.objects.get().supplier_sku == "00128006"


def test_formula_price_is_not_evaluated_or_written_as_zero(boss):
    book = Workbook()
    sheet = book.active
    sheet.title = "priceupdate"
    sheet.append(HEADERS)
    sheet.append(["", "", "", "", "", ""])
    sheet.append(["40 BELOW", "128006", "SM-12750", "Independent", "=1+1", "=2+2"])
    output = BytesIO()
    book.save(output)

    boss.post(
        reverse("catalog_import_upload"),
        {
            "catalog": "analogs",
            "workbook": SimpleUploadedFile("dealer.xlsx", output.getvalue()),
        },
        follow=True,
    )
    batch = CatalogImportBatch.objects.get()
    boss.post(reverse("catalog_import_apply", args=[batch.pk]), follow=True)

    entry = AftermarketCatalogPart.objects.get()
    assert entry.msrp_usd is None
    assert entry.dealer_cost_usd is None


@pytest.mark.parametrize("value", ["-1", "not-a-price"])
def test_invalid_usd_price_is_reported_without_catalog_mutation(boss, value):
    send(boss, [["40 BELOW", "128006", "SM-12750", "Independent", value, "63.31"]])

    batch = CatalogImportBatch.objects.get()
    assert batch.status == CatalogImportBatch.Status.CHECK_FAILED
    assert batch.summary["invalid"] == 1
    assert not AftermarketCatalogPart.objects.exists()
