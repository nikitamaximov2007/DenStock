"""Layer 33.1 — запоминание веса одной штуки для таможенного экспорта.

Распространяет уже принятый принцип (таможенная область применения
запоминается на PartCustomsInfo и переиспользуется во всех следующих
экспортах) на вес брутто/нетто одной штуки. Поля gross_weight_kg/
net_weight_kg уже существовали в модели (Layer 33) — новых полей и миграций
не потребовалось. «Вес брутто сумма» (колонка I / Excel-формула =J*G) не
хранится нигде: считается на лету из сохранённого веса одной штуки и
количества конкретного экспорта.

Фикстуры скопированы (не импортированы) из test_customs_export.py — тот же
паттерн, что и в test_customs_application_area.py: cross-file import ловит
ruff F811, per-file фикстуры — принятая в проекте практика.
"""
from decimal import Decimal
from io import BytesIO

import openpyxl
import pytest
from django.contrib.auth.models import Group
from django.urls import reverse

from apps.accounts import roles
from apps.actions.models import PartCustomsInfo
from apps.actions.services import (
    MANUAL_WEIGHT_NOTE,
    parse_weight_kg,
    part_export_data,
    perform_action,
)
from apps.brp.models import BrpCatalogPart
from apps.brp.services import promote_to_warehouse as promote_brp
from apps.inventory.services import create_stock_lot, receive_stock_lot
from apps.polaris.models import PolarisCatalogPart
from apps.polaris.services import promote_to_warehouse as promote_polaris
from apps.procurement.models import Batch, BatchLine
from apps.procurement.services import finalize_cost
from apps.suppliers.models import Supplier
from apps.warehouse.models import StorageLocation

PASSWORD = "parol-12345"
SHEET = "Лист1"
DATA_ROW = 10

ApplicationArea = PartCustomsInfo.ApplicationArea


@pytest.fixture
def make_user(db, django_user_model):
    def _make(username, *, role=None, is_superuser=False):
        if is_superuser:
            user = django_user_model.objects.create_superuser(username=username, password=PASSWORD)
        else:
            user = django_user_model.objects.create_user(username=username, password=PASSWORD)
        if role:
            user.groups.add(Group.objects.get(name=role))
        return user

    return _make


@pytest.fixture
def admin(make_user):
    return make_user("admin", is_superuser=True)


def _stock(part, location, qty, sup, admin):
    batch = Batch.objects.create(supplier=sup, shipping_cost=Decimal("0"))
    line = BatchLine.objects.create(
        batch=batch, part_type=part, quantity=Decimal(str(qty)),
        unit_cost_currency=Decimal("1"),
    )
    batch.status = Batch.Status.ACCEPTED
    batch.save(update_fields=["status"])
    finalize_cost(batch, admin)
    line.refresh_from_db()
    lot = create_stock_lot(line, location, Decimal(str(qty)))
    receive_stock_lot(lot, by=admin)
    return lot


@pytest.fixture
def env(db, admin):
    sup = Supplier.objects.create(name="Стартовый ввод")
    loc = StorageLocation.objects.create(
        name="C04", code="S04-L03-D01-C04", storage_allowed=True, is_active=True
    )
    return {"sup": sup, "loc": loc, "admin": admin}


def _brp(env, *, material, retail="10", wholesale="7", replacement="", desc="BELT DRIVE", qty=5):
    brp = BrpCatalogPart.objects.create(
        material_no=material, part_desc=desc,
        retail_price_usd=Decimal(retail), replacement_no_1=replacement,
        wholesale_price_usd=Decimal(wholesale) if wholesale is not None else None,
    )
    part = promote_brp(brp, by=env["admin"])
    _stock(part, env["loc"], qty, env["sup"], env["admin"])
    return part, brp


def _polaris(env, *, number, wholesale="6", retail="20", superseded="", desc="SEAL", qty=5):
    pol = PolarisCatalogPart.objects.create(
        part_number=number, part_name=desc, superseded_number=superseded,
        wholesale_price_usd=Decimal(wholesale) if wholesale is not None else None,
        retail_price_usd=Decimal(retail) if retail is not None else None,
    )
    part = promote_polaris(pol, by=env["admin"])
    _stock(part, env["loc"], qty, env["sup"], env["admin"])
    return part, pol


def _sell(env, part, *, qty="1", number="", comment="Иванов", location=None):
    return perform_action(
        part=part, location=location or env["loc"], action_type="sale",
        quantity=qty, customer_comment=comment, scanned_number=number, by=env["admin"],
    )


def _login(client, make_user, *, superuser=True, name="boss"):
    make_user(name, is_superuser=superuser)
    client.login(username=name, password=PASSWORD)


def _sheet(content: bytes):
    return openpyxl.load_workbook(BytesIO(content))[SHEET]


def _save(client, part, **fields):
    fields.setdefault("next", "/inventory/actions/report/")
    return client.post(reverse("actions_customs_quick_save", args=[part.pk]), fields)


# --- 1-3. Сохранение и повторное открытие -----------------------------------------------


def test_post_saves_gross_weight(client, make_user, env):
    part, _ = _brp(env, material="219800345")
    _login(client, make_user)
    resp = _save(client, part, gross_weight_kg="0.350")
    assert resp.status_code == 302
    customs = PartCustomsInfo.objects.get(part_type=part)
    assert customs.gross_weight_kg == Decimal("0.350")


def test_post_saves_net_weight(client, make_user, env):
    part, _ = _brp(env, material="219800345")
    _login(client, make_user)
    resp = _save(client, part, net_weight_kg="0.300")
    assert resp.status_code == 302
    customs = PartCustomsInfo.objects.get(part_type=part)
    assert customs.net_weight_kg == Decimal("0.300")


def test_saved_values_prefilled_on_reopen(client, make_user, env):
    part, _ = _brp(env, material="219800345")
    _sell(env, part, number="219800345")  # без действия строки в отчёте не будет
    _login(client, make_user)
    _save(client, part, gross_weight_kg="0.5", net_weight_kg="0.4",
          application_area=ApplicationArea.SNOWMOBILE)
    html = client.get(reverse("actions_customs_edit", args=[part.pk])).content.decode()
    assert 'value="0.500"' in html
    assert 'value="0.400"' in html
    html = client.get(reverse("actions_report")).content.decode()
    assert "0.500" in html  # значение уже подставлено в форму быстрого сохранения
    assert "0.400" in html


# --- 4-6. Автоматическое переиспользование при следующих продажах/экспортах -------------


def test_next_sale_of_same_part_uses_saved_weights(client, make_user, env):
    part, _ = _brp(env, material="219800345")
    _sell(env, part, number="219800345")  # первая продажа - весов ещё нет
    _login(client, make_user)
    _save(client, part, gross_weight_kg="0.35", net_weight_kg="0.30")
    _sell(env, part, number="219800345")  # вторая продажа той же детали
    row = part_export_data(part)
    assert row["gross_weight_kg"] == Decimal("0.35")
    assert row["net_weight_kg"] == Decimal("0.30")


def test_export_uses_saved_gross(client, make_user, env):
    part, _ = _brp(env, material="219800345")
    PartCustomsInfo.objects.create(part_type=part, gross_weight_kg=Decimal("0.35"))
    _sell(env, part, number="219800345")
    _login(client, make_user)
    sheet = _sheet(client.get(reverse("actions_export")).content)
    assert Decimal(str(sheet[f"G{DATA_ROW}"].value)) == Decimal("0.35")


def test_export_uses_saved_net(client, make_user, env):
    part, _ = _brp(env, material="219800345")
    PartCustomsInfo.objects.create(part_type=part, net_weight_kg=Decimal("0.30"))
    _sell(env, part, number="219800345")
    _login(client, make_user)
    sheet = _sheet(client.get(reverse("actions_export")).content)
    assert Decimal(str(sheet[f"H{DATA_ROW}"].value)) == Decimal("0.30")


# --- 7-9. Общий вес брутто: считается, не хранится -----------------------------------------


def test_gross_total_is_gross_times_quantity(client, make_user, env):
    part, _ = _brp(env, material="219800345")
    PartCustomsInfo.objects.create(part_type=part, gross_weight_kg=Decimal("0.35"))
    _sell(env, part, number="219800345", qty="2")
    row = part_export_data(part)
    row["quantity"] = Decimal("2")
    total = row["gross_weight_kg"] * row["quantity"]
    assert total == Decimal("0.700")  # 0.35 * 2
    _login(client, make_user)
    html = client.get(reverse("actions_report")).content.decode()
    assert "0,700" in html  # ru-локаль: запятая в десятичном разделителе на отображении


def test_gross_total_recalculates_for_different_quantity(client, make_user, env):
    part, _ = _brp(env, material="219800345")
    PartCustomsInfo.objects.create(part_type=part, gross_weight_kg=Decimal("0.350"))
    _login(client, make_user)

    _sell(env, part, number="219800345", qty="2")
    sheet = _sheet(client.get(reverse("actions_export")).content)
    assert sheet[f"I{DATA_ROW}"].value == f"=J{DATA_ROW}*G{DATA_ROW}"
    assert Decimal(str(sheet[f"J{DATA_ROW}"].value)) == Decimal("2")

    _sell(env, part, number="219800345", qty="3")  # итого 5 после второй продажи
    sheet = _sheet(client.get(reverse("actions_export")).content)
    assert Decimal(str(sheet[f"J{DATA_ROW}"].value)) == Decimal("5")
    assert Decimal(str(sheet[f"G{DATA_ROW}"].value)) == Decimal("0.35")  # вес/шт не менялся


def test_gross_total_not_stored_back(client, make_user, env):
    part, _ = _brp(env, material="219800345", qty=10)
    _login(client, make_user)
    _save(client, part, gross_weight_kg="0.35")
    _sell(env, part, number="219800345", qty="7")
    client.get(reverse("actions_export"))  # экспорт с количеством 7
    customs = PartCustomsInfo.objects.get(part_type=part)
    assert customs.gross_weight_kg == Decimal("0.35")  # не 0.35*7 и не что-то другое
    assert not hasattr(customs, "gross_weight_total_kg")  # такого поля в модели нет


# --- 10-15. Валидация: Decimal, точность, границы -------------------------------------------


def test_decimal_used_without_float_errors(client, make_user, env):
    part, _ = _brp(env, material="219800345")
    _login(client, make_user)
    _save(client, part, gross_weight_kg="10.10", net_weight_kg="0.10")
    customs = PartCustomsInfo.objects.get(part_type=part)
    assert customs.gross_weight_kg == Decimal("10.10")  # не 10.099999999999998
    assert customs.net_weight_kg == Decimal("0.10")


def test_tiny_weight_not_rounded_to_zero(client, make_user, env):
    part, _ = _brp(env, material="219800345")
    PartCustomsInfo.objects.create(part_type=part, gross_weight_kg=Decimal("0.001"))
    _sell(env, part, number="219800345")
    _login(client, make_user)
    sheet = _sheet(client.get(reverse("actions_export")).content)
    assert Decimal(str(sheet[f"G{DATA_ROW}"].value)) == Decimal("0.001")


def test_negative_weight_rejected():
    with pytest.raises(ValueError, match="больше нуля"):
        parse_weight_kg("-1")


def test_zero_weight_rejected():
    with pytest.raises(ValueError, match="больше нуля"):
        parse_weight_kg("0")


def test_gross_less_than_net_rejected(client, make_user, env):
    part, _ = _brp(env, material="219800345")
    _login(client, make_user)
    resp = _save(client, part, gross_weight_kg="0.1", net_weight_kg="0.5")
    assert resp.status_code == 302
    assert not PartCustomsInfo.objects.filter(part_type=part).exists()  # ничего не сохранено


def test_gross_less_than_net_rejected_against_existing_value(client, make_user, env):
    """Кросс-проверка учитывает уже сохранённое значение другого поля."""
    part, _ = _brp(env, material="219800345")
    PartCustomsInfo.objects.create(part_type=part, gross_weight_kg=Decimal("0.2"))
    _login(client, make_user)
    _save(client, part, net_weight_kg="0.5")  # 0.2 (старый gross) < 0.5 (новый net)
    customs = PartCustomsInfo.objects.get(part_type=part)
    assert customs.net_weight_kg is None  # net не сохранился


def test_empty_weight_stays_none_not_zero(client, make_user, env):
    part, _ = _brp(env, material="219800345")
    _login(client, make_user)
    _save(client, part, gross_weight_kg="")
    customs = PartCustomsInfo.objects.get(part_type=part)
    assert customs.gross_weight_kg is None  # не Decimal("0")


# --- 16-18. Статус готовности -----------------------------------------------------------


def test_only_gross_filled_is_not_ready(client, make_user, env):
    part, _ = _brp(env, material="219800345")
    PartCustomsInfo.objects.create(
        part_type=part, application_area=ApplicationArea.SNOWMOBILE,
        gross_weight_kg=Decimal("0.5"),
    )
    row = part_export_data(part)
    assert row["customs_ready"] is False
    assert "Не заполнен вес нетто" in row["customs_missing_reasons"]


def test_only_net_filled_is_not_ready(client, make_user, env):
    part, _ = _brp(env, material="219800345")
    PartCustomsInfo.objects.create(
        part_type=part, application_area=ApplicationArea.SNOWMOBILE,
        net_weight_kg=Decimal("0.4"),
    )
    row = part_export_data(part)
    assert row["customs_ready"] is False
    assert "Не заполнен вес брутто" in row["customs_missing_reasons"]


def test_both_weights_and_application_area_is_ready(client, make_user, env):
    part, _ = _brp(env, material="219800345")
    PartCustomsInfo.objects.create(
        part_type=part, application_area=ApplicationArea.SNOWMOBILE,
        gross_weight_kg=Decimal("0.5"), net_weight_kg=Decimal("0.4"),
    )
    row = part_export_data(part)
    assert row["customs_ready"] is True
    assert row["customs_missing_reasons"] == []


# --- 19-20. Частичное сохранение не стирает другое поле -----------------------------------


def test_changing_application_area_does_not_erase_weights(client, make_user, env):
    part, _ = _brp(env, material="219800345")
    PartCustomsInfo.objects.create(
        part_type=part, gross_weight_kg=Decimal("0.5"), net_weight_kg=Decimal("0.4"),
    )
    _login(client, make_user)
    _save(client, part, application_area=ApplicationArea.SNOWMOBILE)  # без ключей веса
    customs = PartCustomsInfo.objects.get(part_type=part)
    assert customs.gross_weight_kg == Decimal("0.5")
    assert customs.net_weight_kg == Decimal("0.4")
    assert customs.application_area == "СНЕГОХОД"


def test_changing_weights_does_not_erase_application_area(client, make_user, env):
    part, _ = _brp(env, material="219800345")
    PartCustomsInfo.objects.create(part_type=part, application_area=ApplicationArea.ATV)
    _login(client, make_user)
    _save(client, part, gross_weight_kg="0.5", net_weight_kg="0.4")  # без ключа области
    customs = PartCustomsInfo.objects.get(part_type=part)
    assert customs.application_area == "КВАДРОЦИКЛ"
    assert customs.gross_weight_kg == Decimal("0.5")
    assert customs.net_weight_kg == Decimal("0.4")


# --- 21-22. GET не создаёт PartCustomsInfo ------------------------------------------------


def test_report_get_does_not_create_customs_row_for_weights(client, make_user, env):
    part, _ = _brp(env, material="219800345")
    _sell(env, part, number="219800345")
    _login(client, make_user)
    client.get(reverse("actions_report"))
    assert not PartCustomsInfo.objects.filter(part_type=part).exists()


def test_export_get_does_not_create_customs_row_for_weights(client, make_user, env):
    part, _ = _brp(env, material="219800345")
    _sell(env, part, number="219800345")
    _login(client, make_user)
    client.get(reverse("actions_export"))
    assert not PartCustomsInfo.objects.filter(part_type=part).exists()


# --- 23-24. Права доступа ------------------------------------------------------------------


def test_unauthenticated_cannot_save_weights(client, env):
    part, _ = _brp(env, material="219800345")
    resp = _save(client, part, gross_weight_kg="0.5")
    assert resp.status_code == 302
    assert "login" in resp.url
    assert not PartCustomsInfo.objects.filter(part_type=part).exists()


def test_viewer_role_cannot_save_weights(client, make_user, env):
    part, _ = _brp(env, material="219800345")
    make_user("viewer", role=roles.VIEWER)
    client.login(username="viewer", password=PASSWORD)
    resp = _save(client, part, gross_weight_kg="0.5")
    assert resp.status_code == 403
    assert not PartCustomsInfo.objects.filter(part_type=part).exists()


# --- 25-26. Точная идентичность ------------------------------------------------------------


def test_brp_and_polaris_same_number_weights_not_mixed(client, make_user, env):
    brp_part, _ = _brp(env, material="3610075")
    pol_part, _ = _polaris(env, number="3610075")  # тот же номер, другой производитель
    _login(client, make_user)
    _save(client, brp_part, gross_weight_kg="0.5", net_weight_kg="0.4")
    _save(client, pol_part, gross_weight_kg="1.2", net_weight_kg="1.0")
    assert PartCustomsInfo.objects.get(part_type=brp_part).gross_weight_kg == Decimal("0.5")
    assert PartCustomsInfo.objects.get(part_type=pol_part).gross_weight_kg == Decimal("1.2")


def test_replacement_does_not_get_exact_part_weight(client, make_user, env):
    exact_part, _ = _brp(
        env, material="250000059", retail="0", wholesale="0", replacement="250000418"
    )
    replacement_part, _ = _brp(env, material="250000418", retail="4.19", wholesale="3.29")
    PartCustomsInfo.objects.create(
        part_type=replacement_part, gross_weight_kg=Decimal("2"), net_weight_kg=Decimal("1.8"),
    )
    row = part_export_data(exact_part)
    assert row["usd_price"] == Decimal("3.29")  # цена - от замены (уже покрыто ранее)
    assert row["gross_weight_kg"] is None  # вес замены не унаследован
    assert row["net_weight_kg"] is None


# --- 27-29. Excel: числовые ячейки, формула ------------------------------------------------


def test_xlsx_with_weights_opens_via_openpyxl(client, make_user, env):
    part, _ = _brp(env, material="219800345")
    PartCustomsInfo.objects.create(
        part_type=part, gross_weight_kg=Decimal("0.35"), net_weight_kg=Decimal("0.30"),
    )
    _sell(env, part, number="219800345")
    _login(client, make_user)
    resp = client.get(reverse("actions_export"))
    workbook = openpyxl.load_workbook(BytesIO(resp.content))  # без ошибок/предупреждений
    assert SHEET in workbook.sheetnames


def test_gross_and_net_cells_are_numeric(client, make_user, env):
    part, _ = _brp(env, material="219800345")
    PartCustomsInfo.objects.create(
        part_type=part, gross_weight_kg=Decimal("0.35"), net_weight_kg=Decimal("0.30"),
    )
    _sell(env, part, number="219800345")
    _login(client, make_user)
    sheet = _sheet(client.get(reverse("actions_export")).content)
    assert isinstance(sheet[f"G{DATA_ROW}"].value, (int, float, Decimal))
    assert isinstance(sheet[f"H{DATA_ROW}"].value, (int, float, Decimal))
    assert "кг" not in str(sheet[f"G{DATA_ROW}"].value)


def test_formula_references_correct_row_and_quantity_column(client, make_user, env):
    part, _ = _brp(env, material="219800345")
    PartCustomsInfo.objects.create(part_type=part, gross_weight_kg=Decimal("0.35"))
    _sell(env, part, number="219800345", qty="4")
    _login(client, make_user)
    sheet = _sheet(client.get(reverse("actions_export")).content)
    assert sheet[f"I{DATA_ROW}"].value == f"=J{DATA_ROW}*G{DATA_ROW}"


# --- weight_verified: ручной ввод обоих весов = ручное подтверждение ------------------------


def test_quick_save_both_weights_sets_verified(client, make_user, env):
    part, _ = _brp(env, material="219800345")
    _login(client, make_user)
    _save(client, part, gross_weight_kg="0.5", net_weight_kg="0.4")
    customs = PartCustomsInfo.objects.get(part_type=part)
    assert customs.weight_verified is True
    assert customs.updated_by is not None


def test_quick_save_only_gross_does_not_verify(client, make_user, env):
    part, _ = _brp(env, material="219800345")
    _login(client, make_user)
    _save(client, part, gross_weight_kg="0.5")  # net не заполнен
    customs = PartCustomsInfo.objects.get(part_type=part)
    assert customs.gross_weight_kg == Decimal("0.5")  # частичное сохранение разрешено
    assert customs.weight_verified is False


def test_quick_save_only_net_does_not_verify(client, make_user, env):
    part, _ = _brp(env, material="219800345")
    _login(client, make_user)
    _save(client, part, net_weight_kg="0.4")
    customs = PartCustomsInfo.objects.get(part_type=part)
    assert customs.net_weight_kg == Decimal("0.4")
    assert customs.weight_verified is False


def test_clearing_one_weight_resets_verified(client, make_user, env):
    part, _ = _brp(env, material="219800345")
    PartCustomsInfo.objects.create(
        part_type=part, application_area=ApplicationArea.SNOWMOBILE,
        gross_weight_kg=Decimal("0.5"), net_weight_kg=Decimal("0.4"),
        weight_verified=True,
    )
    _login(client, make_user)
    _save(client, part, gross_weight_kg="0.5", net_weight_kg="")  # нетто удалили
    customs = PartCustomsInfo.objects.get(part_type=part)
    assert customs.net_weight_kg is None
    assert customs.weight_verified is False
    row = part_export_data(part)
    assert row["customs_ready"] is False  # позиция снова неготова
    assert "Не заполнен вес нетто" in row["customs_missing_reasons"]


def test_changing_only_application_area_keeps_verified(client, make_user, env):
    part, _ = _brp(env, material="219800345")
    PartCustomsInfo.objects.create(
        part_type=part, gross_weight_kg=Decimal("0.5"), net_weight_kg=Decimal("0.4"),
        weight_verified=True,
    )
    _login(client, make_user)
    _save(client, part, application_area=ApplicationArea.ATV)  # ключей веса нет в POST
    customs = PartCustomsInfo.objects.get(part_type=part)
    assert customs.weight_verified is True  # не сброшен
    assert customs.gross_weight_kg == Decimal("0.5")  # веса не стёрты
    assert customs.net_weight_kg == Decimal("0.4")
    assert customs.application_area == "КВАДРОЦИКЛ"


def test_invalid_weights_do_not_change_verified(client, make_user, env):
    part, _ = _brp(env, material="219800345")
    PartCustomsInfo.objects.create(
        part_type=part, gross_weight_kg=Decimal("0.5"), net_weight_kg=Decimal("0.4"),
        weight_verified=True,
    )
    _login(client, make_user)
    _save(client, part, gross_weight_kg="0.1")  # 0.1 < сохранённого нетто 0.4 -> откат
    customs = PartCustomsInfo.objects.get(part_type=part)
    assert customs.gross_weight_kg == Decimal("0.5")  # ничего не изменилось
    assert customs.weight_verified is True
    _save(client, part, gross_weight_kg="абв")  # не число -> ранний отказ до транзакции
    customs.refresh_from_db()
    assert customs.gross_weight_kg == Decimal("0.5")
    assert customs.weight_verified is True


def test_invalid_weights_on_new_card_do_not_verify(client, make_user, env):
    part, _ = _brp(env, material="219800345")
    _login(client, make_user)
    _save(client, part, gross_weight_kg="0.1", net_weight_kg="0.5")  # gross < net
    assert not PartCustomsInfo.objects.filter(part_type=part).exists()


def test_get_does_not_change_verified(client, make_user, env):
    part, _ = _brp(env, material="219800345")
    PartCustomsInfo.objects.create(
        part_type=part, gross_weight_kg=Decimal("0.5"), net_weight_kg=Decimal("0.4"),
        weight_verified=False,
    )
    _sell(env, part, number="219800345")
    _login(client, make_user)
    client.get(reverse("actions_report"))
    client.get(reverse("actions_export"))
    resp = client.get(reverse("actions_customs_quick_save", args=[part.pk]))
    assert resp.status_code == 405  # квик-сохранение — только POST
    customs = PartCustomsInfo.objects.get(part_type=part)
    assert customs.weight_verified is False  # GET ничего не подтвердил


def test_unauthenticated_cannot_verify_weights(client, env):
    part, _ = _brp(env, material="219800345")
    PartCustomsInfo.objects.create(
        part_type=part, gross_weight_kg=Decimal("0.5"), net_weight_kg=Decimal("0.4"),
        weight_verified=False,
    )
    resp = _save(client, part, gross_weight_kg="0.5", net_weight_kg="0.4")
    assert resp.status_code == 302 and "login" in resp.url
    assert PartCustomsInfo.objects.get(part_type=part).weight_verified is False


def test_viewer_cannot_verify_weights(client, make_user, env):
    part, _ = _brp(env, material="219800345")
    PartCustomsInfo.objects.create(
        part_type=part, gross_weight_kg=Decimal("0.5"), net_weight_kg=Decimal("0.4"),
        weight_verified=False,
    )
    make_user("viewer", role=roles.VIEWER)
    client.login(username="viewer", password=PASSWORD)
    resp = _save(client, part, gross_weight_kg="0.5", net_weight_kg="0.4")
    assert resp.status_code == 403
    assert PartCustomsInfo.objects.get(part_type=part).weight_verified is False


# --- Маркер ручного источника: заметка без фиктивного URL ------------------------------------


def test_quick_save_writes_manual_note_without_url(client, make_user, env):
    part, _ = _brp(env, material="219800345")
    _login(client, make_user)
    _save(client, part, gross_weight_kg="0.5", net_weight_kg="0.4")
    customs = PartCustomsInfo.objects.get(part_type=part)
    assert customs.weight_source_note == MANUAL_WEIGHT_NOTE
    assert customs.weight_source_url == ""  # фиктивная ссылка не выдумана
    assert part_export_data(part)["weight_source"] == "manual"  # «Указано вручную»
    html = client.get(reverse("actions_customs_edit", args=[part.pk])).content.decode()
    assert "Указано вручную" in html


def test_quick_save_preserves_real_source_note(client, make_user, env):
    part, _ = _brp(env, material="219800345")
    PartCustomsInfo.objects.create(
        part_type=part, weight_source_url="https://example.com/spec",
        weight_source_note="страница поставщика",
    )
    _login(client, make_user)
    _save(client, part, gross_weight_kg="0.5", net_weight_kg="0.4")
    customs = PartCustomsInfo.objects.get(part_type=part)
    assert customs.weight_verified is True
    assert customs.weight_source_url == "https://example.com/spec"  # источник не затёрт
    assert customs.weight_source_note == "страница поставщика"
    assert part_export_data(part)["weight_source"] == "sourced"


def test_quick_save_partial_pair_does_not_write_manual_note(client, make_user, env):
    part, _ = _brp(env, material="219800345")
    _login(client, make_user)
    _save(client, part, gross_weight_kg="0.5")  # пары нет — подтверждать нечего
    customs = PartCustomsInfo.objects.get(part_type=part)
    assert customs.weight_source_note == ""


# --- Согласованность с детальной формой -------------------------------------------------------


def test_detail_form_cannot_verify_incomplete_pair(client, make_user, env):
    part, _ = _brp(env, material="219800345")
    _login(client, make_user)
    client.post(
        reverse("actions_customs_edit", args=[part.pk]),
        {"gross_weight_kg": "0.5", "weight_verified": "on"},  # нетто нет
    )
    customs = PartCustomsInfo.objects.get(part_type=part)
    assert customs.gross_weight_kg == Decimal("0.5")
    assert customs.weight_verified is False  # неполную пару подтвердить нельзя


def test_detail_form_explicit_unchecked_respected(client, make_user, env):
    """Детальная форма — явный чекбокс: оба веса без отметки НЕ подтверждают.

    Там есть поля источника (вес мог быть записан с непроверенной страницы),
    поэтому решение остаётся за пользователем. Быстрый редактор без полей
    источника — чисто ручной ввод, он подтверждает автоматически.
    """
    part, _ = _brp(env, material="219800345")
    _login(client, make_user)
    client.post(
        reverse("actions_customs_edit", args=[part.pk]),
        {"gross_weight_kg": "0.5", "net_weight_kg": "0.4"},  # чекбокс не отмечен
    )
    customs = PartCustomsInfo.objects.get(part_type=part)
    assert customs.weight_verified is False


def test_report_shows_verified_pill(client, make_user, env):
    part, _ = _brp(env, material="219800345")
    _sell(env, part, number="219800345")
    _login(client, make_user)
    html = client.get(reverse("actions_report")).content.decode()
    assert "Вес подтверждён вручную" not in html  # ещё не подтверждён
    _save(client, part, gross_weight_kg="0.5", net_weight_kg="0.4")
    html = client.get(reverse("actions_report")).content.decode()
    assert "Вес подтверждён вручную" in html
